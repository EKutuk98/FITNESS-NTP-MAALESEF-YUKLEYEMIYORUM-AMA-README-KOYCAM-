#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FITNESS TAKİP SİSTEMİ v1.0 — TIER 1
Login + Dashboard + Sporcular CRUD + Antrenmanlar CRUD
═══════════════════════════════════════════════════════
Teknik:
  - PyQt5 + SQLite3 (@contextmanager)
  - QPainter custom charts (matplotlib YASAK)
  - Dark Luxury Tema (#0f0f1a / #1a1a2e / #6366f1)
  - Soft Delete (durum='Pasif')
  - Auto-migration
  - Live Search
  - Örnek veri (datetime.now() bazlı)
"""

import sys
import os
import sqlite3
import hashlib
import random
from datetime import datetime, timedelta
from contextlib import contextmanager

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel,
    QLineEdit, QComboBox, QMessageBox, QFrame, QSpinBox,
    QHeaderView, QGridLayout, QTextEdit, QStackedWidget,
    QDoubleSpinBox, QSizePolicy, QDateEdit
)
from PyQt5.QtCore import Qt, QTimer, QDate, QRectF, QPointF
from PyQt5.QtGui import (
    QFont, QColor, QPainter, QPen, QBrush,
    QLinearGradient, QPainterPath
)


# ═══════════════════════════════════════════════════════════════════
# RENKLER & TEMA
# ═══════════════════════════════════════════════════════════════════
COLORS = {
    'bg_main':        '#0f0f1a',
    'bg_secondary':   '#1a1a2e',
    'bg_tertiary':    '#1e1e35',
    'bg_card':        '#16213e',
    'bg_hover':       '#1e1e35',
    'primary':        '#6366f1',
    'primary_light':  '#818cf8',
    'primary_dark':   '#4f46e5',
    'accent':         '#f5a623',
    'success':        '#10b981',
    'success_dark':   '#059669',
    'warning':        '#f59e0b',
    'warning_dark':   '#d97706',
    'danger':         '#ef4444',
    'danger_dark':    '#dc2626',
    'info':           '#3b82f6',
    'info_dark':      '#2563eb',
    'text_main':      '#ffffff',
    'text_secondary': '#9ca3af',
    'text_muted':     '#6b7280',
    'border':         '#2d2d44',
    'border_light':   '#3d3d5c',
    'sidebar_bg':     '#0a0a14',
    'sidebar_active': '#6366f1',
    'table_row_alt':  '#12122a',
    'gold':           '#f5a623',
}

C = COLORS  # alias

INPUT_H  = 44
BTN_H    = 44
ROW_H    = 40
SPACING  = 14
PADDING  = 22
SIDEBAR_W = 220

GLOBAL_STYLESHEET = f"""
QMainWindow {{
    background-color: {C['bg_main']};
}}
QWidget {{
    background-color: transparent;
    color: {C['text_main']};
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}}
QPushButton {{
    background-color: {C['primary']};
    color: white; border: none;
    padding: 10px 22px; border-radius: 10px;
    font-weight: bold; font-size: 13px;
    min-height: {BTN_H}px;
}}
QPushButton:hover {{ background-color: {C['primary_light']}; }}
QPushButton:pressed {{ background-color: {C['primary_dark']}; }}
QPushButton:disabled {{ background-color: {C['border']}; color: {C['text_muted']}; }}
QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {{
    background-color: {C['bg_secondary']};
    border: 1.5px solid {C['border']};
    border-radius: 8px; padding: 8px 14px;
    color: {C['text_main']}; font-size: 13px;
    min-height: {INPUT_H}px;
    selection-background-color: {C['primary']};
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus,
QSpinBox:focus, QDoubleSpinBox:focus, QDateEdit:focus {{
    border: 2px solid {C['primary']};
    background-color: #1c1c35;
}}
QComboBox::drop-down {{ border: none; width: 32px; }}
QComboBox QAbstractItemView {{
    background-color: {C['bg_secondary']};
    color: {C['text_main']};
    selection-background-color: {C['primary']};
    border: 1px solid {C['border']};
    font-size: 13px;
}}
QTableWidget {{
    background-color: {C['bg_tertiary']};
    alternate-background-color: {C['table_row_alt']};
    gridline-color: {C['border']};
    border: 1px solid {C['border']};
    border-radius: 10px; font-size: 13px;
    selection-background-color: {C['primary_dark']};
    selection-color: white;
}}
QTableWidget::item {{ padding: 8px 10px; color: {C['text_main']}; }}
QTableWidget::item:selected {{ background-color: {C['primary_dark']}; color: white; }}
QHeaderView::section {{
    background-color: {C['bg_secondary']};
    color: {C['primary_light']}; padding: 10px 10px;
    border: none; border-right: 1px solid {C['border']};
    border-bottom: 2px solid {C['primary']};
    font-weight: bold; font-size: 12px;
}}
QTabWidget::pane {{
    border: 1px solid {C['border']};
    border-radius: 8px; background-color: {C['bg_main']};
}}
QTabBar::tab {{
    background-color: {C['bg_secondary']}; color: {C['text_secondary']};
    padding: 10px 24px; border-top-left-radius: 8px;
    border-top-right-radius: 8px; margin-right: 3px;
    font-weight: bold; font-size: 13px; min-width: 100px;
}}
QTabBar::tab:selected {{ background-color: {C['primary']}; color: white; }}
QTabBar::tab:hover:!selected {{ background-color: {C['bg_hover']}; color: {C['text_main']}; }}
QScrollBar:vertical {{
    background: {C['bg_main']}; width: 8px; border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {C['border_light']}; border-radius: 4px; min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {C['primary']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: {C['bg_main']}; height: 8px; border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {C['border_light']}; border-radius: 4px;
}}
QScrollBar::handle:horizontal:hover {{ background: {C['primary']}; }}
QDialog {{ background-color: {C['bg_main']}; }}
QLabel {{ color: {C['text_main']}; font-size: 13px; }}
QMessageBox {{ background-color: {C['bg_main']}; }}
QMessageBox QLabel {{ color: {C['text_main']}; font-size: 13px; }}
"""


# ═══════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════
def hash_password(pw):
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

def dark_msg(parent, title, text, icon=QMessageBox.Information):
    msg = QMessageBox(parent)
    msg.setWindowTitle(title); msg.setText(text); msg.setIcon(icon)
    msg.setStyleSheet(f"""
        QMessageBox {{ background-color: {C['bg_main']}; }}
        QMessageBox QLabel {{ color: {C['text_main']}; font-size: 13px; min-width: 280px; }}
        QPushButton {{
            background-color: {C['primary']}; color: white; border: none;
            padding: 10px 28px; border-radius: 8px; font-weight: bold;
            font-size: 13px; min-width: 90px;
        }}
        QPushButton:hover {{ background-color: {C['primary_light']}; }}
    """)
    return msg.exec_()

def dark_confirm(parent, title, text):
    msg = QMessageBox(parent)
    msg.setWindowTitle(title); msg.setText(text)
    msg.setIcon(QMessageBox.Question)
    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msg.setDefaultButton(QMessageBox.No)
    msg.setStyleSheet(f"""
        QMessageBox {{ background-color: {C['bg_main']}; }}
        QMessageBox QLabel {{ color: {C['text_main']}; font-size: 13px; min-width: 280px; }}
        QPushButton {{
            background-color: {C['primary']}; color: white; border: none;
            padding: 10px 28px; border-radius: 8px; font-weight: bold;
            font-size: 13px; min-width: 90px;
        }}
        QPushButton:hover {{ background-color: {C['primary_light']}; }}
    """)
    return msg.exec_() == QMessageBox.Yes


# ═══════════════════════════════════════════════════════════════════
# QPainter WIDGET'LAR
# ═══════════════════════════════════════════════════════════════════
class KPICard(QWidget):
    """Gradient KPI kartı — QPainter ile çizilir."""
    def __init__(self, title, value, icon, color_start, color_end, parent=None):
        super().__init__(parent)
        self.title_text = title
        self.value_text = str(value)
        self.icon_text  = icon
        self.color_start = color_start
        self.color_end   = color_end
        self.setFixedHeight(100)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def set_value(self, v):
        self.value_text = str(v); self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        grad = QLinearGradient(0, 0, w, h)
        grad.setColorAt(0, QColor(self.color_start))
        grad.setColorAt(1, QColor(self.color_end))
        path = QPainterPath()
        path.addRoundedRect(QRectF(0, 0, w, h), 14, 14)
        p.fillPath(path, QBrush(grad))
        # Dekoratif daire
        p.setBrush(QBrush(QColor(255, 255, 255, 18)))
        p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(w - 25, 25), 50, 50)
        # Icon — sağ üst
        p.setPen(QPen(QColor(255, 255, 255, 210)))
        p.setFont(QFont('Segoe UI', 22))
        p.drawText(QRectF(w - 55, 8, 48, 36), Qt.AlignCenter, self.icon_text)
        # Başlık — sol üst, tam genişlik, küçük font ki isim sığsın
        p.setPen(QPen(QColor(255, 255, 255, 175)))
        p.setFont(QFont('Segoe UI', 10))
        p.drawText(QRectF(14, 8, w - 20, 24), Qt.AlignLeft | Qt.AlignVCenter, self.title_text)
        # Ayraç
        p.setPen(QPen(QColor(255, 255, 255, 40), 1))
        p.drawLine(14, 38, w - 14, 38)
        # Değer — büyük
        p.setPen(QPen(QColor(255, 255, 255)))
        p.setFont(QFont('Segoe UI', 22, QFont.Bold))
        p.drawText(QRectF(14, 42, w - 60, 54), Qt.AlignLeft | Qt.AlignVCenter, self.value_text)
        p.end()


class BarChartWidget(QWidget):
    """QPainter bar chart."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = []; self.title = ''
        self.setMinimumHeight(220)

    def set_data(self, data, title=''):
        self.data = data; self.title = title; self.update()

    def paintEvent(self, event):
        if not self.data: return
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))
        pad_l, pad_r, pad_t, pad_b = 50, 20, 40, 50
        chart_w = w - pad_l - pad_r
        chart_h = h - pad_t - pad_b
        max_val = max((v for _, v in self.data), default=1) or 1
        # Başlık
        p.setPen(QColor(C['text_main']))
        p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 30), Qt.AlignCenter, self.title)
        n = len(self.data)
        gap = chart_w / n
        bar_w = max(8, int(gap * 0.55))
        for i, (label, val) in enumerate(self.data):
            x = pad_l + i * gap + (gap - bar_w) / 2
            bar_h = int(chart_h * val / max_val) if max_val else 0
            y = pad_t + chart_h - bar_h
            grad = QLinearGradient(x, y, x, y + bar_h)
            grad.setColorAt(0, QColor(C['primary']))
            grad.setColorAt(1, QColor(C['primary_dark']))
            p.setPen(Qt.NoPen); p.setBrush(QBrush(grad))
            p.drawRoundedRect(int(x), y, bar_w, bar_h, 4, 4)
            if bar_h > 18:
                p.setPen(QColor(255, 255, 255))
                p.setFont(QFont('Segoe UI', 10, QFont.Bold))
                p.drawText(int(x), y + 2, bar_w, 16, Qt.AlignCenter, str(int(val)))
            p.setPen(QColor(C['text_secondary']))
            p.setFont(QFont('Segoe UI', 10))
            p.drawText(int(x) - 10, h - pad_b + 6, bar_w + 20, 20, Qt.AlignCenter, str(label))
        p.setPen(QPen(QColor(C['border']), 1))
        p.drawLine(pad_l, pad_t + chart_h, w - pad_r, pad_t + chart_h)
        p.end()


class PieChartWidget(QWidget):
    """QPainter pie chart."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = []; self.title = ''
        self.colors = ['#6366f1','#10b981','#f59e0b','#ef4444','#ec4899','#14b8a6','#8b5cf6']
        self.setMinimumHeight(220)

    def set_data(self, data, title=''):
        self.data = data; self.title = title; self.update()

    def paintEvent(self, event):
        if not self.data: return
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))
        total = sum(v for _, v in self.data)
        if not total: return
        p.setPen(QColor(C['text_main']))
        p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 28), Qt.AlignCenter, self.title)
        pie_size = min(w // 2, h - 60) - 10
        cx = w // 4; cy = (h + 34) // 2
        start = 90 * 16
        for i, (label, val) in enumerate(self.data):
            span = int(val / total * 5760)
            p.setPen(Qt.NoPen); p.setBrush(QColor(self.colors[i % len(self.colors)]))
            p.drawPie(QRectF(cx - pie_size//2, cy - pie_size//2, pie_size, pie_size), start, span)
            start += span
        lx = w // 2 + 10; ly = 38
        for i, (label, val) in enumerate(self.data):
            pct = int(val / total * 100)
            p.setPen(Qt.NoPen); p.setBrush(QColor(self.colors[i % len(self.colors)]))
            p.drawRoundedRect(lx, ly + i * 26, 14, 14, 3, 3)
            p.setPen(QColor(C['text_secondary']))
            p.setFont(QFont('Segoe UI', 11))
            p.drawText(lx + 20, ly + i * 26, w - lx - 30, 18, Qt.AlignVCenter, f"{label} ({pct}%)")
        p.end()


def make_btn(text, color=None, small=False):
    btn = QPushButton(text)
    h = 36 if small else BTN_H
    btn.setFixedHeight(h)
    btn.setFont(QFont('Segoe UI', 12 if small else 13, QFont.Bold))
    c = color or C['primary']
    qc = QColor(c)
    btn.setStyleSheet(f"""
        QPushButton {{
            background-color: {c}; color: white; border: none;
            border-radius: 8px; padding: 0 18px;
            font-size: {'12' if small else '13'}px; font-weight: bold;
        }}
        QPushButton:hover {{ background-color: {qc.lighter(115).name()}; }}
        QPushButton:pressed {{ background-color: {qc.darker(115).name()}; }}
        QPushButton:disabled {{ background-color: {C['border']}; color: {C['text_muted']}; }}
    """)
    return btn


def make_table(headers, col_widths=None):
    t = QTableWidget()
    t.setColumnCount(len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setSelectionBehavior(QTableWidget.SelectRows)
    t.setSelectionMode(QTableWidget.SingleSelection)
    t.setAlternatingRowColors(True)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(ROW_H)
    t.setShowGrid(True)
    hdr = t.horizontalHeader()
    if col_widths:
        for i, cw in enumerate(col_widths):
            if cw == -1: hdr.setSectionResizeMode(i, QHeaderView.Stretch)
            else: hdr.setSectionResizeMode(i, QHeaderView.Interactive); t.setColumnWidth(i, cw)
    else:
        hdr.setSectionResizeMode(QHeaderView.Stretch)
    return t


def make_search(placeholder='Ara...'):
    inp = QLineEdit()
    inp.setPlaceholderText(f'🔍  {placeholder}')
    inp.setFixedHeight(INPUT_H)
    inp.setFont(QFont('Segoe UI', 13))
    return inp


def make_combo(items, width=160):
    cb = QComboBox()
    cb.addItems(items)
    cb.setFixedHeight(INPUT_H)
    if width: cb.setFixedWidth(width)
    cb.setFont(QFont('Segoe UI', 13))
    return cb


# ═══════════════════════════════════════════════════════════════════
# VERİTABANI YÖNETİCİSİ
# ═══════════════════════════════════════════════════════════════════
class DatabaseManager:
    DB_NAME = 'fitness_takip.db'

    def __init__(self, path=None):
        self.db_path = path or self.DB_NAME
        self.create_tables()
        self._maybe_load_sample_data()

    @contextmanager
    def get_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn; conn.commit()
        except Exception as e:
            conn.rollback(); raise e
        finally:
            conn.close()

    def create_tables(self):
        with self.get_connection() as conn:
            conn.executescript("""
            CREATE TABLE IF NOT EXISTS kullanicilar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kullanici_adi TEXT UNIQUE NOT NULL,
                sifre TEXT NOT NULL,
                ad TEXT NOT NULL,
                rol TEXT DEFAULT 'personel',
                durum TEXT DEFAULT 'Aktif',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS sporcular (
                sporcu_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                soyad TEXT NOT NULL,
                yas INTEGER,
                kilo REAL NOT NULL,
                boy REAL NOT NULL,
                cinsiyet TEXT DEFAULT 'Erkek',
                hedef TEXT DEFAULT 'Form Koruma',
                bmi REAL,
                notlar TEXT DEFAULT '',
                durum TEXT DEFAULT 'Aktif',
                kayit_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS antrenmanlar (
                antrenman_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                kategori TEXT NOT NULL,
                saat INTEGER DEFAULT 0,
                dakika INTEGER DEFAULT 30,
                zorluk_seviyesi TEXT DEFAULT 'Orta',
                aciklama TEXT DEFAULT '',
                kalori_tahmini INTEGER DEFAULT 0,
                durum TEXT DEFAULT 'Aktif',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS takipler (
                takip_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                antrenman_id INTEGER NOT NULL,
                kalori INTEGER DEFAULT 0,
                nabiz INTEGER,
                sure_dakika INTEGER DEFAULT 0,
                notlar TEXT DEFAULT '',
                tarih TEXT DEFAULT (datetime('now','localtime')),
                durum TEXT DEFAULT 'Aktif',
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id),
                FOREIGN KEY (antrenman_id) REFERENCES antrenmanlar(antrenman_id)
            );
            CREATE TABLE IF NOT EXISTS ilerlemeler (
                ilerleme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                kilo REAL NOT NULL,
                notlar TEXT DEFAULT '',
                tarih TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS antrenman_programlari (
                program_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                program_adi TEXT NOT NULL,
                hedef TEXT DEFAULT 'Genel',
                sure_hafta INTEGER DEFAULT 4,
                aktif INTEGER DEFAULT 1,
                notlar TEXT DEFAULT '',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS program_gunleri (
                gun_id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_id INTEGER NOT NULL,
                gun_adi TEXT NOT NULL,
                antrenman_id INTEGER NOT NULL,
                set_sayisi INTEGER DEFAULT 3,
                tekrar INTEGER DEFAULT 12,
                agirlik REAL DEFAULT 0,
                dinlenme_saniye INTEGER DEFAULT 60,
                notlar TEXT DEFAULT '',
                FOREIGN KEY (program_id) REFERENCES antrenman_programlari(program_id),
                FOREIGN KEY (antrenman_id) REFERENCES antrenmanlar(antrenman_id)
            );
            CREATE TABLE IF NOT EXISTS gunluk_log (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                antrenman_id INTEGER NOT NULL,
                tarih TEXT DEFAULT (date('now','localtime')),
                sure_dakika INTEGER DEFAULT 0,
                kalori_yakilan INTEGER DEFAULT 0,
                max_nabiz INTEGER DEFAULT 0,
                ort_nabiz INTEGER DEFAULT 0,
                zorluk_hissi INTEGER DEFAULT 5,
                yorgunluk INTEGER DEFAULT 5,
                notlar TEXT DEFAULT '',
                durum TEXT DEFAULT 'Tamamlandi',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id),
                FOREIGN KEY (antrenman_id) REFERENCES antrenmanlar(antrenman_id)
            );
            CREATE TABLE IF NOT EXISTS set_kayitlari (
                set_id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_id INTEGER NOT NULL,
                set_no INTEGER NOT NULL,
                tekrar INTEGER DEFAULT 0,
                agirlik REAL DEFAULT 0,
                tamamlandi INTEGER DEFAULT 1,
                notlar TEXT DEFAULT '',
                FOREIGN KEY (log_id) REFERENCES gunluk_log(log_id)
            );
            CREATE TABLE IF NOT EXISTS vucut_olcumleri (
                olcum_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                tarih TEXT DEFAULT (date('now','localtime')),
                kilo REAL,
                bel REAL,
                gogus REAL,
                kalca REAL,
                sag_kol REAL,
                sol_kol REAL,
                sag_bacak REAL,
                sol_bacak REAL,
                vucut_yag_orani REAL,
                notlar TEXT DEFAULT '',
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS hedefler (
                hedef_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                hedef_tipi TEXT NOT NULL,
                hedef_adi TEXT NOT NULL,
                baslangic_degeri REAL DEFAULT 0,
                hedef_degeri REAL NOT NULL,
                mevcut_deger REAL DEFAULT 0,
                baslangic_tarihi TEXT DEFAULT (date('now','localtime')),
                hedef_tarihi TEXT,
                durum TEXT DEFAULT 'Devam',
                notlar TEXT DEFAULT '',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS streak_kayitlari (
                streak_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                tarih TEXT NOT NULL,
                antrenman_yapildi INTEGER DEFAULT 1,
                UNIQUE(sporcu_id, tarih),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS rozetler (
                rozet_id INTEGER PRIMARY KEY AUTOINCREMENT,
                rozet_adi TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                icon TEXT DEFAULT '🏅',
                kategori TEXT DEFAULT 'Genel',
                puan INTEGER DEFAULT 10,
                kosul_tipi TEXT NOT NULL,
                kosul_deger REAL NOT NULL
            );
            CREATE TABLE IF NOT EXISTS sporcu_rozetleri (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                rozet_id INTEGER NOT NULL,
                kazanma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(sporcu_id, rozet_id),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id),
                FOREIGN KEY (rozet_id) REFERENCES rozetler(rozet_id)
            );
            CREATE TABLE IF NOT EXISTS sporcu_puanlari (
                puan_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL UNIQUE,
                toplam_puan INTEGER DEFAULT 0,
                seviye TEXT DEFAULT 'Başlangıç',
                seviye_no INTEGER DEFAULT 1,
                guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS bir_rm_kayitlari (
                rm_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                egzersiz_adi TEXT NOT NULL,
                agirlik REAL NOT NULL,
                tekrar INTEGER NOT NULL,
                hesaplanan_1rm REAL NOT NULL,
                formul TEXT DEFAULT 'Epley',
                tarih TEXT DEFAULT (date('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS radar_degerleri (
                radar_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                guc_puani REAL DEFAULT 0,
                kardiyo_puani REAL DEFAULT 0,
                esneklik_puani REAL DEFAULT 0,
                denge_puani REAL DEFAULT 0,
                dayaniklilik_puani REAL DEFAULT 0,
                hiz_puani REAL DEFAULT 0,
                hesaplama_tarihi TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(sporcu_id),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS analiz_notlari (
                not_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                not_tipi TEXT DEFAULT 'Gozlem',
                baslik TEXT NOT NULL,
                icerik TEXT DEFAULT '',
                oncelik TEXT DEFAULT 'Normal',
                tarih TEXT DEFAULT (date('now','localtime')),
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS antrenman_sablonlari (
                sablon_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sablon_adi TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                kategori TEXT DEFAULT 'Genel',
                sure_hafta INTEGER DEFAULT 4,
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS sablon_gunleri (
                gun_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sablon_id INTEGER NOT NULL,
                gun_sira INTEGER NOT NULL,
                gun_adi TEXT NOT NULL,
                antrenman_id INTEGER,
                set_sayisi INTEGER DEFAULT 3,
                tekrar INTEGER DEFAULT 12,
                agirlik REAL DEFAULT 0,
                dinlenme_saniye INTEGER DEFAULT 60,
                notlar TEXT DEFAULT '',
                FOREIGN KEY (sablon_id) REFERENCES antrenman_sablonlari(sablon_id),
                FOREIGN KEY (antrenman_id) REFERENCES antrenmanlar(antrenman_id)
            );
            CREATE TABLE IF NOT EXISTS sistem_ayarlari (
                anahtar TEXT PRIMARY KEY,
                deger TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS hatirlaticilar (
                hatirlatici_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER,
                baslik TEXT NOT NULL,
                mesaj TEXT DEFAULT '',
                gun_tipi TEXT DEFAULT 'Gunluk',
                saat TEXT DEFAULT '09:00',
                aktif INTEGER DEFAULT 1,
                son_gosterim TEXT DEFAULT '',
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS aktivite_feed (
                feed_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                aktivite_tipi TEXT NOT NULL,
                baslik TEXT NOT NULL,
                detay TEXT DEFAULT '',
                ikon TEXT DEFAULT '📌',
                tarih TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            CREATE TABLE IF NOT EXISTS besinler (
                besin_id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL,
                kategori TEXT DEFAULT 'Genel',
                porsiyon_gram REAL DEFAULT 100,
                kalori REAL DEFAULT 0,
                protein REAL DEFAULT 0,
                karbonhidrat REAL DEFAULT 0,
                yag REAL DEFAULT 0,
                lif REAL DEFAULT 0,
                durum TEXT DEFAULT 'Aktif'
            );
            CREATE TABLE IF NOT EXISTS gunluk_beslenme (
                beslenme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL,
                besin_id INTEGER NOT NULL,
                tarih TEXT DEFAULT (date('now','localtime')),
                ogun TEXT DEFAULT 'Kahvalti',
                miktar_gram REAL DEFAULT 100,
                kalori_toplam REAL DEFAULT 0,
                protein_toplam REAL DEFAULT 0,
                karb_toplam REAL DEFAULT 0,
                yag_toplam REAL DEFAULT 0,
                notlar TEXT DEFAULT '',
                olusturma_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id),
                FOREIGN KEY (besin_id) REFERENCES besinler(besin_id)
            );
            CREATE TABLE IF NOT EXISTS kalori_hedefleri (
                kalori_hedef_id INTEGER PRIMARY KEY AUTOINCREMENT,
                sporcu_id INTEGER NOT NULL UNIQUE,
                gunluk_kalori INTEGER DEFAULT 2000,
                protein_gram INTEGER DEFAULT 150,
                karb_gram INTEGER DEFAULT 200,
                yag_gram INTEGER DEFAULT 65,
                guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (sporcu_id) REFERENCES sporcular(sporcu_id)
            );
            """)

    def _maybe_load_sample_data(self):
        with self.get_connection() as conn:
            if conn.execute("SELECT COUNT(*) FROM kullanicilar").fetchone()[0] > 0:
                return
            self._insert_sample_data(conn)

    def _insert_sample_data(self, conn):
        c = conn.cursor()
        # Kullanıcılar
        c.execute("INSERT INTO kullanicilar (kullanici_adi,sifre,ad,rol) VALUES (?,?,?,?)",
                  ('admin', hash_password('admin123'), 'Sistem Admin', 'admin'))
        c.execute("INSERT INTO kullanicilar (kullanici_adi,sifre,ad,rol) VALUES (?,?,?,?)",
                  ('personel', hash_password('personel123'), 'Antrenör', 'personel'))

        # Sporcular
        sporcular = [
            ('Ahmet',   'Yılmaz', 28, 75.5, 178, 'Erkek',  'Kilo Verme'),
            ('Ayşe',    'Demir',  25, 62.0, 165, 'Kadın',  'Kas Kazanma'),
            ('Mehmet',  'Kaya',   32, 85.0, 182, 'Erkek',  'Form Koruma'),
            ('Zeynep',  'Çelik',  23, 55.5, 162, 'Kadın',  'Kilo Verme'),
            ('Can',     'Yıldız', 29, 70.0, 175, 'Erkek',  'Dayanıklılık'),
            ('Selin',   'Arslan', 27, 58.5, 168, 'Kadın',  'Kas Kazanma'),
            ('Burak',   'Öztürk', 35, 92.0, 185, 'Erkek',  'Kilo Verme'),
            ('Elif',    'Şahin',  22, 52.0, 160, 'Kadın',  'Form Koruma'),
            ('Murat',   'Doğan',  31, 80.0, 180, 'Erkek',  'Dayanıklılık'),
            ('Fatma',   'Kılıç',  26, 65.0, 170, 'Kadın',  'Kas Kazanma'),
        ]
        for ad, soyad, yas, kilo, boy, cinsiyet, hedef in sporcular:
            bmi = round(kilo / ((boy/100) ** 2), 1)
            c.execute("""INSERT INTO sporcular (ad,soyad,yas,kilo,boy,cinsiyet,hedef,bmi)
                VALUES (?,?,?,?,?,?,?,?)""", (ad,soyad,yas,kilo,boy,cinsiyet,hedef,bmi))

        # Antrenmanlar
        antrenmanlar = [
            ('Koşu',            'Kardiyo',   0, 30, 'Kolay',       'Tempolu koşu 5km',             250),
            ('Yüzme',           'Kardiyo',   1,  0, 'Orta',        'Serbest yüzme 1500m',          400),
            ('Ağırlık Çalışması','Kuvvet',   1,  0, 'Zor',         'Full body kuvvet antrenmanı',  350),
            ('Yoga',            'Esneklik',  1,  0, 'Kolay',       'Esneme ve rahatlama',          150),
            ('Bisiklet',        'Kardiyo',   0, 45, 'Orta',        'Düşük tempo bisiklet',         300),
            ('HIIT',            'Kardiyo',   0, 20, 'Zor',         'Yüksek tempo interval',        400),
            ('Pilates',         'Esneklik',  0, 45, 'Orta',        'Core ve duruş çalışması',      200),
            ('Squat & Deadlift','Kuvvet',    1, 15, 'Zor',         'Alt vücut güç antrenmanı',     380),
            ('Bench Press',     'Kuvvet',    0, 45, 'Orta',        'Üst vücut itme hareketi',      280),
            ('Karışık Dövüş',   'Kardiyo',   1,  0, 'Profesyonel','Boks + kick boks',             500),
        ]
        for ad, kat, saat, dakika, zorluk, aciklama, kalori in antrenmanlar:
            c.execute("""INSERT INTO antrenmanlar
                (ad,kategori,saat,dakika,zorluk_seviyesi,aciklama,kalori_tahmini)
                VALUES (?,?,?,?,?,?,?)""", (ad,kat,saat,dakika,zorluk,aciklama,kalori))

        # Takipler (30 adet)
        now = datetime.now()
        for i in range(30):
            sporcu_id  = random.randint(1, 10)
            antrenman_id = random.randint(1, 10)
            kalori = random.randint(150, 550)
            nabiz  = random.randint(100, 175)
            sure   = random.randint(20, 90)
            tarih  = (now - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d %H:%M')
            c.execute("""INSERT INTO takipler
                (sporcu_id,antrenman_id,kalori,nabiz,sure_dakika,tarih)
                VALUES (?,?,?,?,?,?)""", (sporcu_id, antrenman_id, kalori, nabiz, sure, tarih))

        # İlerleme kayıtları
        for sporcu_id in range(1, 11):
            sporcu = c.execute("SELECT kilo FROM sporcular WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            if not sporcu: continue
            baslangic_kilo = sporcu[0]
            for gun in range(5, 0, -1):
                tarih = (now - timedelta(days=gun*7)).strftime('%Y-%m-%d')
                kilo = round(baslangic_kilo + random.uniform(-1.5, 0.5), 1)
                c.execute("INSERT INTO ilerlemeler (sporcu_id,kilo,tarih) VALUES (?,?,?)",
                          (sporcu_id, kilo, tarih))
        # ── Tier 2: Antrenman Programları ──
        _r = random
        program_verileri = [
            (1, 'Kilo Verme 4 Hafta', 'Kilo Verme', 4),
            (2, 'Kas Yapma Programı', 'Kas Kazanma', 8),
            (3, 'Form Koruma Rutin', 'Form Koruma', 6),
            (4, 'Dayanıklılık Artırma', 'Dayanıklılık', 4),
            (5, 'Başlangıç Programı', 'Genel Sağlık', 4),
        ]
        for sporcu_id, adi, hedef, sure in program_verileri:
            c.execute("""INSERT INTO antrenman_programlari
                (sporcu_id,program_adi,hedef,sure_hafta,aktif) VALUES (?,?,?,?,1)""",
                (sporcu_id, adi, hedef, sure))
            prog_id = c.lastrowid
            gunler = ['Pazartesi','Çarşamba','Cuma']
            for gun in gunler:
                ant_id = _r.randint(1, 10)
                c.execute("""INSERT INTO program_gunleri
                    (program_id,gun_adi,antrenman_id,set_sayisi,tekrar,agirlik,dinlenme_saniye)
                    VALUES (?,?,?,?,?,?,?)""",
                    (prog_id, gun, ant_id, _r.randint(3,5), _r.randint(8,15),
                     _r.choice([0,20,40,60,80,100]), _r.choice([45,60,90,120])))

        # ── Tier 2: Günlük Log Kayıtları ──
        now = datetime.now()
        for i in range(40):
            sporcu_id  = _r.randint(1, 10)
            antrenman_id = _r.randint(1, 10)
            tarih = (now - timedelta(days=_r.randint(0, 30))).strftime('%Y-%m-%d')
            sure  = _r.randint(20, 90)
            kalori = _r.randint(100, 600)
            max_nabiz = _r.randint(140, 185)
            ort_nabiz = _r.randint(110, max_nabiz - 10)
            zorluk = _r.randint(3, 9)
            yorgunluk = _r.randint(2, 8)
            log_id = c.execute("""INSERT INTO gunluk_log
                (sporcu_id,antrenman_id,tarih,sure_dakika,kalori_yakilan,
                 max_nabiz,ort_nabiz,zorluk_hissi,yorgunluk,durum)
                VALUES (?,?,?,?,?,?,?,?,?,'Tamamlandi') RETURNING log_id""",
                (sporcu_id, antrenman_id, tarih, sure, kalori,
                 max_nabiz, ort_nabiz, zorluk, yorgunluk)).fetchone()[0]
            # Set kayıtları
            for set_no in range(1, _r.randint(3,5)):
                c.execute("""INSERT INTO set_kayitlari
                    (log_id,set_no,tekrar,agirlik,tamamlandi) VALUES (?,?,?,?,1)""",
                    (log_id, set_no, _r.randint(8,15), _r.choice([0,20,40,60,80])))

        # ── Tier 3: Vücut Ölçümleri ──
        now_dt = datetime.now()
        for sporcu_id in range(1, 11):
            sporcu_row = c.execute("SELECT kilo FROM sporcular WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            if not sporcu_row: continue
            baz_kilo = sporcu_row[0]
            for ay in range(4, 0, -1):
                tarih = (now_dt - timedelta(days=ay*30)).strftime('%Y-%m-%d')
                kilo = round(baz_kilo + _r.uniform(-2, 1), 1)
                c.execute("""INSERT INTO vucut_olcumleri
                    (sporcu_id,tarih,kilo,bel,gogus,kalca,sag_kol,sol_kol,sag_bacak,sol_bacak,vucut_yag_orani)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (sporcu_id, tarih, kilo,
                     round(70 + _r.uniform(-5,5), 1), round(90 + _r.uniform(-5,5), 1),
                     round(95 + _r.uniform(-5,5), 1), round(35 + _r.uniform(-3,3), 1),
                     round(34 + _r.uniform(-3,3), 1), round(58 + _r.uniform(-5,5), 1),
                     round(57 + _r.uniform(-5,5), 1), round(18 + _r.uniform(-3,5), 1)))

        # ── Tier 3: Hedefler ──
        hedef_verileri = [
            (1, 'Kilo',    'Hedef Kiloya Ulaş',          75.5, 70.0, 72.0),
            (2, 'Kilo',    'Kas Kazanarak Kilo Al',       62.0, 67.0, 64.0),
            (3, 'Güç',     'Bench Press 100kg',           60.0, 100.0, 75.0),
            (4, "Kardiyo", "5km'yi 25dk'da Koş",       32.0, 25.0, 28.0),
            (5, 'Antrenman','30 Gün Kesintisiz Antrenman',0,   30,   18),
            (6, 'Kilo',    'Hedef Kiloya Ulaş',          55.5, 52.0, 53.5),
            (7, 'Kilo',    'Kilo Azalt',                  92.0, 85.0, 89.0),
            (8, 'Güç',     'Squat 80kg',                  40.0, 80.0, 55.0),
            (9, 'Kardiyo', '10km Koşu',                   0,   10,    6),
            (10,"Ölçüm",   "Bel 80cm'ye İndir",         88.0, 80.0, 84.0),
        ]
        hedef_tarihi = (now_dt + timedelta(days=90)).strftime('%Y-%m-%d')
        for sporcu_id, tip, ad, bas, hedef, mevcut in hedef_verileri:
            durum = 'Tamamlandi' if mevcut >= hedef else 'Devam'
            c.execute("""INSERT INTO hedefler
                (sporcu_id,hedef_tipi,hedef_adi,baslangic_degeri,hedef_degeri,mevcut_deger,hedef_tarihi,durum)
                VALUES (?,?,?,?,?,?,?,?)""",
                (sporcu_id, tip, ad, bas, hedef, mevcut, hedef_tarihi, durum))

        # ── Tier 3: Streak Kayıtları — kesintisiz, bugün dahil ──
        streak_gunleri = [19, 18, 17, 13, 13, 12, 10, 10, 8, 7]
        for sporcu_id in range(1, 11):
            gun_sayisi = streak_gunleri[sporcu_id - 1]
            for gun in range(gun_sayisi - 1, -1, -1):
                tarih = (now_dt - timedelta(days=gun)).strftime('%Y-%m-%d')
                try:
                    c.execute("INSERT INTO streak_kayitlari (sporcu_id,tarih,antrenman_yapildi) VALUES (?,?,1)",
                              (sporcu_id, tarih))
                except: pass

        # ── Tier 4: Besinler ──
        besinler = [
            ('Yumurta',          'Protein',    60,  90,  7.5, 0.6,  6.3, 0.0),
            ('Tavuk Göğsü',      'Protein',   100, 165, 31.0, 0.0,  3.6, 0.0),
            ('Somon',            'Protein',   100, 208, 20.0, 0.0, 13.0, 0.0),
            ('Yulaf Ezmesi',     'Karbonhidrat',100, 389, 17.0, 66.0, 7.0, 10.0),
            ('Pirinç (Beyaz)',   'Karbonhidrat',100, 130,  2.4, 28.0, 0.3, 0.4),
            ('Kepekli Ekmek',    'Karbonhidrat', 30,  75,  3.5, 14.0, 1.0, 1.9),
            ('Muz',              'Meyve',        120, 89,  1.1, 23.0, 0.3, 2.6),
            ('Badem',            'Yağ',          30, 170,  6.0,  6.0, 15.0, 3.5),
            ('Zeytinyağı',       'Yağ',          15, 120,  0.0,  0.0, 14.0, 0.0),
            ('Lor Peyniri',      'Protein',     100, 100, 11.0,  3.4,  4.3, 0.0),
            ('Brokoli',          'Sebze',        100,  34,  2.8,  7.0,  0.4, 2.6),
            ('Ispanak',          'Sebze',        100,  23,  2.9,  3.6,  0.4, 2.2),
            ('Domates',          'Sebze',        100,  18,  0.9,  3.9,  0.2, 1.2),
            ('Süt (Tam Yağlı)',  'Süt Ürünleri', 200,  60,  3.2,  4.8,  3.3, 0.0),
            ('Yoğurt (Sade)',    'Süt Ürünleri', 150,  60,  4.0,  4.7,  3.3, 0.0),
            ('Patates',          'Karbonhidrat', 100,  77,  2.0, 17.0,  0.1, 2.2),
            ('Mercimek',         'Protein',      100, 116,  9.0, 20.0,  0.4, 7.9),
            ('Ceviz',            'Yağ',           30, 185,  4.3,  4.0, 18.0, 1.9),
            ('Mısır Gevreği',    'Karbonhidrat', 100, 371,  7.0, 84.0,  1.0, 2.0),
            ('Portakal',         'Meyve',        130,  62,  1.2, 15.0,  0.2, 3.1),
        ]
        for ad, kat, porsiyon, kal, prot, karb, yag, lif in besinler:
            c.execute("""INSERT INTO besinler (ad,kategori,porsiyon_gram,kalori,protein,karbonhidrat,yag,lif)
                VALUES (?,?,?,?,?,?,?,?)""", (ad, kat, porsiyon, kal, prot, karb, yag, lif))

        # ── Tier 4: Kalori Hedefleri ──
        hedef_verileri = [
            (1, 1800, 140, 180, 60),  # Kilo verme
            (2, 2200, 160, 220, 70),  # Kas kazanma
            (3, 2000, 150, 200, 65),  # Form koruma
            (4, 1700, 130, 170, 58),  # Kilo verme
            (5, 2500, 180, 250, 80),  # Dayanıklılık
            (6, 2100, 155, 210, 68),  # Kas kazanma
            (7, 1900, 145, 185, 62),  # Kilo verme
            (8, 2000, 150, 200, 65),  # Form koruma
            (9, 2600, 190, 260, 85),  # Dayanıklılık
            (10,2300, 170, 230, 75),  # Kas kazanma
        ]
        for sporcu_id, kal, prot, karb, yag in hedef_verileri:
            c.execute("""INSERT INTO kalori_hedefleri (sporcu_id,gunluk_kalori,protein_gram,karb_gram,yag_gram)
                VALUES (?,?,?,?,?)""", (sporcu_id, kal, prot, karb, yag))

        # ── Tier 4: Günlük Beslenme (son 14 gün) ──
        ogunler = ['Kahvalti', 'Ogle', 'Aksam', 'Ara Ogun']
        now_dt = datetime.now()
        besin_count = 20
        for sporcu_id in range(1, 11):
            for gun in range(14, 0, -1):
                tarih = (now_dt - timedelta(days=gun)).strftime('%Y-%m-%d')
                ogun_sayisi = _r.randint(2, 4)
                for ogun in _r.sample(ogunler, ogun_sayisi):
                    besin_id = _r.randint(1, besin_count)
                    miktar = _r.choice([100, 150, 200, 250, 80, 120])
                    besin_row = c.execute("SELECT * FROM besinler WHERE besin_id=?", (besin_id,)).fetchone()
                    if not besin_row: continue
                    oran = miktar / besin_row['porsiyon_gram']
                    c.execute("""INSERT INTO gunluk_beslenme
                        (sporcu_id,besin_id,tarih,ogun,miktar_gram,
                         kalori_toplam,protein_toplam,karb_toplam,yag_toplam)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                        (sporcu_id, besin_id, tarih, ogun, miktar,
                         round(besin_row['kalori'] * oran, 1),
                         round(besin_row['protein'] * oran, 1),
                         round(besin_row['karbonhidrat'] * oran, 1),
                         round(besin_row['yag'] * oran, 1)))

        # ── Tier 6: Rozetler (sabit tanımlar) ──
        rozetler = [
            ('İlk Adım',        'İlk antrenmanını tamamla',       '🥾', 'Başlangıç', 10,  'antrenman_sayisi', 1),
            ('Haftanın Sporcusu','Haftada 5 antrenman yap',        '⭐', 'Devam',      30,  'haftalik_antrenman', 5),
            ('Alev Topu',       '7 gün streak yap',                '🔥', 'Streak',     50,  'streak', 7),
            ('Demir Adam',      '30 gün streak yap',               '🦾', 'Streak',     150, 'streak', 30),
            ('Kalorist',        '10.000 kalori yak',               '🔥', 'Kalori',     40,  'toplam_kalori', 10000),
            ('Maraton',         '100 saat antrenman yap',          '🏃', 'Süre',       100, 'toplam_sure_saat', 100),
            ('Güç Ustası',      '50 antrenman tamamla',            '💪', 'Dayanıklılık',80, 'antrenman_sayisi', 50),
            ('Yüzüncü Terleme', '100 antrenman tamamla',           '🏅', 'Dayanıklılık',200,'antrenman_sayisi', 100),
            ('Beslenme Bilinci','7 gün üst üste beslenme takibi',  '🥗', 'Beslenme',   40,  'beslenme_streak', 7),
            ('Hedef Avcısı',    'İlk hedefini tamamla',            '🎯', 'Hedef',      60,  'tamamlanan_hedef', 1),
            ('Çok Yönlü',       '5 farklı kategoride antrenman',   '🌟', 'Çeşitlilik', 70,  'kategori_sayisi', 5),
            ('Sabah Kuşu',      '20 antrenman tamamla',            '🌅', 'Devam',      50,  'antrenman_sayisi', 20),
        ]
        for ad, aciklama, icon, kat, puan, kosul_tipi, kosul_deger in rozetler:
            c.execute("""INSERT INTO rozetler
                (rozet_adi,aciklama,icon,kategori,puan,kosul_tipi,kosul_deger)
                VALUES (?,?,?,?,?,?,?)""",
                (ad, aciklama, icon, kat, puan, kosul_tipi, kosul_deger))

        # ── Tier 6: Sporcu Puanları başlat ──
        seviyeler = [
            (0,    'Başlangıç', 1),
            (100,  'Bronz',     2),
            (300,  'Gümüş',     3),
            (600,  'Altın',     4),
            (1000, 'Platin',    5),
            (2000, 'Elmas',     6),
        ]
        for sporcu_id in range(1, 11):
            puan = _r.randint(50, 800)
            seviye_adi = 'Başlangıç'; seviye_no = 1
            for esik, sev_adi, sev_no in seviyeler:
                if puan >= esik:
                    seviye_adi = sev_adi; seviye_no = sev_no
            c.execute("""INSERT INTO sporcu_puanlari
                (sporcu_id,toplam_puan,seviye,seviye_no) VALUES (?,?,?,?)""",
                (sporcu_id, puan, seviye_adi, seviye_no))

        # ── Tier 6: Sporcu Rozetleri (rastgele) ──
        rozet_ids = list(range(1, len(rozetler)+1))
        for sporcu_id in range(1, 11):
            kazanilan = _r.sample(rozet_ids, _r.randint(2, 6))
            for rozet_id in kazanilan:
                tarih = (datetime.now() - timedelta(days=_r.randint(1,60))).strftime('%Y-%m-%d %H:%M')
                try:
                    c.execute("""INSERT INTO sporcu_rozetleri
                        (sporcu_id,rozet_id,kazanma_tarihi) VALUES (?,?,?)""",
                        (sporcu_id, rozet_id, tarih))
                except: pass

        # ── Tier 7: 1RM Kayıtları — trend için 4 ölçüm/egzersiz ──
        egzersizler = ['Bench Press','Squat','Deadlift','Overhead Press','Barbell Row']
        # Her sporcu için başlangıç ağırlıkları (kg)
        baslangic = {
            'Bench Press':   [50,60,70,80,90,60,80,50,70,65],
            'Squat':         [60,70,80,90,100,70,90,55,80,75],
            'Deadlift':      [70,80,90,100,110,80,100,60,90,85],
            'Overhead Press':[30,35,40,50,60,35,50,30,45,40],
            'Barbell Row':   [50,55,60,70,80,55,70,45,65,60],
        }
        for sporcu_id in range(1, 11):
            for egz in egzersizler:
                baz = baslangic[egz][sporcu_id - 1]
                # 4 kayıt — giderek artan ağırlık (ilerleme trendi)
                for ay in range(4, 0, -1):
                    agirlik = baz + (4 - ay) * _r.randint(2, 5)
                    tekrar  = _r.choice([1, 3, 5, 8])
                    bir_rm  = round(agirlik * (1 + tekrar / 30), 1)
                    tarih   = (now_dt - timedelta(days=ay * 20 + _r.randint(0,5))).strftime('%Y-%m-%d')
                    c.execute("""INSERT INTO bir_rm_kayitlari
                        (sporcu_id,egzersiz_adi,agirlik,tekrar,hesaplanan_1rm,tarih)
                        VALUES (?,?,?,?,?,?)""",
                        (sporcu_id, egz, agirlik, tekrar, bir_rm, tarih))

        # ── Tier 7: Radar Değerleri ──
        for sporcu_id in range(1, 11):
            c.execute("""INSERT INTO radar_degerleri
                (sporcu_id,guc_puani,kardiyo_puani,esneklik_puani,
                 denge_puani,dayaniklilik_puani,hiz_puani)
                VALUES (?,?,?,?,?,?,?)""",
                (sporcu_id,
                 round(_r.uniform(30, 95), 1), round(_r.uniform(30, 95), 1),
                 round(_r.uniform(20, 90), 1), round(_r.uniform(25, 85), 1),
                 round(_r.uniform(30, 95), 1), round(_r.uniform(25, 90), 1)))

        # ── Tier 8: Analiz Notları ──
        not_verileri = [
            (1, 'Öneri', 'Kardiyo Artır', 'Kilo hedefine ulaşmak için haftalık 3 kardiyo hedefle.', 'Yüksek'),
            (2, 'Gözlem', 'Güzel İlerleme', 'Son 2 haftada kas kütlesinde belirgin artış var.', 'Normal'),
            (3, 'Uyarı', 'Overtraining Riski', 'Son 7 günde 6 antrenman — dinlenme günü ekle.', 'Kritik'),
            (4, 'Öneri', 'Protein Artır', 'Günlük protein hedefini tutturamıyor. Beslenmeyi gözden geçir.', 'Yüksek'),
            (5, 'Gözlem', 'Streak Devam', '14 günlük streak! Motivasyon çok yüksek.', 'Normal'),
        ]
        for sporcu_id, tip, baslik, icerik, oncelik in not_verileri:
            c.execute("""INSERT INTO analiz_notlari (sporcu_id,not_tipi,baslik,icerik,oncelik)
                VALUES (?,?,?,?,?)""", (sporcu_id, tip, baslik, icerik, oncelik))


        # ── Tier 9: Antrenman Şablonları + Günleri ──
        # sablon_id: 1=PPL, 2=Arnold, 3=FullBody, 4=Kardiyo+Güç, 5=HIIT+Kuvvet
        sablonlar = [
            ('Push Pull Legs (PPL)',  'Üst vücut push-pull-legs split', 'Kuvvet', 6),
            ('Arnold Split',          '3 gün split — klasik powerbuilding', 'Kuvvet', 8),
            ('Başlangıç Full Body',   'Haftada 3 gün tam vücut', 'Genel', 4),
            ('Kardiyo & Güç Dengesi', 'Kardiyo ve kuvveti dengeler', 'Genel', 4),
            ('HIIT + Kuvvet',         'Yüksek yoğunluk interval + güç', 'HIIT', 6),
        ]
        for ad, aciklama, kat, hafta in sablonlar:
            c.execute("""INSERT INTO antrenman_sablonlari (sablon_adi,aciklama,kategori,sure_hafta)
                VALUES (?,?,?,?)""", (ad, aciklama, kat, hafta))

        # Şablon günleri — her şablona 3-6 gün
        # antrenman_id 1=Koşu,2=Yüzme,3=Ağırlık,4=Yoga,5=Bisiklet,6=HIIT,7=Pilates,8=Squat&DL,9=BenchPress,10=Dövüş
        sablon_gunleri = {
            1: [  # PPL
                (1,'Pazartesi', 9, 4, 12, 60, 90),   # Push — Bench Press
                (2,'Çarşamba',  3, 4, 10, 80, 90),   # Pull — Ağırlık
                (3,'Cuma',      8, 4, 8,  100, 120), # Legs — Squat&DL
                (4,'Pazartesi', 9, 4, 12, 65, 90),
                (5,'Çarşamba',  3, 4, 10, 85, 90),
                (6,'Cuma',      8, 4, 8,  105, 120),
            ],
            2: [  # Arnold Split
                (1,'Pazartesi', 9, 5, 10, 70, 90),
                (2,'Salı',      3, 5, 8,  80, 90),
                (3,'Çarşamba',  8, 5, 8,  90, 120),
                (4,'Perşembe',  9, 5, 10, 72, 90),
                (5,'Cuma',      3, 5, 8,  82, 90),
                (6,'Cumartesi', 8, 5, 8,  92, 120),
            ],
            3: [  # Başlangıç Full Body
                (1,'Pazartesi', 3, 3, 12, 40, 60),
                (2,'Çarşamba',  3, 3, 12, 40, 60),
                (3,'Cuma',      3, 3, 12, 42, 60),
            ],
            4: [  # Kardiyo & Güç
                (1,'Pazartesi', 1, 0, 0,  0,  60),   # Koşu
                (2,'Salı',      3, 3, 12, 50, 90),   # Ağırlık
                (3,'Perşembe',  5, 0, 0,  0,  60),   # Bisiklet
                (4,'Cuma',      3, 3, 12, 52, 90),
            ],
            5: [  # HIIT + Kuvvet
                (1,'Pazartesi', 6, 0, 0,  0,  45),   # HIIT
                (2,'Salı',      9, 4, 10, 60, 90),   # Bench
                (3,'Perşembe',  6, 0, 0,  0,  45),
                (4,'Cuma',      8, 4, 8,  80, 120),  # Squat
                (5,'Cumartesi', 6, 0, 0,  0,  45),
            ],
        }
        for sablon_id, gunler in sablon_gunleri.items():
            for gun_sira, gun_adi, ant_id, set_s, tekrar, agirlik, dinlenme in gunler:
                c.execute("""INSERT INTO sablon_gunleri
                    (sablon_id,gun_sira,gun_adi,antrenman_id,set_sayisi,tekrar,agirlik,dinlenme_saniye)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (sablon_id, gun_sira, gun_adi, ant_id, set_s, tekrar, agirlik, dinlenme))

        # ── Tier 9: Sistem Ayarları ──
        ayarlar = [
            ('birim',           'kg',          'Ağırlık birimi (kg/lbs)'),
            ('tema',            'dark_luxury', 'Uygulama teması'),
            ('dil',             'tr',          'Arayüz dili'),
            ('kalori_birimi',   'kcal',        'Kalori birimi'),
            ('hatirlatici_aktif','1',          'Hatırlatıcı sistemi aktif mi'),
            ('yedekleme_yolu',  '',            'Otomatik yedekleme dosya yolu'),
            ('versiyon',        '1.0.0',       'Uygulama versiyonu'),
        ]
        for anahtar, deger, aciklama in ayarlar:
            c.execute("""INSERT INTO sistem_ayarlari (anahtar,deger,aciklama)
                VALUES (?,?,?)""", (anahtar, deger, aciklama))

        # ── Tier 9: Hatırlatıcılar ──
        hatirlaticilar = [
            (None, 'Günlük Antrenman', 'Bugün antrenman zamanı!', 'Gunluk', '08:00'),
            (None, 'Su İç', 'Günde 2-3 litre su iç!', 'Gunluk', '10:00'),
            (None, 'Haftalık Ölçüm', 'Bu hafta vücut ölçümünü almayı unutma!', 'Haftalik', '09:00'),
        ]
        for sporcu_id, baslik, mesaj, gun_tipi, saat in hatirlaticilar:
            c.execute("""INSERT INTO hatirlaticilar (sporcu_id,baslik,mesaj,gun_tipi,saat,aktif)
                VALUES (?,?,?,?,?,1)""", (sporcu_id, baslik, mesaj, gun_tipi, saat))

        # ── Tier 10: Aktivite Feed ──
        feed_verileri = [
            (1, 'antrenman', '🏋️ Antrenman Tamamlandı', 'Bench Press — 80kg × 5 tekrar', '💪'),
            (1, 'rozet',     '🏅 Yeni Rozet!',           'Alev Topu rozeti kazanıldı',   '🔥'),
            (2, 'hedef',     '🎯 Hedef Güncellendi',     'Kilo hedefi: 67kg → 65kg',     '🎯'),
            (3, 'olcum',     '📏 Ölçüm Alındı',          'Kilo: 84.5kg — 0.5kg azaldı',  '📉'),
            (4, 'streak',    '🔥 14 Gün Streak!',        'Harika devam, durma!',          '🏆'),
            (5, 'beslenme',  '🥗 Kalori Hedefinde',      'Günlük kalori %95 tamamlandı', '✅'),
        ]
        for sporcu_id, tip, baslik, detay, ikon in feed_verileri:
            c.execute("""INSERT INTO aktivite_feed (sporcu_id,aktivite_tipi,baslik,detay,ikon)
                VALUES (?,?,?,?,?)""", (sporcu_id, tip, baslik, detay, ikon))

    # ── TİER 10: PRODUCTION DB METODLARI ───────────────────────
    def add_aktivite_feed(self, sporcu_id, tip, baslik, detay='', ikon='📌'):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO aktivite_feed (sporcu_id,aktivite_tipi,baslik,detay,ikon)
                VALUES (?,?,?,?,?)""", (sporcu_id, tip, baslik, detay, ikon))

    def get_aktivite_feed(self, sporcu_id=None, limit=20):
        with self.get_connection() as conn:
            q = """SELECT f.*, s.ad||' '||s.soyad as sporcu_adi
                   FROM aktivite_feed f
                   JOIN sporcular s ON f.sporcu_id=s.sporcu_id"""
            p = []
            if sporcu_id:
                q += " WHERE f.sporcu_id=?"; p.append(sporcu_id)
            q += f" ORDER BY f.tarih DESC LIMIT {limit}"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def get_gelismis_dashboard_stats(self):
        """Tier 10 dashboard için genişletilmiş istatistikler."""
        with self.get_connection() as conn:
            s = {}
            s['toplam_sporcu']      = conn.execute("SELECT COUNT(*) FROM sporcular WHERE durum='Aktif'").fetchone()[0]
            s['toplam_antrenman']   = conn.execute("SELECT COUNT(*) FROM antrenmanlar WHERE durum='Aktif'").fetchone()[0]
            s['toplam_log']         = conn.execute("SELECT COUNT(*) FROM gunluk_log WHERE durum='Tamamlandi'").fetchone()[0]
            s['toplam_kalori_yak']  = conn.execute("SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log WHERE durum='Tamamlandi'").fetchone()[0]
            s['toplam_sure']        = conn.execute("SELECT COALESCE(SUM(sure_dakika),0) FROM gunluk_log WHERE durum='Tamamlandi'").fetchone()[0]
            s['toplam_rozet']       = conn.execute("SELECT COUNT(*) FROM sporcu_rozetleri").fetchone()[0]
            s['aktif_streak']       = conn.execute("SELECT COUNT(DISTINCT sporcu_id) FROM streak_kayitlari WHERE tarih=date('now')").fetchone()[0]
            s['bu_hafta_log']       = conn.execute("SELECT COUNT(*) FROM gunluk_log WHERE durum='Tamamlandi' AND tarih>=date('now','-7 days')").fetchone()[0]
            # Kategori dağılımı
            rows = conn.execute("""SELECT a.kategori, COUNT(*) as sayi
                FROM gunluk_log g JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.durum='Tamamlandi' GROUP BY a.kategori""").fetchall()
            s['kategori_dagilimi'] = [(r[0], r[1]) for r in rows]
            # Haftalık log trendi (son 8 hafta)
            rows2 = conn.execute("""SELECT strftime('%W',tarih) as hafta, COUNT(*) as sayi
                FROM gunluk_log WHERE durum='Tamamlandi' AND tarih>=date('now','-56 days')
                GROUP BY hafta ORDER BY hafta""").fetchall()
            s['haftalik_trend'] = [(f"H{i+1}", r[1]) for i, r in enumerate(rows2)]
            # Hedef dağılımı
            rows3 = conn.execute("SELECT hedef, COUNT(*) FROM sporcular WHERE durum='Aktif' GROUP BY hedef").fetchall()
            s['hedef_dagilimi'] = [(r[0], r[1]) for r in rows3]
            # En aktif sporcu
            rows4 = conn.execute("""SELECT s.ad||' '||s.soyad as ad, COUNT(*) as sayi
                FROM gunluk_log g JOIN sporcular s ON g.sporcu_id=s.sporcu_id
                WHERE g.durum='Tamamlandi' AND g.tarih>=date('now','-30 days')
                GROUP BY g.sporcu_id ORDER BY sayi DESC LIMIT 5""").fetchall()
            s['en_aktif'] = [(r[0], r[1]) for r in rows4]
            # Toplam puan
            s['toplam_puan'] = conn.execute("SELECT COALESCE(SUM(toplam_puan),0) FROM sporcu_puanlari").fetchone()[0]
            return s

    def get_son_aktiviteler(self, limit=10):
        """Son kullanıcı aktiviteleri."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT g.tarih, s.ad||' '||s.soyad as sporcu_adi,
                       a.ad as antrenman_adi, g.kalori_yakilan, g.sure_dakika
                FROM gunluk_log g
                JOIN sporcular s ON g.sporcu_id=s.sporcu_id
                JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.durum='Tamamlandi'
                ORDER BY g.olusturma_tarihi DESC LIMIT ?""", (limit,)).fetchall()
            return [dict(r) for r in rows]

    def export_excel_ozet(self, dosya_yolu):
        """Tüm sistem verisini openpyxl ile Excel'e aktar."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()

            def stil_header(cell, bg='1a1a2e', fg='ffffff'):
                cell.font = Font(bold=True, color=fg, size=11)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Sayfa 1: Sporcular
            ws1 = wb.active; ws1.title = 'Sporcular'
            headers1 = ['ID','Ad','Soyad','Yaş','Kilo','Boy','BMI','Cinsiyet','Hedef','Kayıt']
            for col, h in enumerate(headers1, 1):
                cell = ws1.cell(1, col, h); stil_header(cell)
                ws1.column_dimensions[chr(64+col)].width = 14
            with self.get_connection() as conn:
                for row_idx, row in enumerate(conn.execute("SELECT sporcu_id,ad,soyad,yas,kilo,boy,bmi,cinsiyet,hedef,kayit_tarihi FROM sporcular WHERE durum='Aktif' ORDER BY ad").fetchall(), 2):
                    for col, val in enumerate(row, 1):
                        ws1.cell(row_idx, col, val)

            # Sayfa 2: Antrenman Logları
            ws2 = wb.create_sheet('Antrenman Logları')
            headers2 = ['Tarih','Sporcu','Antrenman','Kategori','Süre(dk)','Kalori','Zorluk','Yorgunluk']
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(1, col, h); stil_header(cell)
                ws2.column_dimensions[chr(64+col)].width = 16
            with self.get_connection() as conn:
                rows = conn.execute("""SELECT g.tarih, s.ad||' '||s.soyad, a.ad, a.kategori,
                    g.sure_dakika, g.kalori_yakilan, g.zorluk_hissi, g.yorgunluk
                    FROM gunluk_log g
                    JOIN sporcular s ON g.sporcu_id=s.sporcu_id
                    JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                    WHERE g.durum='Tamamlandi' ORDER BY g.tarih DESC""").fetchall()
                for row_idx, row in enumerate(rows, 2):
                    for col, val in enumerate(row, 1):
                        ws2.cell(row_idx, col, val)

            # Sayfa 3: Beslenme Özeti
            ws3 = wb.create_sheet('Beslenme')
            headers3 = ['Tarih','Sporcu','Öğün','Besin','Miktar(g)','Kalori','Protein','Karb','Yağ']
            for col, h in enumerate(headers3, 1):
                cell = ws3.cell(1, col, h); stil_header(cell)
                ws3.column_dimensions[chr(64+col)].width = 14
            with self.get_connection() as conn:
                rows3 = conn.execute("""SELECT gb.tarih, s.ad||' '||s.soyad, gb.ogun, b.ad,
                    gb.miktar_gram, gb.kalori_toplam, gb.protein_toplam, gb.karb_toplam, gb.yag_toplam
                    FROM gunluk_beslenme gb
                    JOIN sporcular s ON gb.sporcu_id=s.sporcu_id
                    JOIN besinler b ON gb.besin_id=b.besin_id
                    ORDER BY gb.tarih DESC LIMIT 500""").fetchall()
                for row_idx, row in enumerate(rows3, 2):
                    for col, val in enumerate(row, 1):
                        ws3.cell(row_idx, col, val)

            # Sayfa 4: Liderboard
            ws4 = wb.create_sheet('Liderboard')
            headers4 = ['Sıra','Sporcu','Seviye','Puan','Rozet Sayısı']
            for col, h in enumerate(headers4, 1):
                cell = ws4.cell(1, col, h); stil_header(cell)
                ws4.column_dimensions[chr(64+col)].width = 18
            lb = self.get_leaderboard()
            for row_idx, d in enumerate(lb, 2):
                ws4.cell(row_idx, 1, row_idx-1)
                ws4.cell(row_idx, 2, f"{d['ad']} {d['soyad']}")
                ws4.cell(row_idx, 3, d['seviye'])
                ws4.cell(row_idx, 4, d['toplam_puan'])
                ws4.cell(row_idx, 5, d['rozet_sayisi'])

            wb.save(dosya_yolu)
            return True
        except ImportError:
            return 'openpyxl_yok'
        except Exception as e:
            return str(e)

    # ── TİER 9: PRODUCTION DB METODLARI ────────────────────────
    def get_sistem_ayari(self, anahtar, varsayilan=''):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT deger FROM sistem_ayarlari WHERE anahtar=?", (anahtar,)).fetchone()
            return row[0] if row else varsayilan

    def set_sistem_ayari(self, anahtar, deger, aciklama=''):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO sistem_ayarlari (anahtar,deger,aciklama)
                VALUES (?,?,?)
                ON CONFLICT(anahtar) DO UPDATE SET deger=excluded.deger,
                guncelleme_tarihi=datetime('now','localtime')""",
                (anahtar, deger, aciklama))

    def get_tum_ayarlar(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM sistem_ayarlari ORDER BY anahtar").fetchall()]

    def get_antrenman_sablonlari(self):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM antrenman_sablonlari ORDER BY olusturma_tarihi DESC").fetchall()]

    def get_sablon_gunleri(self, sablon_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT sg.*, a.ad as antrenman_adi, a.kategori
                FROM sablon_gunleri sg
                LEFT JOIN antrenmanlar a ON sg.antrenman_id=a.antrenman_id
                WHERE sg.sablon_id=? ORDER BY sg.gun_sira""", (sablon_id,)).fetchall()]

    def sablon_sporcu_uygula(self, sablon_id, sporcu_id):
        """Şablonu sporcu programına uygula."""
        with self.get_connection() as conn:
            sablon = conn.execute(
                "SELECT * FROM antrenman_sablonlari WHERE sablon_id=?", (sablon_id,)).fetchone()
            if not sablon: return False
            sablon = dict(sablon)
            gunler = self.get_sablon_gunleri(sablon_id)
            prog_id = conn.execute("""INSERT INTO antrenman_programlari
                (sporcu_id,program_adi,hedef,sure_hafta)
                VALUES (?,?,?,?)""",
                (sporcu_id, sablon['sablon_adi'],
                 sablon['kategori'], sablon['sure_hafta'])).lastrowid
            for g in gunler:
                conn.execute("""INSERT INTO program_gunleri
                    (program_id,gun_adi,antrenman_id,set_sayisi,tekrar,agirlik,dinlenme_saniye,notlar)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (prog_id, g['gun_adi'], g['antrenman_id'],
                     g['set_sayisi'], g['tekrar'], g['agirlik'],
                     g['dinlenme_saniye'], g['notlar']))
        return prog_id

    def get_hatirlaticilar(self, sporcu_id=None):
        with self.get_connection() as conn:
            q = "SELECT * FROM hatirlaticilar WHERE aktif=1"
            p = []
            if sporcu_id:
                q += " AND (sporcu_id=? OR sporcu_id IS NULL)"; p.append(sporcu_id)
            return [dict(r) for r in conn.execute(q + " ORDER BY saat", p).fetchall()]

    def add_hatirlatici(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO hatirlaticilar
                (sporcu_id,baslik,mesaj,gun_tipi,saat,aktif)
                VALUES (?,?,?,?,?,1)""",
                (data.get('sporcu_id'), data['baslik'], data.get('mesaj',''),
                 data.get('gun_tipi','Gunluk'), data.get('saat','09:00')))

    def toggle_hatirlatici(self, hatirlatici_id):
        with self.get_connection() as conn:
            conn.execute("""UPDATE hatirlaticilar SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END
                WHERE hatirlatici_id=?""", (hatirlatici_id,))

    def delete_hatirlatici(self, hatirlatici_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM hatirlaticilar WHERE hatirlatici_id=?", (hatirlatici_id,))

    def export_sporcu_csv(self, sporcu_id, dosya_yolu):
        """Sporcu verilerini CSV'ye aktar."""
        import csv
        with open(dosya_yolu, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            with self.get_connection() as conn:
                # Sporcu bilgisi
                writer.writerow(['=== SPORCU BİLGİSİ ==='])
                sporcu = conn.execute("SELECT * FROM sporcular WHERE sporcu_id=?",
                                      (sporcu_id,)).fetchone()
                if sporcu:
                    writer.writerow(list(sporcu.keys()))
                    writer.writerow(list(dict(sporcu).values()))
                writer.writerow([])
                # Antrenman logu
                writer.writerow(['=== ANTRENMAN KAYITLARI ==='])
                writer.writerow(['Tarih','Antrenman','Kategori','Süre(dk)','Kalori','Zorluk'])
                rows = conn.execute("""
                    SELECT g.tarih, a.ad, a.kategori, g.sure_dakika, g.kalori_yakilan, g.zorluk_hissi
                    FROM gunluk_log g JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                    WHERE g.sporcu_id=? AND g.durum='Tamamlandi' ORDER BY g.tarih DESC""",
                    (sporcu_id,)).fetchall()
                for r in rows: writer.writerow(list(r))
                writer.writerow([])
                # Vücut ölçümleri
                writer.writerow(['=== VÜCUT ÖLÇÜMLERİ ==='])
                writer.writerow(['Tarih','Kilo','Bel','Göğüs','Kalça','Yağ%'])
                rows2 = conn.execute("""SELECT tarih,kilo,bel,gogus,kalca,vucut_yag_orani
                    FROM vucut_olcumleri WHERE sporcu_id=? ORDER BY tarih DESC""",
                    (sporcu_id,)).fetchall()
                for r in rows2: writer.writerow(list(r))
                writer.writerow([])
                # Beslenme
                writer.writerow(['=== BESLENME KAYITLARI ==='])
                writer.writerow(['Tarih','Öğün','Besin','Miktar(g)','Kalori','Protein','Karb','Yağ'])
                rows3 = conn.execute("""
                    SELECT gb.tarih, gb.ogun, b.ad, gb.miktar_gram,
                           gb.kalori_toplam, gb.protein_toplam, gb.karb_toplam, gb.yag_toplam
                    FROM gunluk_beslenme gb JOIN besinler b ON gb.besin_id=b.besin_id
                    WHERE gb.sporcu_id=? ORDER BY gb.tarih DESC LIMIT 100""",
                    (sporcu_id,)).fetchall()
                for r in rows3: writer.writerow(list(r))
        return True

    def get_db_boyutu(self):
        import os
        try:
            return os.path.getsize(self.db_path)
        except: return 0

    def yedekle(self, hedef_yol):
        import shutil
        try:
            shutil.copy2(self.db_path, hedef_yol)
            return True
        except: return False

    def get_sistem_ozet(self):
        """Dashboard için sistem özeti."""
        with self.get_connection() as conn:
            return {
                'sporcu': conn.execute("SELECT COUNT(*) FROM sporcular WHERE durum='Aktif'").fetchone()[0],
                'antrenman': conn.execute("SELECT COUNT(*) FROM antrenmanlar WHERE durum='Aktif'").fetchone()[0],
                'log': conn.execute("SELECT COUNT(*) FROM gunluk_log WHERE durum='Tamamlandi'").fetchone()[0],
                'besin': conn.execute("SELECT COUNT(*) FROM besinler WHERE durum='Aktif'").fetchone()[0],
                'rozet': conn.execute("SELECT COUNT(*) FROM rozetler").fetchone()[0],
                'sablon': conn.execute("SELECT COUNT(*) FROM antrenman_sablonlari").fetchone()[0],
                'db_boyut': round(self.get_db_boyutu() / 1024, 1),
            }

    # ── TİER 8: GELİŞMİŞ ANALİZ DB METODLARI ──────────────────
    def get_kas_grubu_volum(self, sporcu_id, gun=30):
        """Kas grubu başına haftalık set/tekrar volümü."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT a.kategori,
                       COUNT(g.log_id) as antrenman_sayisi,
                       COALESCE(SUM(g.sure_dakika),0) as toplam_sure,
                       COALESCE(SUM(g.kalori_yakilan),0) as toplam_kalori,
                       COALESCE(AVG(g.zorluk_hissi),0) as ort_zorluk
                FROM gunluk_log g
                JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.sporcu_id=? AND g.durum='Tamamlandi'
                AND g.tarih >= date('now',?)
                GROUP BY a.kategori ORDER BY antrenman_sayisi DESC""",
                (sporcu_id, f'-{gun} days')).fetchall()
            return [dict(r) for r in rows]

    def get_overtraining_uyari(self, sporcu_id):
        """Son 7 günde aşırı antrenman uyarıları."""
        with self.get_connection() as conn:
            haftalik = conn.execute(
                """SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=?
                   AND durum='Tamamlandi' AND tarih >= date('now','-7 days')""",
                (sporcu_id,)).fetchone()[0]
            art_arda = conn.execute(
                """SELECT COUNT(*) FROM (
                   SELECT tarih FROM gunluk_log WHERE sporcu_id=?
                   AND durum='Tamamlandi'
                   AND tarih >= date('now','-3 days') GROUP BY tarih)""",
                (sporcu_id,)).fetchone()[0]
            ort_zorluk = conn.execute(
                """SELECT COALESCE(AVG(zorluk_hissi),0) FROM gunluk_log
                   WHERE sporcu_id=? AND durum='Tamamlandi'
                   AND tarih >= date('now','-7 days')""",
                (sporcu_id,)).fetchone()[0]
            uyarilar = []
            if haftalik >= 6:
                uyarilar.append({
                    'seviye': 'Kritik', 'ikon': '🚨',
                    'mesaj': f'Haftada {haftalik} antrenman — overtraining riski yüksek!',
                    'oneri': 'En az 1-2 dinlenme günü ekle.'
                })
            elif haftalik >= 5:
                uyarilar.append({
                    'seviye': 'Uyarı', 'ikon': '⚠️',
                    'mesaj': f'Haftada {haftalik} antrenman — dinlenme günü önerilir.',
                    'oneri': 'Bu hafta 1 aktif dinlenme günü planla.'
                })
            if art_arda >= 3:
                uyarilar.append({
                    'seviye': 'Uyarı', 'ikon': '⚠️',
                    'mesaj': f'Son 3 gün üst üste antrenman.',
                    'oneri': 'Yarın dinlenme veya hafif stretching yap.'
                })
            if ort_zorluk >= 8:
                uyarilar.append({
                    'seviye': 'Bilgi', 'ikon': '💡',
                    'mesaj': f'Ort. zorluk {ort_zorluk:.1f}/10 — çok yoğun.',
                    'oneri': 'Bir antrenmanı daha hafif planla.'
                })
            if not uyarilar:
                uyarilar.append({
                    'seviye': 'İyi', 'ikon': '✅',
                    'mesaj': 'Antrenman yoğunluğu uygun seviyede.',
                    'oneri': 'Devam et!'
                })
            return {'haftalik': haftalik, 'art_arda': art_arda,
                    'ort_zorluk': round(ort_zorluk, 1), 'uyarilar': uyarilar}

    def get_karsilastirmali_analiz(self, sporcu_id1, sporcu_id2):
        """İki sporcu arasında karşılaştırmalı istatistikler."""
        def stats(sid):
            with self.get_connection() as conn:
                ant = conn.execute(
                    "SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                    (sid,)).fetchone()[0]
                kalori = conn.execute(
                    "SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                    (sid,)).fetchone()[0]
                sure = conn.execute(
                    "SELECT COALESCE(SUM(sure_dakika),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                    (sid,)).fetchone()[0]
                streak = self.get_streak(sid)
                rm = conn.execute(
                    "SELECT COALESCE(MAX(hesaplanan_1rm),0) FROM bir_rm_kayitlari WHERE sporcu_id=?",
                    (sid,)).fetchone()[0]
                puan = conn.execute(
                    "SELECT COALESCE(toplam_puan,0) FROM sporcu_puanlari WHERE sporcu_id=?",
                    (sid,)).fetchone()
                rozet = conn.execute(
                    "SELECT COUNT(*) FROM sporcu_rozetleri WHERE sporcu_id=?",
                    (sid,)).fetchone()[0]
                sporcu = conn.execute(
                    "SELECT ad, soyad, hedef FROM sporcular WHERE sporcu_id=?",
                    (sid,)).fetchone()
                return {
                    'ad': f"{sporcu[0]} {sporcu[1]}" if sporcu else '—',
                    'hedef': sporcu[2] if sporcu else '—',
                    'antrenman': ant, 'kalori': int(kalori),
                    'sure': int(sure), 'streak': streak,
                    'max_1rm': rm, 'puan': puan[0] if puan else 0,
                    'rozet': rozet,
                }
        return stats(sporcu_id1), stats(sporcu_id2)

    def get_aylik_rapor(self, sporcu_id, ay_offset=0):
        """Belirli ay için özet rapor."""
        with self.get_connection() as conn:
            ay_baslangic = f"date('now','start of month','{'-' if ay_offset else ''}{abs(ay_offset)} months')"
            ay_bitis     = f"date('now','start of month','{1-abs(ay_offset)} months','-1 day')"
            ant = conn.execute(
                f"SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi' AND tarih BETWEEN {ay_baslangic} AND {ay_bitis}",
                (sporcu_id,)).fetchone()[0]
            kalori = conn.execute(
                f"SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi' AND tarih BETWEEN {ay_baslangic} AND {ay_bitis}",
                (sporcu_id,)).fetchone()[0]
            sure = conn.execute(
                f"SELECT COALESCE(SUM(sure_dakika),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi' AND tarih BETWEEN {ay_baslangic} AND {ay_bitis}",
                (sporcu_id,)).fetchone()[0]
            kilo_rows = conn.execute(
                f"SELECT kilo FROM vucut_olcumleri WHERE sporcu_id=? AND tarih BETWEEN {ay_baslangic} AND {ay_bitis} ORDER BY tarih",
                (sporcu_id,)).fetchall()
            kilo_fark = 0
            if len(kilo_rows) >= 2:
                kilo_fark = round(kilo_rows[-1][0] - kilo_rows[0][0], 1)
            return {
                'antrenman': ant, 'kalori': int(kalori),
                'sure': int(sure), 'kilo_fark': kilo_fark,
            }

    def get_analiz_notlari(self, sporcu_id=None):
        with self.get_connection() as conn:
            q = """SELECT n.*, s.ad||' '||s.soyad as sporcu_adi
                   FROM analiz_notlari n
                   JOIN sporcular s ON n.sporcu_id=s.sporcu_id WHERE 1=1"""
            p = []
            if sporcu_id:
                q += " AND n.sporcu_id=?"; p.append(sporcu_id)
            q += " ORDER BY n.olusturma_tarihi DESC"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def add_analiz_notu(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO analiz_notlari
                (sporcu_id,not_tipi,baslik,icerik,oncelik,tarih)
                VALUES (?,?,?,?,?,?)""",
                (data['sporcu_id'], data.get('not_tipi','Gözlem'),
                 data['baslik'], data.get('icerik',''),
                 data.get('oncelik','Normal'),
                 data.get('tarih', datetime.now().strftime('%Y-%m-%d'))))

    def delete_analiz_notu(self, not_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM analiz_notlari WHERE not_id=?", (not_id,))

    # ── TİER 7: 1RM & RADAR DB METODLARI ───────────────────────
    # Epley: 1RM = w * (1 + r/30)
    # Brzycki: 1RM = w * (36 / (37 - r))
    # Lander: 1RM = (100 * w) / (101.3 - 2.67123 * r)

    def hesapla_1rm(self, agirlik, tekrar, formul='Epley'):
        if tekrar == 1: return agirlik
        if formul == 'Epley':
            return round(agirlik * (1 + tekrar / 30), 1)
        elif formul == 'Brzycki':
            return round(agirlik * 36 / (37 - tekrar), 1)
        elif formul == 'Lander':
            return round((100 * agirlik) / (101.3 - 2.67123 * tekrar), 1)
        return round(agirlik * (1 + tekrar / 30), 1)

    def add_1rm_kaydi(self, sporcu_id, egzersiz, agirlik, tekrar, formul='Epley'):
        bir_rm = self.hesapla_1rm(agirlik, tekrar, formul)
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO bir_rm_kayitlari
                (sporcu_id,egzersiz_adi,agirlik,tekrar,hesaplanan_1rm,formul)
                VALUES (?,?,?,?,?,?)""",
                (sporcu_id, egzersiz, agirlik, tekrar, bir_rm, formul))
        return bir_rm

    def get_1rm_kayitlari(self, sporcu_id=None, egzersiz=None):
        with self.get_connection() as conn:
            q = """SELECT r.*, s.ad||' '||s.soyad as sporcu_adi
                   FROM bir_rm_kayitlari r
                   JOIN sporcular s ON r.sporcu_id=s.sporcu_id WHERE 1=1"""
            p = []
            if sporcu_id:
                q += " AND r.sporcu_id=?"; p.append(sporcu_id)
            if egzersiz and egzersiz != 'Tümü':
                q += " AND r.egzersiz_adi=?"; p.append(egzersiz)
            q += " ORDER BY r.tarih DESC, r.hesaplanan_1rm DESC"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def get_1rm_en_iyi(self, sporcu_id):
        """Her egzersiz için en iyi 1RM."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT egzersiz_adi, MAX(hesaplanan_1rm) as max_1rm,
                       agirlik, tekrar, tarih
                FROM bir_rm_kayitlari WHERE sporcu_id=?
                GROUP BY egzersiz_adi ORDER BY max_1rm DESC""",
                (sporcu_id,)).fetchall()
            return [dict(r) for r in rows]

    def get_1rm_trend(self, sporcu_id, egzersiz):
        """Belirli egzersiz için 1RM zaman trendi."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT tarih, MAX(hesaplanan_1rm) as max_1rm
                FROM bir_rm_kayitlari
                WHERE sporcu_id=? AND egzersiz_adi=?
                GROUP BY tarih ORDER BY tarih""",
                (sporcu_id, egzersiz)).fetchall()
            return [(r[0][-5:], r[1]) for r in rows]

    def get_radar_degerleri(self, sporcu_id):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM radar_degerleri WHERE sporcu_id=?",
                (sporcu_id,)).fetchone()
            return dict(row) if row else None

    def hesapla_radar_otomatik(self, sporcu_id):
        """Mevcut verilerden radar değerlerini otomatik hesapla."""
        with self.get_connection() as conn:
            # Güç: en iyi 1RM bazlı
            max_rm = conn.execute(
                "SELECT MAX(hesaplanan_1rm) FROM bir_rm_kayitlari WHERE sporcu_id=?",
                (sporcu_id,)).fetchone()[0] or 0
            guc = min(100, max_rm / 2)

            # Kardiyo: kardiyo antrenman sayısı
            kardiyo = conn.execute("""
                SELECT COUNT(*) FROM gunluk_log g
                JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.sporcu_id=? AND a.kategori='Kardiyo'
                AND g.durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            kardiyo_puan = min(100, kardiyo * 5)

            # Esneklik: esneklik antrenman sayısı
            esneklik = conn.execute("""
                SELECT COUNT(*) FROM gunluk_log g
                JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.sporcu_id=? AND a.kategori IN ('Esneklik','Denge')
                AND g.durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            esneklik_puan = min(100, esneklik * 8)

            # Dayanıklılık: toplam süre bazlı
            toplam_sure = conn.execute(
                "SELECT COALESCE(SUM(sure_dakika),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            dayaniklilik = min(100, toplam_sure / 30)

            # Hız: HIIT antrenman sayısı
            hiit = conn.execute("""
                SELECT COUNT(*) FROM gunluk_log g
                JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.sporcu_id=? AND a.kategori='HIIT'
                AND g.durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            hiz_puan = min(100, hiit * 10)

            # Denge: streak bazlı
            streak = self.get_streak(sporcu_id)
            denge_puan = min(100, streak * 5)

        vals = {
            'guc_puani': round(guc, 1),
            'kardiyo_puani': round(kardiyo_puan, 1),
            'esneklik_puani': round(esneklik_puan, 1),
            'denge_puani': round(denge_puan, 1),
            'dayaniklilik_puani': round(dayaniklilik, 1),
            'hiz_puani': round(hiz_puan, 1),
        }
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO radar_degerleri
                (sporcu_id,guc_puani,kardiyo_puani,esneklik_puani,
                 denge_puani,dayaniklilik_puani,hiz_puani)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(sporcu_id) DO UPDATE SET
                guc_puani=excluded.guc_puani,
                kardiyo_puani=excluded.kardiyo_puani,
                esneklik_puani=excluded.esneklik_puani,
                denge_puani=excluded.denge_puani,
                dayaniklilik_puani=excluded.dayaniklilik_puani,
                hiz_puani=excluded.hiz_puani,
                hesaplama_tarihi=datetime('now','localtime')""",
                (sporcu_id, vals['guc_puani'], vals['kardiyo_puani'],
                 vals['esneklik_puani'], vals['denge_puani'],
                 vals['dayaniklilik_puani'], vals['hiz_puani']))
        return vals

    # ── TİER 6: GAMİFİCATİON DB METODLARI ──────────────────────
    SEVIYE_ESLEME = [
        (0,    'Başlangıç', 1, '#6b7280'),
        (100,  'Bronz',     2, '#92400e'),
        (300,  'Gümüş',     3, '#6b7280'),
        (600,  'Altın',     4, '#f59e0b'),
        (1000, 'Platin',    5, '#818cf8'),
        (2000, 'Elmas',     6, '#06b6d4'),
    ]

    def get_rozetler(self, kategori='Tümü'):
        with self.get_connection() as conn:
            q = "SELECT * FROM rozetler"
            if kategori != 'Tümü':
                q += " WHERE kategori=?"
                return [dict(r) for r in conn.execute(q, (kategori,)).fetchall()]
            return [dict(r) for r in conn.execute(q).fetchall()]

    def get_sporcu_rozetleri(self, sporcu_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT sr.*, r.rozet_adi, r.aciklama, r.icon, r.kategori, r.puan
                FROM sporcu_rozetleri sr
                JOIN rozetler r ON sr.rozet_id=r.rozet_id
                WHERE sr.sporcu_id=?
                ORDER BY sr.kazanma_tarihi DESC""", (sporcu_id,)).fetchall()]

    def get_sporcu_puan(self, sporcu_id):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM sporcu_puanlari WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            return dict(row) if row else {'toplam_puan': 0, 'seviye': 'Başlangıç', 'seviye_no': 1}

    def _hesapla_seviye(self, puan):
        seviye_adi = 'Başlangıç'; seviye_no = 1; renk = '#6b7280'
        for esik, sev_adi, sev_no, sev_renk in self.SEVIYE_ESLEME:
            if puan >= esik:
                seviye_adi = sev_adi; seviye_no = sev_no; renk = sev_renk
        return seviye_adi, seviye_no, renk

    def puan_ekle(self, sporcu_id, puan, sebep=''):
        with self.get_connection() as conn:
            mevcut = conn.execute(
                "SELECT toplam_puan FROM sporcu_puanlari WHERE sporcu_id=?",
                (sporcu_id,)).fetchone()
            if mevcut:
                yeni_puan = mevcut[0] + puan
                sev_adi, sev_no, _ = self._hesapla_seviye(yeni_puan)
                conn.execute("""UPDATE sporcu_puanlari SET
                    toplam_puan=?, seviye=?, seviye_no=?,
                    guncelleme_tarihi=datetime('now','localtime')
                    WHERE sporcu_id=?""", (yeni_puan, sev_adi, sev_no, sporcu_id))
            else:
                sev_adi, sev_no, _ = self._hesapla_seviye(puan)
                conn.execute("""INSERT INTO sporcu_puanlari
                    (sporcu_id,toplam_puan,seviye,seviye_no) VALUES (?,?,?,?)""",
                    (sporcu_id, puan, sev_adi, sev_no))

    def rozet_ver(self, sporcu_id, rozet_id):
        """Sporcu zaten sahip değilse rozet ver ve puanı ekle."""
        with self.get_connection() as conn:
            mevcut = conn.execute(
                "SELECT id FROM sporcu_rozetleri WHERE sporcu_id=? AND rozet_id=?",
                (sporcu_id, rozet_id)).fetchone()
            if mevcut: return False
            rozet = conn.execute("SELECT puan FROM rozetler WHERE rozet_id=?",
                                 (rozet_id,)).fetchone()
            conn.execute("""INSERT INTO sporcu_rozetleri (sporcu_id,rozet_id)
                VALUES (?,?)""", (sporcu_id, rozet_id))
        if rozet:
            self.puan_ekle(sporcu_id, rozet[0])
        return True

    def rozet_kontrol_ve_ver(self, sporcu_id):
        """Sporcu koşullarını kontrol et, hak ettiği rozetleri ver."""
        with self.get_connection() as conn:
            rozetler = conn.execute("SELECT * FROM rozetler").fetchall()
            ant_sayisi = conn.execute(
                "SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            toplam_kalori = conn.execute(
                "SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            toplam_sure = conn.execute(
                "SELECT COALESCE(SUM(sure_dakika),0) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            haftalik = conn.execute(
                """SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'
                   AND tarih >= date('now','-7 days')""", (sporcu_id,)).fetchone()[0]
            tamamlanan_hedef = conn.execute(
                "SELECT COUNT(*) FROM hedefler WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            kategori_sayisi = conn.execute(
                """SELECT COUNT(DISTINCT a.kategori) FROM gunluk_log g
                   JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                   WHERE g.sporcu_id=? AND g.durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            beslenme_streak = conn.execute(
                """SELECT COUNT(*) FROM (
                   SELECT tarih FROM gunluk_beslenme WHERE sporcu_id=?
                   AND tarih >= date('now','-7 days') GROUP BY tarih)""",
                (sporcu_id,)).fetchone()[0]

        streak = self.get_streak(sporcu_id)
        degerler = {
            'antrenman_sayisi': ant_sayisi,
            'streak': streak,
            'toplam_kalori': toplam_kalori,
            'toplam_sure_saat': toplam_sure / 60,
            'haftalik_antrenman': haftalik,
            'tamamlanan_hedef': tamamlanan_hedef,
            'kategori_sayisi': kategori_sayisi,
            'beslenme_streak': beslenme_streak,
        }
        yeni_rozetler = []
        for r in rozetler:
            deger = degerler.get(r['kosul_tipi'], 0)
            if deger >= r['kosul_deger']:
                if self.rozet_ver(sporcu_id, r['rozet_id']):
                    yeni_rozetler.append(dict(r))
        return yeni_rozetler

    def get_leaderboard(self):
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT sp.sporcu_id, s.ad, s.soyad, sp.toplam_puan, sp.seviye, sp.seviye_no,
                       COUNT(sr.id) as rozet_sayisi
                FROM sporcu_puanlari sp
                JOIN sporcular s ON sp.sporcu_id=s.sporcu_id
                LEFT JOIN sporcu_rozetleri sr ON sp.sporcu_id=sr.sporcu_id
                WHERE s.durum='Aktif'
                GROUP BY sp.sporcu_id
                ORDER BY sp.toplam_puan DESC""").fetchall()
            return [dict(r) for r in rows]

    def get_tum_sporcu_rozetleri_ozet(self):
        """Rozet sayısı ve toplam puan özeti."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT s.sporcu_id, s.ad||' '||s.soyad as ad,
                       COUNT(sr.id) as rozet_sayisi,
                       COALESCE(SUM(r.puan),0) as rozet_puani
                FROM sporcular s
                LEFT JOIN sporcu_rozetleri sr ON s.sporcu_id=sr.sporcu_id
                LEFT JOIN rozetler r ON sr.rozet_id=r.rozet_id
                WHERE s.durum='Aktif'
                GROUP BY s.sporcu_id ORDER BY rozet_sayisi DESC""").fetchall()
            return [dict(r) for r in rows]

    # ── BESİNLER ──────────────────────────────────────────────────
    def get_besinler(self, search='', kategori='Tümü'):
        with self.get_connection() as conn:
            q = "SELECT * FROM besinler WHERE durum='Aktif'"
            p = []
            if kategori != 'Tümü':
                q += " AND kategori=?"; p.append(kategori)
            if search:
                q += " AND (ad LIKE ? OR kategori LIKE ?)"; p += [f'%{search}%']*2
            return [dict(r) for r in conn.execute(q + " ORDER BY ad", p).fetchall()]

    def get_besin(self, besin_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM besinler WHERE besin_id=?", (besin_id,)).fetchone()
            return dict(row) if row else None

    def add_besin(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO besinler
                (ad,kategori,porsiyon_gram,kalori,protein,karbonhidrat,yag,lif)
                VALUES (?,?,?,?,?,?,?,?)""",
                (data['ad'], data.get('kategori','Genel'), data.get('porsiyon_gram',100),
                 data.get('kalori',0), data.get('protein',0), data.get('karbonhidrat',0),
                 data.get('yag',0), data.get('lif',0)))

    def delete_besin(self, besin_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE besinler SET durum='Pasif' WHERE besin_id=?", (besin_id,))

    # ── GÜNLÜK BESLENME ───────────────────────────────────────────
    def get_gunluk_beslenme(self, sporcu_id, tarih=None):
        tarih = tarih or datetime.now().strftime('%Y-%m-%d')
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT gb.*, b.ad as besin_adi, b.kategori
                FROM gunluk_beslenme gb
                JOIN besinler b ON gb.besin_id=b.besin_id
                WHERE gb.sporcu_id=? AND gb.tarih=?
                ORDER BY gb.ogun, gb.olusturma_tarihi""",
                (sporcu_id, tarih)).fetchall()]

    def get_beslenme_tarihleri(self, sporcu_id, limit=14):
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT tarih, SUM(kalori_toplam) as toplam_kalori,
                       SUM(protein_toplam) as toplam_protein,
                       SUM(karb_toplam) as toplam_karb,
                       SUM(yag_toplam) as toplam_yag
                FROM gunluk_beslenme WHERE sporcu_id=?
                GROUP BY tarih ORDER BY tarih DESC LIMIT ?""",
                (sporcu_id, limit)).fetchall()
            return [dict(r) for r in rows]

    def add_beslenme_kaydi(self, data):
        besin = self.get_besin(data['besin_id'])
        if not besin: return
        oran = data.get('miktar_gram', 100) / (besin['porsiyon_gram'] or 100)
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO gunluk_beslenme
                (sporcu_id,besin_id,tarih,ogun,miktar_gram,
                 kalori_toplam,protein_toplam,karb_toplam,yag_toplam,notlar)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (data['sporcu_id'], data['besin_id'],
                 data.get('tarih', datetime.now().strftime('%Y-%m-%d')),
                 data.get('ogun','Kahvalti'), data.get('miktar_gram',100),
                 round(besin['kalori'] * oran, 1),
                 round(besin['protein'] * oran, 1),
                 round(besin['karbonhidrat'] * oran, 1),
                 round(besin['yag'] * oran, 1),
                 data.get('notlar','')))

    def delete_beslenme_kaydi(self, beslenme_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM gunluk_beslenme WHERE beslenme_id=?", (beslenme_id,))

    def get_gunluk_makro_ozet(self, sporcu_id, tarih=None):
        tarih = tarih or datetime.now().strftime('%Y-%m-%d')
        with self.get_connection() as conn:
            row = conn.execute("""
                SELECT COALESCE(SUM(kalori_toplam),0) as kalori,
                       COALESCE(SUM(protein_toplam),0) as protein,
                       COALESCE(SUM(karb_toplam),0) as karb,
                       COALESCE(SUM(yag_toplam),0) as yag
                FROM gunluk_beslenme WHERE sporcu_id=? AND tarih=?""",
                (sporcu_id, tarih)).fetchone()
            hedef = conn.execute(
                "SELECT * FROM kalori_hedefleri WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            return {
                'kalori': round(row[0],1), 'protein': round(row[1],1),
                'karb': round(row[2],1), 'yag': round(row[3],1),
                'hedef': dict(hedef) if hedef else {'gunluk_kalori':2000,'protein_gram':150,'karb_gram':200,'yag_gram':65}
            }

    def get_set_kalori_hedefi(self, sporcu_id):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM kalori_hedefleri WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            return dict(row) if row else None

    def set_kalori_hedefi(self, sporcu_id, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO kalori_hedefleri
                (sporcu_id,gunluk_kalori,protein_gram,karb_gram,yag_gram)
                VALUES (?,?,?,?,?)
                ON CONFLICT(sporcu_id) DO UPDATE SET
                gunluk_kalori=excluded.gunluk_kalori,
                protein_gram=excluded.protein_gram,
                karb_gram=excluded.karb_gram,
                yag_gram=excluded.yag_gram,
                guncelleme_tarihi=datetime('now','localtime')""",
                (sporcu_id, data['gunluk_kalori'], data['protein_gram'],
                 data['karb_gram'], data['yag_gram']))

    def get_haftalik_kalori(self, sporcu_id):
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT tarih, SUM(kalori_toplam) as toplam
                FROM gunluk_beslenme WHERE sporcu_id=?
                AND tarih >= date('now','-7 days')
                GROUP BY tarih ORDER BY tarih""", (sporcu_id,)).fetchall()
            return [(r[0][-5:], round(r[1],1)) for r in rows]


    # ── TİER 5: İLERLEME ANALİZ & HAFTALIK ÖZET ─────────────────
    def get_ilerleme_grafik_data(self, sporcu_id, gun=30):
        """Son N gün antrenman + kalori + kilo verisi — grafik için."""
        with self.get_connection() as conn:
            # Antrenman log (günlük)
            ant_rows = conn.execute("""
                SELECT tarih, COUNT(*) as sayi, SUM(sure_dakika) as sure,
                       SUM(kalori_yakilan) as kalori
                FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'
                AND tarih >= date('now',?)
                GROUP BY tarih ORDER BY tarih""",
                (sporcu_id, f'-{gun} days')).fetchall()
            antrenman_data = [(r[0][-5:], r[1], r[2] or 0, r[3] or 0) for r in ant_rows]

            # Kilo trendi (vücut ölçümleri)
            kilo_rows = conn.execute("""
                SELECT tarih, kilo FROM vucut_olcumleri
                WHERE sporcu_id=? AND kilo IS NOT NULL
                AND tarih >= date('now',?)
                ORDER BY tarih""",
                (sporcu_id, f'-{gun} days')).fetchall()
            kilo_data = [(r[0][-5:], r[1]) for r in kilo_rows]

            # Kalori alımı (beslenme)
            kal_rows = conn.execute("""
                SELECT tarih, SUM(kalori_toplam) as toplam
                FROM gunluk_beslenme WHERE sporcu_id=?
                AND tarih >= date('now',?)
                GROUP BY tarih ORDER BY tarih""",
                (sporcu_id, f'-{gun} days')).fetchall()
            kalori_data = [(r[0][-5:], round(r[1],1)) for r in kal_rows]

            return {
                'antrenman': antrenman_data,
                'kilo':      kilo_data,
                'kalori':    kalori_data,
            }

    def get_haftalik_ozet(self, sporcu_id):
        """Son 7 günün tam özeti."""
        with self.get_connection() as conn:
            ant = conn.execute("""
                SELECT COUNT(*) as sayi, COALESCE(SUM(sure_dakika),0) as sure,
                       COALESCE(SUM(kalori_yakilan),0) as kalori,
                       COALESCE(AVG(zorluk_hissi),0) as ort_zorluk
                FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'
                AND tarih >= date('now','-7 days')""", (sporcu_id,)).fetchone()

            bes = conn.execute("""
                SELECT COALESCE(AVG(gun_kalori),0) as ort_kalori
                FROM (SELECT tarih, SUM(kalori_toplam) as gun_kalori
                      FROM gunluk_beslenme WHERE sporcu_id=?
                      AND tarih >= date('now','-7 days') GROUP BY tarih)""",
                (sporcu_id,)).fetchone()

            streak = self.get_streak(sporcu_id)

            # Kategorilere göre antrenman dağılımı bu hafta
            kat_rows = conn.execute("""
                SELECT a.kategori, COUNT(*) as sayi
                FROM gunluk_log g JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.sporcu_id=? AND g.durum='Tamamlandi'
                AND g.tarih >= date('now','-7 days')
                GROUP BY a.kategori ORDER BY sayi DESC""", (sporcu_id,)).fetchall()

            # En çok çalışılan kas grubu / kategori
            en_cok = kat_rows[0][0] if kat_rows else '—'

            sporcu = conn.execute(
                "SELECT * FROM sporcular WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            hedef_row = conn.execute(
                "SELECT * FROM kalori_hedefleri WHERE sporcu_id=?", (sporcu_id,)).fetchone()

            return {
                'antrenman_sayisi':  ant[0],
                'toplam_sure':       int(ant[1]),
                'toplam_kalori_yak': int(ant[2]),
                'ort_zorluk':        round(ant[3], 1),
                'ort_gunluk_kalori': int(bes[0]),
                'streak':            streak,
                'en_cok_kategori':   en_cok,
                'kategori_dagilimi': [(r[0], r[1]) for r in kat_rows],
                'sporcu':            dict(sporcu) if sporcu else {},
                'kalori_hedef':      dict(hedef_row) if hedef_row else {},
            }

    def get_kilo_degisim(self, sporcu_id):
        """İlk ve son ölçüm karşılaştırması."""
        with self.get_connection() as conn:
            ilk = conn.execute(
                "SELECT kilo, tarih FROM vucut_olcumleri WHERE sporcu_id=? ORDER BY tarih ASC LIMIT 1",
                (sporcu_id,)).fetchone()
            son = conn.execute(
                "SELECT kilo, tarih FROM vucut_olcumleri WHERE sporcu_id=? ORDER BY tarih DESC LIMIT 1",
                (sporcu_id,)).fetchone()
            if not ilk or not son:
                return None
            return {
                'ilk_kilo': ilk[0], 'ilk_tarih': ilk[1],
                'son_kilo': son[0], 'son_tarih': son[1],
                'fark': round(son[0] - ilk[0], 1),
            }

    def get_performans_trend(self, sporcu_id, hafta=4):
        """Son N haftanın antrenman yoğunluk trendi."""
        with self.get_connection() as conn:
            rows = conn.execute("""
                SELECT strftime('%W', tarih) as hafta_no,
                       COUNT(*) as sayi, AVG(zorluk_hissi) as ort_zorluk,
                       SUM(kalori_yakilan) as toplam_kalori
                FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'
                AND tarih >= date('now',?)
                GROUP BY hafta_no ORDER BY hafta_no""",
                (sporcu_id, f'-{hafta*7} days')).fetchall()
            return [{'hafta': f"H{i+1}", 'sayi': r[1],
                     'zorluk': round(r[2] or 0, 1), 'kalori': int(r[3] or 0)}
                    for i, r in enumerate(rows)]

    # ── VÜCUT ÖLÇÜMLERİ ──────────────────────────────────────────
    def get_vucut_olcumleri(self, sporcu_id, limit=20):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM vucut_olcumleri WHERE sporcu_id=? ORDER BY tarih DESC LIMIT ?",
                (sporcu_id, limit)).fetchall()]

    def add_vucut_olcumu(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO vucut_olcumleri
                (sporcu_id,tarih,kilo,bel,gogus,kalca,sag_kol,sol_kol,
                 sag_bacak,sol_bacak,vucut_yag_orani,notlar)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (data['sporcu_id'], data.get('tarih', datetime.now().strftime('%Y-%m-%d')),
                 data.get('kilo'), data.get('bel'), data.get('gogus'), data.get('kalca'),
                 data.get('sag_kol'), data.get('sol_kol'), data.get('sag_bacak'),
                 data.get('sol_bacak'), data.get('vucut_yag_orani'), data.get('notlar','')))

    def get_vucut_trend(self, sporcu_id, alan='kilo'):
        with self.get_connection() as conn:
            rows = conn.execute(
                f"SELECT tarih, {alan} FROM vucut_olcumleri WHERE sporcu_id=? AND {alan} IS NOT NULL ORDER BY tarih",
                (sporcu_id,)).fetchall()
            return [(r[0], r[1]) for r in rows]

    # ── HEDEFLER ──────────────────────────────────────────────────
    def get_hedefler(self, sporcu_id=None, durum='Tümü'):
        with self.get_connection() as conn:
            q = """SELECT h.*, s.ad||' '||s.soyad as sporcu_adi
                   FROM hedefler h JOIN sporcular s ON h.sporcu_id=s.sporcu_id WHERE 1=1"""
            p = []
            if sporcu_id:
                q += " AND h.sporcu_id=?"; p.append(sporcu_id)
            if durum != 'Tümü':
                q += " AND h.durum=?"; p.append(durum)
            q += " ORDER BY h.olusturma_tarihi DESC"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def get_hedef(self, hedef_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM hedefler WHERE hedef_id=?", (hedef_id,)).fetchone()
            return dict(row) if row else None

    def add_hedef(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO hedefler
                (sporcu_id,hedef_tipi,hedef_adi,baslangic_degeri,hedef_degeri,
                 mevcut_deger,baslangic_tarihi,hedef_tarihi,notlar)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (data['sporcu_id'], data['hedef_tipi'], data['hedef_adi'],
                 data.get('baslangic_degeri',0), data['hedef_degeri'],
                 data.get('mevcut_deger',0),
                 data.get('baslangic_tarihi', datetime.now().strftime('%Y-%m-%d')),
                 data.get('hedef_tarihi'), data.get('notlar','')))

    def update_hedef_ilerleme(self, hedef_id, mevcut_deger):
        with self.get_connection() as conn:
            h = conn.execute("SELECT * FROM hedefler WHERE hedef_id=?", (hedef_id,)).fetchone()
            if not h: return
            durum = 'Tamamlandi' if mevcut_deger >= h['hedef_degeri'] else 'Devam'
            conn.execute("UPDATE hedefler SET mevcut_deger=?, durum=? WHERE hedef_id=?",
                         (mevcut_deger, durum, hedef_id))

    def delete_hedef(self, hedef_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM hedefler WHERE hedef_id=?", (hedef_id,))

    # ── STREAK ────────────────────────────────────────────────────
    def add_streak(self, sporcu_id, tarih=None):
        tarih = tarih or datetime.now().strftime('%Y-%m-%d')
        with self.get_connection() as conn:
            try:
                conn.execute(
                    "INSERT INTO streak_kayitlari (sporcu_id,tarih,antrenman_yapildi) VALUES (?,?,1)",
                    (sporcu_id, tarih))
                return True
            except: return False

    def get_streak(self, sporcu_id):
        """Mevcut streak sayısını hesapla — ardışık gün sayısı."""
        with self.get_connection() as conn:
            rows = conn.execute(
                """SELECT tarih FROM streak_kayitlari WHERE sporcu_id=? AND antrenman_yapildi=1
                   ORDER BY tarih DESC""", (sporcu_id,)).fetchall()
            if not rows: return 0
            streak = 0
            bugun = datetime.now().date()
            for i, row in enumerate(rows):
                beklenen = bugun - timedelta(days=i)
                kayit = datetime.strptime(row[0], '%Y-%m-%d').date()
                if kayit == beklenen:
                    streak += 1
                else:
                    break
            return streak

    def get_tum_streakler(self):
        """Tüm sporcuların streak sayıları."""
        with self.get_connection() as conn:
            sporcular = conn.execute(
                "SELECT sporcu_id, ad, soyad FROM sporcular WHERE durum='Aktif'").fetchall()
            result = []
            for s in sporcular:
                streak = self.get_streak(s[0])
                result.append({'sporcu_id': s[0], 'ad': s[1], 'soyad': s[2], 'streak': streak})
            return sorted(result, key=lambda x: x['streak'], reverse=True)

    # ── ANTRENMAN PROGRAMLARI ────────────────────────────────────
    def get_programlar(self, sporcu_id=None):
        with self.get_connection() as conn:
            q = """SELECT p.*, s.ad||' '||s.soyad as sporcu_adi
                   FROM antrenman_programlari p
                   JOIN sporcular s ON p.sporcu_id=s.sporcu_id WHERE 1=1"""
            params = []
            if sporcu_id:
                q += " AND p.sporcu_id=?"; params.append(sporcu_id)
            q += " ORDER BY p.olusturma_tarihi DESC"
            return [dict(r) for r in conn.execute(q, params).fetchall()]

    def get_program(self, program_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM antrenman_programlari WHERE program_id=?",
                               (program_id,)).fetchone()
            return dict(row) if row else None

    def add_program(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO antrenman_programlari
                (sporcu_id,program_adi,hedef,sure_hafta,notlar)
                VALUES (?,?,?,?,?)""",
                (data['sporcu_id'], data['program_adi'], data.get('hedef','Genel'),
                 data.get('sure_hafta',4), data.get('notlar','')))

    def delete_program(self, program_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM program_gunleri WHERE program_id=?", (program_id,))
            conn.execute("DELETE FROM antrenman_programlari WHERE program_id=?", (program_id,))

    def get_program_gunleri(self, program_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute("""
                SELECT pg.*, a.ad as antrenman_adi, a.kategori
                FROM program_gunleri pg
                JOIN antrenmanlar a ON pg.antrenman_id=a.antrenman_id
                WHERE pg.program_id=? ORDER BY pg.gun_id""", (program_id,)).fetchall()]

    def add_program_gunu(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO program_gunleri
                (program_id,gun_adi,antrenman_id,set_sayisi,tekrar,agirlik,dinlenme_saniye,notlar)
                VALUES (?,?,?,?,?,?,?,?)""",
                (data['program_id'], data['gun_adi'], data['antrenman_id'],
                 data.get('set_sayisi',3), data.get('tekrar',12),
                 data.get('agirlik',0), data.get('dinlenme_saniye',60), data.get('notlar','')))

    def delete_program_gunu(self, gun_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM program_gunleri WHERE gun_id=?", (gun_id,))

    # ── GÜNLÜK LOG ────────────────────────────────────────────────
    def get_gunluk_log(self, sporcu_id=None, search='', limit=100):
        with self.get_connection() as conn:
            q = """SELECT g.*, s.ad||' '||s.soyad as sporcu_adi,
                   a.ad as antrenman_adi, a.kategori
                   FROM gunluk_log g
                   JOIN sporcular s ON g.sporcu_id=s.sporcu_id
                   JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                   WHERE g.durum='Tamamlandi'"""
            p = []
            if sporcu_id:
                q += " AND g.sporcu_id=?"; p.append(sporcu_id)
            if search:
                q += " AND (s.ad LIKE ? OR a.ad LIKE ?)"; p += [f'%{search}%']*2
            q += " ORDER BY g.tarih DESC, g.olusturma_tarihi DESC"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def add_gunluk_log(self, data):
        with self.get_connection() as conn:
            log_id = conn.execute("""INSERT INTO gunluk_log
                (sporcu_id,antrenman_id,tarih,sure_dakika,kalori_yakilan,
                 max_nabiz,ort_nabiz,zorluk_hissi,yorgunluk,notlar,durum)
                VALUES (?,?,?,?,?,?,?,?,?,?,'Tamamlandi') RETURNING log_id""",
                (data['sporcu_id'], data['antrenman_id'], data['tarih'],
                 data.get('sure_dakika',0), data.get('kalori_yakilan',0),
                 data.get('max_nabiz',0), data.get('ort_nabiz',0),
                 data.get('zorluk_hissi',5), data.get('yorgunluk',5),
                 data.get('notlar',''))).fetchone()[0]
            # Set kayıtları
            for s in data.get('setler', []):
                conn.execute("""INSERT INTO set_kayitlari
                    (log_id,set_no,tekrar,agirlik,tamamlandi) VALUES (?,?,?,?,?)""",
                    (log_id, s['set_no'], s.get('tekrar',0), s.get('agirlik',0),
                     s.get('tamamlandi',1)))
            return log_id

    def delete_gunluk_log(self, log_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM set_kayitlari WHERE log_id=?", (log_id,))
            conn.execute("UPDATE gunluk_log SET durum='Silindi' WHERE log_id=?", (log_id,))

    def get_set_kayitlari(self, log_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM set_kayitlari WHERE log_id=? ORDER BY set_no",
                (log_id,)).fetchall()]

    def get_sporcu_log_stats(self, sporcu_id):
        with self.get_connection() as conn:
            toplam = conn.execute(
                "SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'",
                (sporcu_id,)).fetchone()[0]
            bu_hafta = conn.execute(
                """SELECT COUNT(*) FROM gunluk_log WHERE sporcu_id=? AND durum='Tamamlandi'
                   AND tarih >= date('now','-7 days')""", (sporcu_id,)).fetchone()[0]
            toplam_kalori = conn.execute(
                """SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log
                   WHERE sporcu_id=? AND durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            ort_sure = conn.execute(
                """SELECT COALESCE(AVG(sure_dakika),0) FROM gunluk_log
                   WHERE sporcu_id=? AND durum='Tamamlandi'""", (sporcu_id,)).fetchone()[0]
            return {'toplam': toplam, 'bu_hafta': bu_hafta,
                    'toplam_kalori': toplam_kalori, 'ort_sure': round(ort_sure, 1)}

    # ── AUTH ───────────────────────────────────────────────────────
    def authenticate(self, username, password):
        with self.get_connection() as conn:
            row = conn.execute(
                "SELECT * FROM kullanicilar WHERE kullanici_adi=? AND sifre=? AND durum='Aktif'",
                (username, hash_password(password))).fetchone()
            return dict(row) if row else None

    # ── SPORCULAR ──────────────────────────────────────────────────
    def get_sporcular(self, search='', durum='Tümü'):
        with self.get_connection() as conn:
            q = "SELECT * FROM sporcular WHERE durum!='Silindi'"
            p = []
            if durum != 'Tümü':
                q += " AND durum=?"; p.append(durum)
            if search:
                q += " AND (ad LIKE ? OR soyad LIKE ? OR hedef LIKE ?)"
                p += [f'%{search}%'] * 3
            return [dict(r) for r in conn.execute(q + ' ORDER BY kayit_tarihi DESC', p).fetchall()]

    def get_sporcu(self, sporcu_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM sporcular WHERE sporcu_id=?", (sporcu_id,)).fetchone()
            return dict(row) if row else None

    def add_sporcu(self, data):
        bmi = round(data['kilo'] / ((data['boy']/100) ** 2), 1)
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO sporcular
                (ad,soyad,yas,kilo,boy,cinsiyet,hedef,bmi,notlar)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (data['ad'], data['soyad'], data['yas'], data['kilo'], data['boy'],
                 data['cinsiyet'], data['hedef'], bmi, data.get('notlar', '')))

    def update_sporcu(self, sporcu_id, data):
        bmi = round(data['kilo'] / ((data['boy']/100) ** 2), 1)
        with self.get_connection() as conn:
            conn.execute("""UPDATE sporcular SET ad=?,soyad=?,yas=?,kilo=?,boy=?,
                cinsiyet=?,hedef=?,bmi=?,notlar=?,durum=? WHERE sporcu_id=?""",
                (data['ad'], data['soyad'], data['yas'], data['kilo'], data['boy'],
                 data['cinsiyet'], data['hedef'], bmi, data.get('notlar',''),
                 data.get('durum','Aktif'), sporcu_id))

    def delete_sporcu(self, sporcu_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE sporcular SET durum='Pasif' WHERE sporcu_id=?", (sporcu_id,))

    def get_sporcu_stats(self, sporcu_id):
        with self.get_connection() as conn:
            toplam_antrenman = conn.execute(
                "SELECT COUNT(*) FROM takipler WHERE sporcu_id=? AND durum='Aktif'",
                (sporcu_id,)).fetchone()[0]
            toplam_kalori = conn.execute(
                "SELECT COALESCE(SUM(kalori),0) FROM takipler WHERE sporcu_id=? AND durum='Aktif'",
                (sporcu_id,)).fetchone()[0]
            son_ilerleme = conn.execute(
                "SELECT kilo FROM ilerlemeler WHERE sporcu_id=? ORDER BY tarih DESC LIMIT 1",
                (sporcu_id,)).fetchone()
            return {
                'toplam_antrenman': toplam_antrenman,
                'toplam_kalori': toplam_kalori,
                'son_kilo': son_ilerleme[0] if son_ilerleme else None
            }

    # ── ANTRENMANLAR ───────────────────────────────────────────────
    def get_antrenmanlar(self, search='', kategori='Tümü'):
        with self.get_connection() as conn:
            q = "SELECT * FROM antrenmanlar WHERE durum='Aktif'"
            p = []
            if kategori != 'Tümü':
                q += " AND kategori=?"; p.append(kategori)
            if search:
                q += " AND (ad LIKE ? OR aciklama LIKE ?)"
                p += [f'%{search}%'] * 2
            return [dict(r) for r in conn.execute(q + ' ORDER BY antrenman_id', p).fetchall()]

    def get_antrenman(self, antrenman_id):
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM antrenmanlar WHERE antrenman_id=?",
                               (antrenman_id,)).fetchone()
            return dict(row) if row else None

    def add_antrenman(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO antrenmanlar
                (ad,kategori,saat,dakika,zorluk_seviyesi,aciklama,kalori_tahmini)
                VALUES (?,?,?,?,?,?,?)""",
                (data['ad'], data['kategori'], data['saat'], data['dakika'],
                 data['zorluk'], data.get('aciklama',''), data.get('kalori_tahmini', 0)))

    def update_antrenman(self, antrenman_id, data):
        with self.get_connection() as conn:
            conn.execute("""UPDATE antrenmanlar SET ad=?,kategori=?,saat=?,dakika=?,
                zorluk_seviyesi=?,aciklama=?,kalori_tahmini=? WHERE antrenman_id=?""",
                (data['ad'], data['kategori'], data['saat'], data['dakika'],
                 data['zorluk'], data.get('aciklama',''), data.get('kalori_tahmini',0),
                 antrenman_id))

    def delete_antrenman(self, antrenman_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE antrenmanlar SET durum='Pasif' WHERE antrenman_id=?",
                         (antrenman_id,))

    # ── TAKIPLER ───────────────────────────────────────────────────
    def get_takipler(self, search='', limit=50):
        with self.get_connection() as conn:
            q = """SELECT t.*, s.ad||' '||s.soyad as sporcu_adi,
                   a.ad as antrenman_adi, a.kategori
                   FROM takipler t
                   JOIN sporcular s ON t.sporcu_id=s.sporcu_id
                   JOIN antrenmanlar a ON t.antrenman_id=a.antrenman_id
                   WHERE t.durum='Aktif'"""
            p = []
            if search:
                q += " AND (s.ad LIKE ? OR s.soyad LIKE ? OR a.ad LIKE ?)"
                p += [f'%{search}%'] * 3
            q += " ORDER BY t.tarih DESC"
            return [dict(r) for r in conn.execute(q, p).fetchall()]

    def add_takip(self, data):
        with self.get_connection() as conn:
            conn.execute("""INSERT INTO takipler
                (sporcu_id,antrenman_id,kalori,nabiz,sure_dakika,notlar,tarih)
                VALUES (?,?,?,?,?,?,?)""",
                (data['sporcu_id'], data['antrenman_id'], data['kalori'],
                 data.get('nabiz',0), data.get('sure_dakika',0),
                 data.get('notlar',''), data.get('tarih', datetime.now().strftime('%Y-%m-%d %H:%M'))))

    # ── İLERLEMELER ────────────────────────────────────────────────
    def add_ilerleme(self, sporcu_id, kilo, notlar=''):
        with self.get_connection() as conn:
            conn.execute("UPDATE sporcular SET kilo=? WHERE sporcu_id=?", (kilo, sporcu_id))
            conn.execute("INSERT INTO ilerlemeler (sporcu_id,kilo,notlar) VALUES (?,?,?)",
                         (sporcu_id, kilo, notlar))

    def get_ilerlemeler(self, sporcu_id):
        with self.get_connection() as conn:
            return [dict(r) for r in conn.execute(
                "SELECT * FROM ilerlemeler WHERE sporcu_id=? ORDER BY tarih",
                (sporcu_id,)).fetchall()]

    # ── DASHBOARD ──────────────────────────────────────────────────
    def get_dashboard_stats(self):
        with self.get_connection() as conn:
            s = {}
            s['toplam_sporcu']    = conn.execute("SELECT COUNT(*) FROM sporcular WHERE durum='Aktif'").fetchone()[0]
            s['toplam_antrenman'] = conn.execute("SELECT COUNT(*) FROM antrenmanlar WHERE durum='Aktif'").fetchone()[0]
            s['toplam_takip']     = conn.execute("SELECT COUNT(*) FROM takipler WHERE durum='Aktif'").fetchone()[0]
            s['toplam_kalori']    = conn.execute("SELECT COALESCE(SUM(kalori),0) FROM takipler WHERE durum='Aktif'").fetchone()[0]
            # Kategori dağılımı
            rows = conn.execute("""SELECT a.kategori, COUNT(*) as sayi
                FROM takipler t JOIN antrenmanlar a ON t.antrenman_id=a.antrenman_id
                WHERE t.durum='Aktif' GROUP BY a.kategori""").fetchall()
            s['kategori_dagilimi'] = [(r[0], r[1]) for r in rows]
            # Hedef dağılımı
            rows2 = conn.execute("""SELECT hedef, COUNT(*) as sayi FROM sporcular
                WHERE durum='Aktif' GROUP BY hedef""").fetchall()
            s['hedef_dagilimi'] = [(r[0], r[1]) for r in rows2]
            # Son 7 gün takip
            rows3 = conn.execute("""SELECT DATE(tarih) as gun, COUNT(*) as sayi
                FROM takipler WHERE durum='Aktif'
                AND tarih >= date('now','-7 days')
                GROUP BY gun ORDER BY gun""").fetchall()
            s['haftalik_takip'] = [(r[0][-5:], r[1]) for r in rows3]
            # Cinsiyet dağılımı
            rows4 = conn.execute("""SELECT cinsiyet, COUNT(*) as sayi FROM sporcular
                WHERE durum='Aktif' GROUP BY cinsiyet""").fetchall()
            s['cinsiyet_dagilimi'] = [(r[0], r[1]) for r in rows4]
            # Tier 2 eklemeleri
            s['toplam_program'] = conn.execute(
                "SELECT COUNT(*) FROM antrenman_programlari WHERE aktif=1").fetchone()[0]
            s['bu_hafta_antrenman'] = conn.execute(
                """SELECT COUNT(*) FROM gunluk_log WHERE durum='Tamamlandi'
                   AND tarih >= date('now','-7 days')""").fetchone()[0]
            rows5 = conn.execute("""SELECT a.kategori, COUNT(*) as sayi
                FROM gunluk_log g JOIN antrenmanlar a ON g.antrenman_id=a.antrenman_id
                WHERE g.durum='Tamamlandi' GROUP BY a.kategori""").fetchall()
            s['log_kategori'] = [(r[0], r[1]) for r in rows5]
            return s


# ═══════════════════════════════════════════════════════════════════
# DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════
class BaseDialog(QDialog):
    def __init__(self, title, parent=None, min_width=580):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(min_width)
        self.setModal(True)
        self.setStyleSheet(f"""
            QDialog {{ background-color: {C['bg_card']}; }}
            QLabel {{ color: {C['text_secondary']}; font-size: 13px; font-weight: bold; }}
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox, QTextEdit, QDateEdit {{
                background-color: {C['bg_secondary']};
                border: 1.5px solid {C['border']}; border-radius: 6px;
                padding: 8px 12px; font-size: 13px; color: {C['text_main']};
                min-height: {INPUT_H}px;
            }}
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus,
            QDoubleSpinBox:focus, QDateEdit:focus {{
                border: 2px solid {C['primary']};
            }}
            QComboBox::drop-down {{ border: none; width: 28px; }}
            QComboBox QAbstractItemView {{
                background-color: {C['bg_secondary']}; color: {C['text_main']};
                selection-background-color: {C['primary']}; font-size: 13px;
            }}
        """)
        self._main_layout = QVBoxLayout(self)
        self._main_layout.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        self._main_layout.setSpacing(SPACING)
        # Başlık
        lbl = QLabel(title)
        lbl.setFont(QFont('Segoe UI', 18, QFont.Bold))
        lbl.setStyleSheet(f"color: {C['text_main']};")
        self._main_layout.addWidget(lbl)
        # Ayraç
        line = QFrame(); line.setFrameShape(QFrame.HLine)
        line.setStyleSheet(f"background: {C['border']}; max-height: 1px;")
        self._main_layout.addWidget(line)
        # Form grid
        self.form = QGridLayout(); self.form.setSpacing(12)
        self._main_layout.addLayout(self.form)
        self._form_row = 0
        self._main_layout.addStretch()
        # Buton alanı
        self._btn_layout = QHBoxLayout(); self._btn_layout.addStretch()
        self._main_layout.addLayout(self._btn_layout)

    def add_row(self, label, widget):
        lbl = QLabel(label)
        lbl.setFont(QFont('Segoe UI', 13, QFont.Bold))
        lbl.setStyleSheet(f"color: {C['text_secondary']};")
        self.form.addWidget(lbl, self._form_row, 0)
        self.form.addWidget(widget, self._form_row, 1)
        self._form_row += 1
        return widget

    def add_btn(self, text, color, callback):
        btn = make_btn(text, color)
        btn.clicked.connect(callback)
        self._btn_layout.addWidget(btn)
        return btn

    def inp(self, ph=''):
        w = QLineEdit(); w.setPlaceholderText(ph)
        w.setFont(QFont('Segoe UI', 13)); return w

    def spin(self, mn=0, mx=999, val=0):
        w = QSpinBox(); w.setRange(mn, mx); w.setValue(val)
        w.setFont(QFont('Segoe UI', 13)); return w

    def dspin(self, mn=0.0, mx=999.0, val=0.0, dec=1, prefix=''):
        w = QDoubleSpinBox(); w.setRange(mn, mx); w.setValue(val)
        w.setDecimals(dec); w.setSingleStep(0.5)
        if prefix: w.setPrefix(prefix)
        w.setFont(QFont('Segoe UI', 13)); return w

    def combo(self, items):
        w = QComboBox(); w.addItems(items)
        w.setFont(QFont('Segoe UI', 13)); return w

    def txt(self, ph='', h=80):
        w = QTextEdit(); w.setPlaceholderText(ph)
        w.setFixedHeight(h); w.setFont(QFont('Segoe UI', 13)); return w


class SporcuDialog(BaseDialog):
    def __init__(self, db, sporcu_data=None, parent=None):
        title = 'Sporcu Düzenle' if sporcu_data else 'Yeni Sporcu Ekle'
        super().__init__(title, parent, 600)
        self.db = db; self.sporcu_data = sporcu_data
        self._build()
        if sporcu_data: self._load()

    def _build(self):
        self.f_ad      = self.add_row('Ad *',         self.inp('Örn: Ahmet'))
        self.f_soyad   = self.add_row('Soyad *',      self.inp('Örn: Yılmaz'))
        self.f_yas     = self.add_row('Yaş',          self.spin(12, 100, 25))
        self.f_kilo    = self.add_row('Kilo (kg) *',  self.dspin(20, 300, 70.0))
        self.f_boy     = self.add_row('Boy (cm) *',   self.dspin(100, 250, 170.0))
        self.f_cinsiyet= self.add_row('Cinsiyet',     self.combo(['Erkek', 'Kadın']))
        self.f_hedef   = self.add_row('Hedef',        self.combo(
            ['Kilo Verme', 'Kas Kazanma', 'Form Koruma', 'Dayanıklılık', 'Genel Sağlık']))
        self.f_notlar  = self.add_row('Notlar',       self.txt('Sporcu hakkında notlar...'))
        if self.sporcu_data:
            self.f_durum = self.add_row('Durum', self.combo(['Aktif', 'Pasif']))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.sporcu_data
        self.f_ad.setText(d['ad']); self.f_soyad.setText(d['soyad'])
        self.f_yas.setValue(d['yas'] or 25)
        self.f_kilo.setValue(d['kilo']); self.f_boy.setValue(d['boy'])
        idx = self.f_cinsiyet.findText(d.get('cinsiyet','Erkek'))
        if idx >= 0: self.f_cinsiyet.setCurrentIndex(idx)
        idx2 = self.f_hedef.findText(d.get('hedef','Form Koruma'))
        if idx2 >= 0: self.f_hedef.setCurrentIndex(idx2)
        self.f_notlar.setPlainText(d.get('notlar',''))
        if hasattr(self, 'f_durum'):
            idx3 = self.f_durum.findText(d.get('durum','Aktif'))
            if idx3 >= 0: self.f_durum.setCurrentIndex(idx3)

    def _save(self):
        if not self.f_ad.text().strip() or not self.f_soyad.text().strip():
            dark_msg(self, 'Hata', 'Ad ve Soyad zorunludur!', QMessageBox.Warning); return
        data = {
            'ad': self.f_ad.text().strip(), 'soyad': self.f_soyad.text().strip(),
            'yas': self.f_yas.value(), 'kilo': self.f_kilo.value(),
            'boy': self.f_boy.value(), 'cinsiyet': self.f_cinsiyet.currentText(),
            'hedef': self.f_hedef.currentText(), 'notlar': self.f_notlar.toPlainText().strip(),
            'durum': self.f_durum.currentText() if hasattr(self,'f_durum') else 'Aktif',
        }
        if self.sporcu_data:
            self.db.update_sporcu(self.sporcu_data['sporcu_id'], data)
        else:
            self.db.add_sporcu(data)
        self.accept()


class AntrenmanDialog(BaseDialog):
    def __init__(self, db, antrenman_data=None, parent=None):
        title = 'Antrenman Düzenle' if antrenman_data else 'Yeni Antrenman Ekle'
        super().__init__(title, parent, 600)
        self.db = db; self.antrenman_data = antrenman_data
        self._build()
        if antrenman_data: self._load()

    def _build(self):
        self.f_ad      = self.add_row('Antrenman Adı *', self.inp('Örn: Koşu, Bench Press'))
        self.f_kat     = self.add_row('Kategori',         self.combo(
            ['Kardiyo','Kuvvet','Esneklik','Denge','HIIT','Spor']))
        # Süre satırı - saat + dakika yan yana
        sure_w = QWidget(); sure_l = QHBoxLayout(sure_w); sure_l.setContentsMargins(0,0,0,0)
        self.f_saat = QSpinBox(); self.f_saat.setRange(0,5); self.f_saat.setPrefix('Saat: ')
        self.f_saat.setFont(QFont('Segoe UI',13)); self.f_saat.setFixedHeight(INPUT_H)
        self.f_dakika = QSpinBox(); self.f_dakika.setRange(0,59); self.f_dakika.setPrefix('Dakika: ')
        self.f_dakika.setValue(30); self.f_dakika.setFont(QFont('Segoe UI',13)); self.f_dakika.setFixedHeight(INPUT_H)
        sure_l.addWidget(self.f_saat); sure_l.addWidget(self.f_dakika)
        self.add_row('Süre *', sure_w)
        self.f_zorluk  = self.add_row('Zorluk',    self.combo(['Kolay','Orta','Zor','Profesyonel']))
        self.f_kalori  = self.add_row('Tahmini Kalori', self.spin(0, 1000, 250))
        self.f_aciklama= self.add_row('Açıklama',  self.txt('Antrenman detayları...'))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _load(self):
        d = self.antrenman_data
        self.f_ad.setText(d['ad'])
        idx = self.f_kat.findText(d.get('kategori','Kardiyo'))
        if idx >= 0: self.f_kat.setCurrentIndex(idx)
        self.f_saat.setValue(d.get('saat',0)); self.f_dakika.setValue(d.get('dakika',30))
        idx2 = self.f_zorluk.findText(d.get('zorluk_seviyesi','Orta'))
        if idx2 >= 0: self.f_zorluk.setCurrentIndex(idx2)
        self.f_kalori.setValue(d.get('kalori_tahmini',0))
        self.f_aciklama.setPlainText(d.get('aciklama',''))

    def _save(self):
        if not self.f_ad.text().strip():
            dark_msg(self, 'Hata', 'Antrenman adı zorunludur!', QMessageBox.Warning); return
        if self.f_saat.value() == 0 and self.f_dakika.value() == 0:
            dark_msg(self, 'Hata', 'Süre giriniz!', QMessageBox.Warning); return
        data = {
            'ad': self.f_ad.text().strip(), 'kategori': self.f_kat.currentText(),
            'saat': self.f_saat.value(), 'dakika': self.f_dakika.value(),
            'zorluk': self.f_zorluk.currentText(), 'kalori_tahmini': self.f_kalori.value(),
            'aciklama': self.f_aciklama.toPlainText().strip(),
        }
        if self.antrenman_data:
            self.db.update_antrenman(self.antrenman_data['antrenman_id'], data)
        else:
            self.db.add_antrenman(data)
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# SAYFALAR
# ═══════════════════════════════════════════════════════════════════
class DashboardPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Dashboard')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_clock = QLabel()
        self.lbl_clock.setFont(QFont('Segoe UI', 14, QFont.Bold))
        self.lbl_clock.setStyleSheet(f"color: {C['primary_light']};")
        hdr.addWidget(self.lbl_clock)
        hdr.addSpacing(16)
        self.lbl_update = QLabel()
        self.lbl_update.setStyleSheet(f"color: {C['text_muted']}; font-size: 11px;")
        hdr.addWidget(self.lbl_update)
        hdr.addSpacing(10)
        btn_refresh = make_btn('⟳  Yenile', C['primary'], small=True)
        btn_refresh.setMinimumWidth(100)
        btn_refresh.clicked.connect(self.refresh)
        hdr.addWidget(btn_refresh)
        lay.addLayout(hdr)

        # KPI satırı
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_sporcu    = KPICard('Aktif Sporcu',    '—', '🏃', '#6366f1', '#4f46e5')
        self.k_antrenman = KPICard('Antrenman Tipi',  '—', '💪', '#10b981', '#059669')
        self.k_takip     = KPICard('Toplam Takip',    '—', '📋', '#3b82f6', '#2563eb')
        self.k_kalori    = KPICard('Toplam Kalori',   '—', '🔥', '#f59e0b', '#d97706')
        for w in [self.k_sporcu, self.k_antrenman, self.k_takip, self.k_kalori]:
            kpi.addWidget(w)
        lay.addLayout(kpi)

        # Grafikler
        charts = QHBoxLayout(); charts.setSpacing(SPACING)

        def chart_frame(widget):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(10,10,10,10)
            fl.addWidget(widget); return f

        self.chart_kategori = PieChartWidget()
        self.chart_kategori.setMinimumHeight(240)
        charts.addWidget(chart_frame(self.chart_kategori), 1)

        self.chart_haftalik = BarChartWidget()
        self.chart_haftalik.setMinimumHeight(240)
        charts.addWidget(chart_frame(self.chart_haftalik), 1)

        self.chart_hedef = PieChartWidget()
        self.chart_hedef.setMinimumHeight(240)
        charts.addWidget(chart_frame(self.chart_hedef), 1)

        lay.addLayout(charts)

        # Clock timer
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(1000)
        self._tick()

    def _tick(self):
        self.lbl_clock.setText(datetime.now().strftime('%H:%M:%S'))

    def refresh(self):
        s = self.db.get_dashboard_stats()
        self.k_sporcu.set_value(str(s['toplam_sporcu']))
        self.k_antrenman.set_value(str(s['toplam_antrenman']))
        self.k_takip.set_value(str(s['toplam_takip']))
        kal = s['toplam_kalori']
        self.k_kalori.set_value(f"{kal:,} kcal")
        self.chart_kategori.set_data(s['kategori_dagilimi'], 'Antrenman Kategorileri')
        self.chart_hedef.set_data(s['hedef_dagilimi'], 'Sporcu Hedefleri')
        self.chart_haftalik.set_data(s['haftalik_takip'], 'Son 7 Gün Takip')
        self.lbl_update.setText(f"Son güncelleme: {datetime.now().strftime('%d.%m.%Y %H:%M')}")


class SporcularPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Sporcular')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('0 sporcu')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Sporcu', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # Filtreler
        filt = QHBoxLayout(); filt.setSpacing(10)
        self.search = make_search('Ad, soyad veya hedef ara...')
        self.search.textChanged.connect(lambda: self._timer.start(300))
        filt.addWidget(self.search, 1)
        filt.addWidget(QLabel('Durum:'))
        self.durum_cb = make_combo(['Tümü','Aktif','Pasif'], 130)
        self.durum_cb.currentTextChanged.connect(self.refresh)
        filt.addWidget(self.durum_cb)
        lay.addLayout(filt)

        # Tablo
        self.table = make_table(
            ['ID','Ad','Soyad','Yaş','Kilo','Boy','BMI','Cinsiyet','Hedef','Durum'],
            [0, -1, 130, 55, 80, 80, 65, 80, 130, 75]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        # Butonlar
        bot = QHBoxLayout()
        self.btn_edit = make_btn('Düzenle', C['primary']); self.btn_edit.setEnabled(False)
        self.btn_edit.clicked.connect(self._edit)
        self.btn_del = make_btn('Pasife Al', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_edit); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_sporcular(self.search.text(), self.durum_cb.currentText())
        self.lbl_count.setText(f"{len(data)} sporcu")
        self.table.setRowCount(0)
        hedef_c = {
            'Kilo Verme': C['danger'], 'Kas Kazanma': C['success'],
            'Form Koruma': C['primary'], 'Dayanıklılık': C['warning'],
            'Genel Sağlık': C['info'],
        }
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            bmi = d.get('bmi') or 0
            bmi_color = C['success'] if 18.5 <= bmi <= 24.9 else (C['warning'] if bmi < 18.5 else C['danger'])
            vals = [str(d['sporcu_id']), d['ad'], d['soyad'], str(d.get('yas','—')),
                    f"{d['kilo']} kg", f"{d['boy']} cm", f"{bmi:.1f}",
                    d.get('cinsiyet','—'), d.get('hedef','—'), d.get('durum','Aktif')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 6:  # BMI rengi
                    item.setForeground(QColor(bmi_color))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 8:  # Hedef rengi
                    item.setForeground(QColor(hedef_c.get(val, C['text_main'])))
                if col == 9:  # Durum
                    item.setForeground(QColor(C['success'] if val == 'Aktif' else C['text_muted']))
                self.table.setItem(row, col, item)
        self.selected_id = None
        self.btn_edit.setEnabled(False); self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(), 0).text())
            self.btn_edit.setEnabled(True); self.btn_del.setEnabled(True)

    def _add(self):
        if SporcuDialog(self.db, parent=self).exec_() == QDialog.Accepted:
            self.refresh()

    def _edit(self):
        if not self.selected_id: return
        data = self.db.get_sporcu(self.selected_id)
        if data and SporcuDialog(self.db, data, self).exec_() == QDialog.Accepted:
            self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        name = f"{self.table.item(row,1).text()} {self.table.item(row,2).text()}"
        if dark_confirm(self, 'Pasife Al', f'"{name}" pasife alınacak. Emin misiniz?'):
            self.db.delete_sporcu(self.selected_id); self.refresh()


class AntrenmanlarPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Antrenmanlar')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('0 antrenman')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Antrenman', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        filt = QHBoxLayout(); filt.setSpacing(10)
        self.search = make_search('Antrenman adı veya açıklama ara...')
        self.search.textChanged.connect(lambda: self._timer.start(300))
        filt.addWidget(self.search, 1)
        filt.addWidget(QLabel('Kategori:'))
        self.kat_cb = make_combo(['Tümü','Kardiyo','Kuvvet','Esneklik','Denge','HIIT','Spor'], 140)
        self.kat_cb.currentTextChanged.connect(self.refresh)
        filt.addWidget(self.kat_cb)
        lay.addLayout(filt)

        self.table = make_table(
            ['ID','Antrenman Adı','Kategori','Süre','Zorluk','Tahmini Kalori','Açıklama'],
            [0, -1, 110, 90, 110, 130, 220]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        self.table.doubleClicked.connect(self._edit)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_edit = make_btn('Düzenle', C['primary']); self.btn_edit.setEnabled(False)
        self.btn_edit.clicked.connect(self._edit)
        self.btn_del = make_btn('Sil', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_edit); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_antrenmanlar(self.search.text(), self.kat_cb.currentText())
        self.lbl_count.setText(f"{len(data)} antrenman")
        self.table.setRowCount(0)
        zorluk_c = {'Kolay': C['success'], 'Orta': C['warning'],
                    'Zor': C['danger'], 'Profesyonel': '#ec4899'}
        kat_c = {'Kardiyo': C['info'], 'Kuvvet': C['danger'],
                 'Esneklik': C['success'], 'HIIT': '#f59e0b',
                 'Denge': C['primary'], 'Spor': '#ec4899'}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            sure = f"{d['saat']}s {d['dakika']}dk" if d['saat'] > 0 else f"{d['dakika']} dk"
            vals = [str(d['antrenman_id']), d['ad'], d['kategori'], sure,
                    d['zorluk_seviyesi'], f"{d.get('kalori_tahmini',0)} kcal",
                    d.get('aciklama','')[:50]]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 2:
                    item.setForeground(QColor(kat_c.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 4:
                    item.setForeground(QColor(zorluk_c.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_id = None
        self.btn_edit.setEnabled(False); self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(), 0).text())
            self.btn_edit.setEnabled(True); self.btn_del.setEnabled(True)

    def _add(self):
        if AntrenmanDialog(self.db, parent=self).exec_() == QDialog.Accepted:
            self.refresh()

    def _edit(self):
        if not self.selected_id: return
        data = self.db.get_antrenman(self.selected_id)
        if data and AntrenmanDialog(self.db, data, self).exec_() == QDialog.Accepted:
            self.refresh()

    def _delete(self):
        if not self.selected_id: return
        name = self.table.item(self.table.currentRow(), 1).text()
        if dark_confirm(self, 'Sil', f'"{name}" silinecek. Emin misiniz?'):
            self.db.delete_antrenman(self.selected_id); self.refresh()


# ═══════════════════════════════════════════════════════════════════
# LOGIN PENCERESİ
# ═══════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════
# TIER 2 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — QPainter Line Chart (İlerleme Grafikleri)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 4 — Makro Pie Chart Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 5 — Multi-Line Chart Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 6 — Rozet Kartı Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 7 — Radar (Spider) Chart Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════
import math


# ═══════════════════════════════════════════════════════════════════
# TIER 8 — GELİŞMİŞ ANALİZ SAYFALARI
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 9 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Aktivite Feed Widget (QPainter)
# ═══════════════════════════════════════════════════════════════════

class AktiviteFeedWidget(QWidget):
    """Son aktiviteler listesi — QPainter kartları."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = []
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_items(self, items):
        self.items = items[:8]; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        if not self.items:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Henüz aktivite yok')
            p.end(); return

        row_h = min(40, (h - 8) // max(len(self.items), 1))
        tip_renk = {
            'antrenman': C['primary'], 'rozet': C['accent'],
            'hedef': C['success'], 'olcum': C['info'],
            'streak': C['danger'], 'beslenme': C['warning'],
        }

        for i, item in enumerate(self.items):
            y = 4 + i * row_h
            renk = tip_renk.get(item.get('aktivite_tipi',''), C['text_muted'])

            # Satır arka planı (alternatif)
            if i % 2 == 0:
                p.fillRect(0, y, w, row_h, QColor(C['bg_secondary']))

            # Sol renkli şerit
            p.fillRect(0, y+3, 3, row_h-6, QColor(renk))

            # İkon
            p.setPen(QColor(renk)); p.setFont(QFont('Segoe UI', 14))
            p.drawText(QRectF(8, y, 28, row_h), Qt.AlignCenter, item.get('ikon','📌'))

            # Başlık
            p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 11, QFont.Bold))
            p.drawText(QRectF(40, y, w * 0.4, row_h), Qt.AlignLeft | Qt.AlignVCenter, item.get('baslik',''))

            # Sporcu adı
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 10))
            p.drawText(QRectF(w * 0.42, y, w * 0.3, row_h),
                       Qt.AlignLeft | Qt.AlignVCenter, item.get('sporcu_adi',''))

            # Tarih
            tarih = item.get('tarih','')[:16]
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(w * 0.75, y, w * 0.24, row_h),
                       Qt.AlignRight | Qt.AlignVCenter, tarih)

        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Gelişmiş Dashboard (8 KPI + 3 grafik + feed)
# ═══════════════════════════════════════════════════════════════════

class GelismisDashboardPage(QWidget):
    """Tier 10 Dashboard — tam featured."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Dashboard')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_clock = QLabel()
        self.lbl_clock.setFont(QFont('Segoe UI', 14, QFont.Bold))
        self.lbl_clock.setStyleSheet(f"color:{C['primary_light']};")
        hdr.addWidget(self.lbl_clock)
        hdr.addSpacing(16)
        self.lbl_update = QLabel()
        self.lbl_update.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        hdr.addWidget(self.lbl_update)
        hdr.addSpacing(10)
        btn_ref = make_btn('⟳  Yenile', C['primary'], small=True)
        btn_ref.setMinimumWidth(100); btn_ref.clicked.connect(self.refresh)
        hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        # KPI satırı 1 (4 kart)
        kpi1 = QHBoxLayout(); kpi1.setSpacing(SPACING)
        self.k_sporcu   = KPICard('Aktif Sporcu',    '—', '🏃', '#6366f1', '#4f46e5')
        self.k_log      = KPICard('Toplam Antrenman', '—', '💪', '#10b981', '#059669')
        self.k_kalori   = KPICard('Yakılan Kalori',   '—', '🔥', '#f59e0b', '#d97706')
        self.k_sure     = KPICard('Toplam Süre',      '—', '⏱', '#3b82f6', '#2563eb')
        for k in [self.k_sporcu, self.k_log, self.k_kalori, self.k_sure]:
            kpi1.addWidget(k)
        lay.addLayout(kpi1)

        # KPI satırı 2 (4 kart)
        kpi2 = QHBoxLayout(); kpi2.setSpacing(SPACING)
        self.k_rozet    = KPICard('Kazanılan Rozet',  '—', '🏅', '#8b5cf6', '#7c3aed')
        self.k_streak   = KPICard('Bugün Streak',     '—', '🔥', '#ef4444', '#dc2626')
        self.k_puan     = KPICard('Toplam Puan',      '—', '⭐', '#f59e0b', '#d97706')
        self.k_hafta    = KPICard('Bu Hafta Antrenman','—', '📅', '#14b8a6', '#0d9488')
        for k in [self.k_rozet, self.k_streak, self.k_puan, self.k_hafta]:
            kpi2.addWidget(k)
        lay.addLayout(kpi2)

        # Grafikler (3'lü)
        charts = QHBoxLayout(); charts.setSpacing(SPACING)

        def chart_frame(widget, h=220):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            widget.setMinimumHeight(h); fl.addWidget(widget)
            return f

        self.chart_kat    = PieChartWidget()
        self.chart_trend  = BarChartWidget()
        self.chart_hedef  = PieChartWidget()
        charts.addWidget(chart_frame(self.chart_kat), 1)
        charts.addWidget(chart_frame(self.chart_trend), 1)
        charts.addWidget(chart_frame(self.chart_hedef), 1)
        lay.addLayout(charts)

        # Alt: En aktif + Aktivite feed — stretch=1 ile kalan alanı al
        bot = QHBoxLayout(); bot.setSpacing(SPACING)

        # Sol: En aktif sporcular
        left = QFrame()
        left.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ll = QVBoxLayout(left); ll.setContentsMargins(12,10,12,10); ll.setSpacing(6)
        lbl_aktif = QLabel('🏆 En Aktif (Son 30 Gün)')
        lbl_aktif.setFont(QFont('Segoe UI', 12, QFont.Bold)); ll.addWidget(lbl_aktif)
        self.tbl_aktif = make_table(['Sporcu','Antrenman'], [-1, 100])
        ll.addWidget(self.tbl_aktif, 1)
        bot.addWidget(left, 1)

        # Sağ: Son aktiviteler feed
        right = QFrame()
        right.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        rl = QVBoxLayout(right); rl.setContentsMargins(12,10,12,10); rl.setSpacing(6)
        lbl_feed = QLabel('📡 Son Aktiviteler')
        lbl_feed.setFont(QFont('Segoe UI', 12, QFont.Bold)); rl.addWidget(lbl_feed)
        self.feed_widget = AktiviteFeedWidget()
        rl.addWidget(self.feed_widget, 1)
        bot.addWidget(right, 2)
        lay.addLayout(bot, 1)

        # Clock timer
        self._timer = QTimer(self); self._timer.timeout.connect(self._tick); self._timer.start(1000)
        self._tick()

    def _tick(self):
        self.lbl_clock.setText(datetime.now().strftime('%H:%M:%S'))

    def refresh(self):
        s = self.db.get_gelismis_dashboard_stats()

        self.k_sporcu.set_value(str(s['toplam_sporcu']))
        self.k_log.set_value(str(s['toplam_log']))
        self.k_kalori.set_value(f"{int(s['toplam_kalori_yak']):,}")
        sure_saat = s['toplam_sure'] // 60
        self.k_sure.set_value(f"{sure_saat}s")
        self.k_rozet.set_value(str(s['toplam_rozet']))
        self.k_streak.set_value(str(s['aktif_streak']))
        self.k_puan.set_value(f"{s['toplam_puan']:,}")
        self.k_hafta.set_value(str(s['bu_hafta_log']))

        self.chart_kat.set_data(s['kategori_dagilimi'], 'Antrenman Kategorileri')
        self.chart_trend.set_data(s['haftalik_trend'], '8 Haftalık Trend')
        self.chart_hedef.set_data(s['hedef_dagilimi'], 'Sporcu Hedefleri')

        # En aktif
        self.tbl_aktif.setRowCount(0)
        for ad, sayi in s['en_aktif']:
            row = self.tbl_aktif.rowCount(); self.tbl_aktif.insertRow(row)
            item_ad = QTableWidgetItem(ad); item_ad.setFont(QFont('Segoe UI', 12))
            item_s  = QTableWidgetItem(str(sayi)); item_s.setFont(QFont('Segoe UI', 12, QFont.Bold))
            item_s.setForeground(QColor(C['accent']))
            self.tbl_aktif.setItem(row, 0, item_ad)
            self.tbl_aktif.setItem(row, 1, item_s)

        # Aktivite feed
        feed = self.db.get_aktivite_feed(limit=10)
        if not feed:
            # Son aktivitelerden feed oluştur
            son = self.db.get_son_aktiviteler(8)
            feed = [{'ikon':'💪','baslik':f"{r['antrenman_adi']} tamamlandı",
                     'sporcu_adi': r['sporcu_adi'], 'tarih': r['tarih'],
                     'aktivite_tipi':'antrenman'} for r in son]
        self.feed_widget.set_items(feed)
        self.lbl_update.setText(f"Son güncelleme: {datetime.now().strftime('%d.%m.%Y %H:%M')}")


# ═══════════════════════════════════════════════════════════════════
# TIER 10 — Excel Export Sayfası
# ═══════════════════════════════════════════════════════════════════

class ExcelExportPage(QWidget):
    """Tüm sistem Excel export."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Excel & Raporlar')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # Excel export
        ex_frame = QFrame()
        ex_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ex_l = QVBoxLayout(ex_frame); ex_l.setContentsMargins(24,20,24,20); ex_l.setSpacing(14)

        lbl_ex = QLabel('📊 Excel Özet Raporu')
        lbl_ex.setFont(QFont('Segoe UI', 16, QFont.Bold)); ex_l.addWidget(lbl_ex)

        info = QLabel('Tüm sporcular, antrenman logları, beslenme verileri ve liderboard tablosunu\n'
                      'tek bir Excel dosyasına (4 sayfa) aktarır.')
        info.setStyleSheet(f"color:{C['text_secondary']};font-size:13px;")
        info.setWordWrap(True); ex_l.addWidget(info)

        icerik = QLabel('📋 İçerik:\n'
                        '  • Sayfa 1 — Sporcular listesi\n'
                        '  • Sayfa 2 — Tüm antrenman logları\n'
                        '  • Sayfa 3 — Beslenme kayıtları (son 500)\n'
                        '  • Sayfa 4 — Liderboard & puan tablosu')
        icerik.setStyleSheet(f"color:{C['text_secondary']};font-size:12px;")
        ex_l.addWidget(icerik)

        btn_excel = make_btn('📥 Excel İndir', C['success'])
        btn_excel.setMaximumWidth(200); btn_excel.clicked.connect(self._excel_export)
        ex_l.addWidget(btn_excel)
        lay.addWidget(ex_frame)

        # CSV export sporcu bazlı
        csv_frame = QFrame()
        csv_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        cv_l = QVBoxLayout(csv_frame); cv_l.setContentsMargins(24,20,24,20); cv_l.setSpacing(14)

        lbl_csv = QLabel('📤 Sporcu CSV Export')
        lbl_csv.setFont(QFont('Segoe UI', 16, QFont.Bold)); cv_l.addWidget(lbl_csv)

        info2 = QLabel('Seçilen sporcunun tüm verisini (antrenman, beslenme, ölçümler) CSV formatında aktar.')
        info2.setStyleSheet(f"color:{C['text_secondary']};font-size:13px;")
        info2.setWordWrap(True); cv_l.addWidget(info2)

        row_csv = QHBoxLayout()
        row_csv.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 220)
        self._sporcu_ids = [None]
        row_csv.addWidget(self.sporcu_cb)
        btn_csv = make_btn('📥 CSV İndir', C['primary'], small=True)
        btn_csv.clicked.connect(self._csv_export)
        row_csv.addWidget(btn_csv); row_csv.addStretch()
        cv_l.addLayout(row_csv)
        lay.addWidget(csv_frame)

        lay.addStretch()
        self._load_sporcular()

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.clear(); self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])

    def refresh(self):
        self._load_sporcular()

    def _excel_export(self):
        from PyQt5.QtWidgets import QFileDialog
        dosya, _ = QFileDialog.getSaveFileName(
            self, 'Excel Kaydet',
            f"fitness_rapor_{datetime.now().strftime('%Y%m%d')}.xlsx",
            'Excel (*.xlsx)')
        if not dosya: return
        sonuc = self.db.export_excel_ozet(dosya)
        if sonuc is True:
            dark_msg(self,'Başarılı',f'Excel raporu oluşturuldu:\n{dosya}\n\n4 sayfa: Sporcular, Loglar, Beslenme, Liderboard')
        elif sonuc == 'openpyxl_yok':
            dark_msg(self,'Hata',
                     'openpyxl kütüphanesi yüklü değil.\n\nKurulum:\n  pip install openpyxl',
                     QMessageBox.Warning)
        else:
            dark_msg(self,'Hata',f'Dışa aktarma başarısız:\n{sonuc}',QMessageBox.Warning)

    def _csv_export(self):
        from PyQt5.QtWidgets import QFileDialog
        idx = self.sporcu_cb.currentIndex()
        if idx == 0:
            dark_msg(self,'Uyarı','Sporcu seçin!',QMessageBox.Warning); return
        sporcu_id = self._sporcu_ids[idx]
        ad = self.sporcu_cb.currentText().replace(' ','_')
        dosya, _ = QFileDialog.getSaveFileName(
            self, 'CSV Kaydet', f"fitness_{ad}.csv", 'CSV (*.csv)')
        if dosya:
            if self.db.export_sporcu_csv(sporcu_id, dosya):
                dark_msg(self,'Başarılı',f'CSV dosyası kaydedildi:\n{dosya}')
            else:
                dark_msg(self,'Hata','Dışa aktarma başarısız.',QMessageBox.Warning)



class SablonlarPage(QWidget):
    """Antrenman şablonları — görüntüle, sporcu programına uygula."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sablon_id = None
        self._build(); self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Antrenman Şablonları')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")
        splitter.setChildrenCollapsible(False)

        # Sol: Şablon listesi
        left = QWidget()
        ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)
        lbl_s = QLabel('Hazır Şablonlar')
        lbl_s.setFont(QFont('Segoe UI', 13, QFont.Bold)); ll.addWidget(lbl_s)
        self.tbl_sablon = make_table(
            ['ID','Şablon Adı','Kategori','Süre'],
            [0, -1, 110, 80]
        )
        self.tbl_sablon.hideColumn(0)
        self.tbl_sablon.itemSelectionChanged.connect(self._on_sablon_select)
        ll.addWidget(self.tbl_sablon, 1)

        # Sporcu uygula
        bot_l = QHBoxLayout()
        bot_l.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 180)
        self._sporcu_ids = [None]
        bot_l.addWidget(self.sporcu_cb)
        self.btn_uygula = make_btn('✅ Uygula', C['success'], small=True)
        self.btn_uygula.setEnabled(False)
        self.btn_uygula.clicked.connect(self._uygula)
        bot_l.addWidget(self.btn_uygula); bot_l.addStretch()
        ll.addLayout(bot_l)
        splitter.addWidget(left)

        # Sağ: Şablon günleri
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0); rl.setSpacing(8)
        lbl_g = QLabel('Şablon Detayı')
        lbl_g.setFont(QFont('Segoe UI', 13, QFont.Bold)); rl.addWidget(lbl_g)
        self.lbl_sablon_aciklama = QLabel('')
        self.lbl_sablon_aciklama.setStyleSheet(f"color:{C['text_muted']};font-size:12px;")
        self.lbl_sablon_aciklama.setWordWrap(True)
        self.lbl_sablon_aciklama.setFixedHeight(20)
        rl.addWidget(self.lbl_sablon_aciklama)
        self.tbl_gunler = make_table(
            ['Sıra','Gün','Antrenman','Kategori','Set','Tekrar','Ağırlık'],
            [55, 110, -1, 110, 55, 65, 80]
        )
        rl.addWidget(self.tbl_gunler, 1)
        splitter.addWidget(right)
        splitter.setSizes([400, 580])
        lay.addWidget(splitter, 1)

    def refresh(self):
        # Sporcuları yükle
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True); self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --'); self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

        # Şablonları yükle
        sablonlar = self.db.get_antrenman_sablonlari()
        self.tbl_sablon.setRowCount(0)
        kat_renk = {'Kuvvet':C['danger'],'Genel':C['primary'],'HIIT':C['warning']}
        for s in sablonlar:
            row = self.tbl_sablon.rowCount(); self.tbl_sablon.insertRow(row)
            vals = [str(s['sablon_id']), s['sablon_adi'],
                    s['kategori'], f"{s['sure_hafta']} Hafta"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 2:
                    item.setForeground(QColor(kat_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_sablon.setItem(row, col, item)

    def _on_sablon_select(self):
        if not self.tbl_sablon.selectedItems(): return
        self.selected_sablon_id = int(self.tbl_sablon.item(
            self.tbl_sablon.currentRow(), 0).text())
        self.btn_uygula.setEnabled(True)

        # Günleri yükle
        sablonlar = self.db.get_antrenman_sablonlari()
        aciklama = ''
        for s in sablonlar:
            if s['sablon_id'] == self.selected_sablon_id:
                aciklama = s.get('aciklama',''); break
        self.lbl_sablon_aciklama.setText(aciklama)

        gunler = self.db.get_sablon_gunleri(self.selected_sablon_id)
        self.tbl_gunler.setRowCount(0)
        gun_renk = {'Pazartesi':C['primary'],'Salı':C['success'],'Çarşamba':C['warning'],
                    'Perşembe':C['info'],'Cuma':C['danger'],'Cumartesi':'#ec4899','Pazar':'#8b5cf6'}
        for g in gunler:
            row = self.tbl_gunler.rowCount(); self.tbl_gunler.insertRow(row)
            ag = f"{g['agirlik']} kg" if g.get('agirlik') else '—'
            vals = [str(g['gun_sira']), g['gun_adi'],
                    g.get('antrenman_adi','—'), g.get('kategori','—'),
                    str(g['set_sayisi']), str(g['tekrar']), ag]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(gun_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_gunler.setItem(row, col, item)

    def _uygula(self):
        if not self.selected_sablon_id: return
        idx = self.sporcu_cb.currentIndex()
        if idx == 0:
            dark_msg(self,'Uyarı','Sporcu seçin!',QMessageBox.Warning); return
        sporcu_id = self._sporcu_ids[idx]
        prog_id = self.db.sablon_sporcu_uygula(self.selected_sablon_id, sporcu_id)
        if prog_id:
            dark_msg(self,'Başarılı','Şablon sporcu programına eklendi! ✅\nProgramlar sayfasından görüntüleyebilirsiniz.')
        else:
            dark_msg(self,'Hata','Şablon uygulanamadı.',QMessageBox.Warning)


class SistemAyarlariPage(QWidget):
    """Sistem ayarları, CSV export, yedekleme, hatırlatıcılar."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()
        # Hatırlatıcı timer — her dakika kontrol
        self._hat_timer = QTimer(self)
        self._hat_timer.timeout.connect(self._hatirlatici_kontrol)
        self._hat_timer.start(60000)

    def _build(self):
        from PyQt5.QtWidgets import QTabWidget, QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Sistem Ayarları')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        tabs = QTabWidget()

        # ── Tab 1: Genel Ayarlar ──────────────────────────────────
        tab_genel = QWidget()
        gl = QVBoxLayout(tab_genel); gl.setContentsMargins(16,16,16,16); gl.setSpacing(14)

        # Sistem özeti kartları
        kpi_s = QHBoxLayout(); kpi_s.setSpacing(SPACING)
        self.k_sporcu_s  = KPICard('Sporcu',     '—', '🏃', C['primary'], C['primary_dark'])
        self.k_log_s     = KPICard('Log',        '—', '📋', C['info'],    C['info_dark'])
        self.k_besin_s   = KPICard('Besin',      '—', '🥗', C['success'], C['success_dark'])
        self.k_db_boyut  = KPICard('DB Boyutu',  '—', '💾', C['warning'], C['warning_dark'])
        for k in [self.k_sporcu_s, self.k_log_s, self.k_besin_s, self.k_db_boyut]:
            kpi_s.addWidget(k)
        gl.addLayout(kpi_s)

        # Ayarlar grid
        ayar_frame = QFrame()
        ayar_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        af_l = QGridLayout(ayar_frame); af_l.setContentsMargins(16,16,16,16); af_l.setSpacing(14)

        def ayar_row(row, lbl_text, widget):
            lbl = QLabel(lbl_text)
            lbl.setFont(QFont('Segoe UI', 13, QFont.Bold))
            lbl.setStyleSheet(f"color:{C['text_secondary']};")
            af_l.addWidget(lbl, row, 0)
            af_l.addWidget(widget, row, 1)

        self.cb_birim = make_combo(['kg', 'lbs'], 120)
        ayar_row(0, 'Ağırlık Birimi', self.cb_birim)

        self.cb_dil = make_combo(['Türkçe', 'English'], 120)
        ayar_row(1, 'Dil', self.cb_dil)

        self.cb_kalori = make_combo(['kcal', 'kJ'], 120)
        ayar_row(2, 'Kalori Birimi', self.cb_kalori)

        btn_kaydet_ayar = make_btn('Ayarları Kaydet', C['primary'], small=True)
        btn_kaydet_ayar.clicked.connect(self._kaydet_ayarlar)
        af_l.addWidget(btn_kaydet_ayar, 3, 0, 1, 2)
        gl.addWidget(ayar_frame)
        gl.addStretch()
        tabs.addTab(tab_genel, '⚙️ Genel')

        # ── Tab 2: Export & Yedekleme ─────────────────────────────
        tab_export = QWidget()
        el = QVBoxLayout(tab_export); el.setContentsMargins(16,16,16,16); el.setSpacing(14)

        # CSV Export
        exp_frame = QFrame()
        exp_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        ex_l = QVBoxLayout(exp_frame); ex_l.setContentsMargins(16,14,16,14); ex_l.setSpacing(10)
        lbl_exp = QLabel('📤 CSV Export')
        lbl_exp.setFont(QFont('Segoe UI', 14, QFont.Bold)); ex_l.addWidget(lbl_exp)
        lbl_exp_info = QLabel('Sporcu verilerini (antrenman, beslenme, ölçümler) CSV olarak dışa aktar.')
        lbl_exp_info.setStyleSheet(f"color:{C['text_muted']};font-size:12px;")
        lbl_exp_info.setWordWrap(True); ex_l.addWidget(lbl_exp_info)
        exp_row = QHBoxLayout()
        exp_row.addWidget(QLabel('Sporcu:'))
        self.export_sporcu_cb = make_combo(['-- Sporcu Seç --'], 200)
        self._export_sporcu_ids = [None]
        exp_row.addWidget(self.export_sporcu_cb)
        btn_export = make_btn('📥 CSV İndir', C['success'], small=True)
        btn_export.clicked.connect(self._export_csv)
        exp_row.addWidget(btn_export); exp_row.addStretch()
        ex_l.addLayout(exp_row)
        el.addWidget(exp_frame)

        # Yedekleme
        ydk_frame = QFrame()
        ydk_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        yk_l = QVBoxLayout(ydk_frame); yk_l.setContentsMargins(16,14,16,14); yk_l.setSpacing(10)
        lbl_ydk = QLabel('💾 Veritabanı Yedekleme')
        lbl_ydk.setFont(QFont('Segoe UI', 14, QFont.Bold)); yk_l.addWidget(lbl_ydk)
        lbl_ydk_info = QLabel('SQLite veritabanını güvenli bir konuma yedekle.')
        lbl_ydk_info.setStyleSheet(f"color:{C['text_muted']};font-size:12px;")
        yk_l.addWidget(lbl_ydk_info)
        btn_ydk = make_btn('💾 Yedekle', C['primary'], small=True)
        btn_ydk.clicked.connect(self._yedekle); yk_l.addWidget(btn_ydk)
        el.addWidget(ydk_frame)
        el.addStretch()
        tabs.addTab(tab_export, '📤 Export')

        # ── Tab 3: Hatırlatıcılar ─────────────────────────────────
        tab_hat = QWidget()
        hl_tab = QVBoxLayout(tab_hat); hl_tab.setContentsMargins(16,16,16,16); hl_tab.setSpacing(10)

        hdr_h = QHBoxLayout()
        lbl_h = QLabel('Hatırlatıcılar')
        lbl_h.setFont(QFont('Segoe UI', 14, QFont.Bold)); hdr_h.addWidget(lbl_h); hdr_h.addStretch()
        btn_add_hat = make_btn('+ Ekle', C['success'], small=True)
        btn_add_hat.clicked.connect(self._add_hatirlatici); hdr_h.addWidget(btn_add_hat)
        hl_tab.addLayout(hdr_h)

        self.tbl_hat = make_table(
            ['ID','Başlık','Mesaj','Gün Tipi','Saat','Aktif'],
            [0, -1, 200, 100, 70, 65]
        )
        self.tbl_hat.hideColumn(0)
        self.tbl_hat.itemSelectionChanged.connect(self._on_hat_select)
        hl_tab.addWidget(self.tbl_hat)

        bot_h = QHBoxLayout()
        self.btn_toggle_hat = make_btn('Aç/Kapat', C['warning'], small=True)
        self.btn_toggle_hat.setEnabled(False)
        self.btn_toggle_hat.clicked.connect(self._toggle_hat)
        self.btn_del_hat = make_btn('Sil', C['danger'], small=True)
        self.btn_del_hat.setEnabled(False)
        self.btn_del_hat.clicked.connect(self._del_hat)
        bot_h.addWidget(self.btn_toggle_hat); bot_h.addWidget(self.btn_del_hat); bot_h.addStretch()
        hl_tab.addLayout(bot_h)
        tabs.addTab(tab_hat, '🔔 Hatırlatıcılar')

        # ── Tab 4: Hakkında ───────────────────────────────────────
        tab_hak = QWidget()
        hakl = QVBoxLayout(tab_hak); hakl.setContentsMargins(40,40,40,40); hakl.setSpacing(20)
        hakl.addStretch()
        ico = QLabel('🏋️'); ico.setFont(QFont('Segoe UI',52)); ico.setAlignment(Qt.AlignCenter)
        hakl.addWidget(ico)
        lbl_ad = QLabel('Fitness Takip Sistemi'); lbl_ad.setFont(QFont('Segoe UI',20,QFont.Bold))
        lbl_ad.setAlignment(Qt.AlignCenter); hakl.addWidget(lbl_ad)
        lbl_ver = QLabel('v1.0.0 — Dark Luxury Edition')
        lbl_ver.setFont(QFont('Segoe UI',12)); lbl_ver.setAlignment(Qt.AlignCenter)
        lbl_ver.setStyleSheet(f"color:{C['text_muted']};"); hakl.addWidget(lbl_ver)
        for bilgi in [
            '10 Tier — Tam Özellikli Fitness Platformu',
            'PyQt5 + SQLite3 | QPainter Charts',
            'Gamification | 1RM | Radar | Beslenme | Analiz',
        ]:
            lbl_b = QLabel(bilgi); lbl_b.setAlignment(Qt.AlignCenter)
            lbl_b.setStyleSheet(f"color:{C['text_secondary']};font-size:12px;")
            hakl.addWidget(lbl_b)
        hakl.addStretch()
        tabs.addTab(tab_hak, 'ℹ️ Hakkında')

        lay.addWidget(tabs)
        self._hat_selected_id = None

    def refresh(self):
        # Sistem özeti
        ozet = self.db.get_sistem_ozet()
        self.k_sporcu_s.set_value(str(ozet['sporcu']))
        self.k_log_s.set_value(str(ozet['log']))
        self.k_besin_s.set_value(str(ozet['besin']))
        self.k_db_boyut.set_value(f"{ozet['db_boyut']} KB")

        # Ayarlar
        birim = self.db.get_sistem_ayari('birim', 'kg')
        idx = self.cb_birim.findText(birim)
        if idx >= 0: self.cb_birim.setCurrentIndex(idx)

        # Export sporcu listesi
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.export_sporcu_cb.blockSignals(True); self.export_sporcu_cb.clear()
        self.export_sporcu_cb.addItem('-- Sporcu Seç --'); self._export_sporcu_ids = [None]
        for s in sporcular:
            self.export_sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._export_sporcu_ids.append(s['sporcu_id'])
        self.export_sporcu_cb.blockSignals(False)

        # Hatırlatıcılar
        self._refresh_hatirlaticilar()

    def _kaydet_ayarlar(self):
        self.db.set_sistem_ayari('birim', self.cb_birim.currentText(), 'Ağırlık birimi')
        self.db.set_sistem_ayari('dil', 'tr' if self.cb_dil.currentIndex()==0 else 'en', 'Dil')
        self.db.set_sistem_ayari('kalori_birimi', self.cb_kalori.currentText(), 'Kalori birimi')
        dark_msg(self,'Başarılı','Ayarlar kaydedildi! ✅')

    def _export_csv(self):
        from PyQt5.QtWidgets import QFileDialog
        idx = self.export_sporcu_cb.currentIndex()
        if idx == 0:
            dark_msg(self,'Uyarı','Sporcu seçin!',QMessageBox.Warning); return
        sporcu_id = self._export_sporcu_ids[idx]
        sporcu_adi = self.export_sporcu_cb.currentText().replace(' ','_')
        dosya, _ = QFileDialog.getSaveFileName(
            self, 'CSV Kaydet', f"fitness_{sporcu_adi}.csv", 'CSV (*.csv)')
        if dosya:
            if self.db.export_sporcu_csv(sporcu_id, dosya):
                dark_msg(self,'Başarılı',f'CSV dosyası kaydedildi:\n{dosya}')
            else:
                dark_msg(self,'Hata','Dışa aktarma başarısız.',QMessageBox.Warning)

    def _yedekle(self):
        from PyQt5.QtWidgets import QFileDialog
        dosya, _ = QFileDialog.getSaveFileName(
            self, 'Yedek Kaydet', f"fitness_yedek_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
            'SQLite DB (*.db)')
        if dosya:
            if self.db.yedekle(dosya):
                dark_msg(self,'Başarılı',f'Veritabanı yedeklendi:\n{dosya}')
            else:
                dark_msg(self,'Hata','Yedekleme başarısız.',QMessageBox.Warning)

    def _refresh_hatirlaticilar(self):
        with self.db.get_connection() as conn:
            rows = conn.execute("SELECT * FROM hatirlaticilar ORDER BY saat").fetchall()
            data = [dict(r) for r in rows]
        self.tbl_hat.setRowCount(0)
        for d in data:
            row = self.tbl_hat.rowCount(); self.tbl_hat.insertRow(row)
            aktif = '✅ Aktif' if d['aktif'] else '⏸ Pasif'
            vals = [str(d['hatirlatici_id']), d['baslik'],
                    d.get('mesaj','')[:40], d.get('gun_tipi','Günlük'),
                    d.get('saat','—'), aktif]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 5:
                    item.setForeground(QColor(C['success'] if d['aktif'] else C['text_muted']))
                self.tbl_hat.setItem(row, col, item)
        self._hat_selected_id = None
        self.btn_toggle_hat.setEnabled(False); self.btn_del_hat.setEnabled(False)

    def _on_hat_select(self):
        if self.tbl_hat.selectedItems():
            self._hat_selected_id = int(self.tbl_hat.item(self.tbl_hat.currentRow(),0).text())
            self.btn_toggle_hat.setEnabled(True); self.btn_del_hat.setEnabled(True)

    def _add_hatirlatici(self):
        dlg = HatirlaticiDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_hatirlatici(dlg.result_data)
            self._refresh_hatirlaticilar()

    def _toggle_hat(self):
        if self._hat_selected_id:
            self.db.toggle_hatirlatici(self._hat_selected_id)
            self._refresh_hatirlaticilar()

    def _del_hat(self):
        if self._hat_selected_id:
            if dark_confirm(self,'Sil','Bu hatırlatıcı silinecek. Emin misiniz?'):
                self.db.delete_hatirlatici(self._hat_selected_id)
                self._refresh_hatirlaticilar()

    def _hatirlatici_kontrol(self):
        """Her dakika çağrılır — saat gelmiş mi kontrol et."""
        simdi = datetime.now().strftime('%H:%M')
        hatirlaticilar = self.db.get_hatirlaticilar()
        for h in hatirlaticilar:
            if h.get('saat','') == simdi and h.get('son_gosterim','') != simdi:
                dark_msg(self, f"🔔 {h['baslik']}", h.get('mesaj',''))
                with self.db.get_connection() as conn:
                    conn.execute("UPDATE hatirlaticilar SET son_gosterim=? WHERE hatirlatici_id=?",
                                 (simdi, h['hatirlatici_id']))


class HatirlaticiDialog(BaseDialog):
    def __init__(self, db, parent=None):
        super().__init__('Hatırlatıcı Ekle', parent, 480)
        self.db = db; self._build()

    def _build(self):
        self.f_baslik  = self.add_row('Başlık *', self.inp('Hatırlatıcı başlığı'))
        self.f_mesaj   = self.add_row('Mesaj',    self.inp('Kısa mesaj'))
        self.f_gun     = self.add_row('Tekrar',   self.combo(['Gunluk','Haftalik','Tek Sefer']))
        self.f_saat    = self.add_row('Saat',     self.inp('09:00'))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Ekle', C['success'], self._save)

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        self.result_data = {
            'baslik':   self.f_baslik.text().strip(),
            'mesaj':    self.f_mesaj.text().strip(),
            'gun_tipi': self.f_gun.currentText(),
            'saat':     self.f_saat.text().strip() or '09:00',
        }
        self.accept()



class GelismisAnalizPage(QWidget):
    """Kas grubu volüm analizi + overtraining uyarıları + analiz notları."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self.selected_not_id    = None
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QTabWidget, QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Gelişmiş Analiz')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # Sporcu seç
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 220)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Süre:'))
        self.sure_cb = make_combo(['Son 30 Gün','Son 60 Gün','Son 90 Gün'], 150)
        self.sure_cb.currentTextChanged.connect(self._refresh_analiz)
        fil.addWidget(self.sure_cb); fil.addStretch()
        lay.addLayout(fil)

        # Tab widget
        tabs = QTabWidget()

        # ── Tab 1: Volüm Analizi ──────────────────────────────────
        tab_volum = QWidget()
        vl = QVBoxLayout(tab_volum); vl.setContentsMargins(12,12,12,12); vl.setSpacing(10)

        kpi_v = QHBoxLayout(); kpi_v.setSpacing(SPACING)
        self.k_toplam_ant  = KPICard('Toplam Antrenman', '—', '💪', C['primary'], C['primary_dark'])
        self.k_toplam_sure = KPICard('Toplam Süre',      '—', '⏱', C['info'],    C['info_dark'])
        self.k_toplam_kal  = KPICard('Toplam Kalori',    '—', '🔥', C['danger'],  C['danger_dark'])
        self.k_ort_zorluk  = KPICard('Ort. Zorluk',      '—', '⚡', C['warning'], C['warning_dark'])
        for k in [self.k_toplam_ant, self.k_toplam_sure, self.k_toplam_kal, self.k_ort_zorluk]:
            kpi_v.addWidget(k)
        vl.addLayout(kpi_v)

        splitter_v = QSplitter(Qt.Horizontal)
        splitter_v.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Volüm pie chart
        frame_pie = QFrame()
        frame_pie.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        fp_l = QVBoxLayout(frame_pie); fp_l.setContentsMargins(8,8,8,8)
        self.volum_pie = PieChartWidget()
        self.volum_pie.setMinimumHeight(260)
        fp_l.addWidget(self.volum_pie)
        splitter_v.addWidget(frame_pie)

        # Volüm bar chart
        frame_bar = QFrame()
        frame_bar.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        fb_l = QVBoxLayout(frame_bar); fb_l.setContentsMargins(8,8,8,8)
        self.volum_bar = BarChartWidget()
        self.volum_bar.setMinimumHeight(260)
        fb_l.addWidget(self.volum_bar)
        splitter_v.addWidget(frame_bar)
        splitter_v.setSizes([400, 500])
        vl.addWidget(splitter_v)

        # Detay tablo
        self.tbl_volum = make_table(
            ['Kategori','Antrenman','Süre (dk)','Kalori','Ort. Zorluk'],
            [-1, 110, 110, 110, 120]
        )
        self.tbl_volum.setMaximumHeight(180)
        vl.addWidget(self.tbl_volum)
        tabs.addTab(tab_volum, '📊 Volüm Analizi')

        # ── Tab 2: Overtraining Uyarıları ─────────────────────────
        tab_ovt = QWidget()
        ol = QVBoxLayout(tab_ovt); ol.setContentsMargins(12,12,12,12); ol.setSpacing(10)

        kpi_o = QHBoxLayout(); kpi_o.setSpacing(SPACING)
        self.k_haftalik   = KPICard('Bu Hafta', '—', '🗓', C['primary'], C['primary_dark'])
        self.k_art_arda   = KPICard('Art Arda', '—', '🔥', C['danger'],  C['danger_dark'])
        self.k_ort_zorluk2= KPICard('Ort. Zorluk','—','⚡', C['warning'], C['warning_dark'])
        for k in [self.k_haftalik, self.k_art_arda, self.k_ort_zorluk2]:
            kpi_o.addWidget(k)
        ol.addLayout(kpi_o)

        # Uyarı kartları scroll alanı
        scroll_o = QScrollArea(); scroll_o.setWidgetResizable(True)
        scroll_o.setStyleSheet(f"QScrollArea{{border:none;background:{C['bg_main']};}}")
        self.uyari_widget = QWidget()
        self.uyari_widget.setStyleSheet(f"background:{C['bg_main']};")
        self.uyari_lay = QVBoxLayout(self.uyari_widget)
        self.uyari_lay.setSpacing(8); self.uyari_lay.setContentsMargins(0,0,0,0)
        scroll_o.setWidget(self.uyari_widget)
        ol.addWidget(scroll_o)
        tabs.addTab(tab_ovt, '🚨 Overtraining')

        # ── Tab 3: Analiz Notları ──────────────────────────────────
        tab_notlar = QWidget()
        nl = QVBoxLayout(tab_notlar); nl.setContentsMargins(12,12,12,12); nl.setSpacing(10)

        nh = QHBoxLayout()
        lbl_n = QLabel('Antrenör Notları')
        lbl_n.setFont(QFont('Segoe UI', 14, QFont.Bold)); nh.addWidget(lbl_n)
        nh.addStretch()
        btn_not = make_btn('+ Not Ekle', C['success'], small=True)
        btn_not.clicked.connect(self._add_not); nh.addWidget(btn_not)
        nl.addLayout(nh)

        self.tbl_notlar = make_table(
            ['ID','Sporcu','Tip','Başlık','Öncelik','Tarih'],
            [0, 150, 90, -1, 90, 100]
        )
        self.tbl_notlar.hideColumn(0)
        self.tbl_notlar.itemSelectionChanged.connect(self._on_not_select)
        nl.addWidget(self.tbl_notlar)

        # Not detay alanı
        self.txt_not_detay = QTextEdit()
        self.txt_not_detay.setReadOnly(True)
        self.txt_not_detay.setMaximumHeight(100)
        self.txt_not_detay.setStyleSheet(f"""
            QTextEdit{{background:{C['bg_card']};border:1px solid {C['border']};
            border-radius:8px;padding:8px;color:{C['text_main']};font-size:13px;}}
        """)
        nl.addWidget(self.txt_not_detay)

        bot_n = QHBoxLayout()
        self.btn_del_not = make_btn('Sil', C['danger'], small=True)
        self.btn_del_not.setEnabled(False); self.btn_del_not.clicked.connect(self._del_not)
        bot_n.addWidget(self.btn_del_not); bot_n.addStretch()
        nl.addLayout(bot_n)
        tabs.addTab(tab_notlar, '📝 Analiz Notları')

        lay.addWidget(tabs)
        self._sporcu_ids = [None]

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self._refresh_analiz()
        self._refresh_notlar()

    def _refresh_analiz(self):
        self._load_sporcular()
        if not self.selected_sporcu_id: return
        gun_map = {'Son 30 Gün':30, 'Son 60 Gün':60, 'Son 90 Gün':90}
        gun = gun_map.get(self.sure_cb.currentText(), 30)
        data = self.db.get_kas_grubu_volum(self.selected_sporcu_id, gun)

        # KPI
        toplam_ant  = sum(d['antrenman_sayisi'] for d in data)
        toplam_sure = sum(d['toplam_sure'] for d in data)
        toplam_kal  = sum(d['toplam_kalori'] for d in data)
        ort_zorluk  = sum(d['ort_zorluk']*d['antrenman_sayisi'] for d in data) / max(toplam_ant, 1)

        self.k_toplam_ant.set_value(str(toplam_ant))
        self.k_toplam_sure.set_value(f"{toplam_sure} dk")
        self.k_toplam_kal.set_value(f"{int(toplam_kal):,}")
        self.k_ort_zorluk.set_value(f"{ort_zorluk:.1f}/10")

        # Pie ve bar
        pie_data = [(d['kategori'], d['antrenman_sayisi']) for d in data]
        self.volum_pie.set_data(pie_data, 'Kategori Dağılımı')
        bar_data = [(d['kategori'][:7], d['toplam_kalori']) for d in data]
        self.volum_bar.set_data(bar_data, 'Kategoriye Göre Kalori')

        # Tablo
        self.tbl_volum.setRowCount(0)
        kat_renk = {'Kardiyo':C['info'],'Kuvvet':C['danger'],'Esneklik':C['success'],
                    'HIIT':C['warning'],'Denge':C['primary']}
        for d in data:
            row = self.tbl_volum.rowCount(); self.tbl_volum.insertRow(row)
            vals = [d['kategori'], str(d['antrenman_sayisi']),
                    str(int(d['toplam_sure'])), str(int(d['toplam_kalori'])),
                    f"{d['ort_zorluk']:.1f}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 0:
                    item.setForeground(QColor(kat_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_volum.setItem(row, col, item)

        # Overtraining
        self._refresh_overtraining()

    def _refresh_overtraining(self):
        if not self.selected_sporcu_id: return
        ozet = self.db.get_overtraining_uyari(self.selected_sporcu_id)
        self.k_haftalik.set_value(f"{ozet['haftalik']}/7")
        self.k_art_arda.set_value(str(ozet['art_arda']))
        self.k_ort_zorluk2.set_value(f"{ozet['ort_zorluk']}/10")

        # Uyarı kartlarını temizle
        while self.uyari_lay.count():
            item = self.uyari_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        seviye_renk = {
            'Kritik': C['danger'], 'Uyarı': C['warning'],
            'Bilgi': C['info'], 'İyi': C['success']
        }
        for u in ozet['uyarilar']:
            frame = QFrame()
            renk = seviye_renk.get(u['seviye'], C['primary'])
            frame.setStyleSheet(f"""
                QFrame{{background:{C['bg_card']};border-radius:10px;
                border-left:4px solid {renk};margin:2px 0;}}
            """)
            fl = QVBoxLayout(frame); fl.setContentsMargins(14,10,14,10); fl.setSpacing(4)
            hdr_l = QHBoxLayout()
            lbl_sev = QLabel(f"{u['ikon']} {u['seviye']}")
            lbl_sev.setFont(QFont('Segoe UI', 12, QFont.Bold))
            lbl_sev.setStyleSheet(f"color:{renk};")
            hdr_l.addWidget(lbl_sev); hdr_l.addStretch()
            fl.addLayout(hdr_l)
            lbl_msg = QLabel(u['mesaj'])
            lbl_msg.setFont(QFont('Segoe UI', 12))
            lbl_msg.setStyleSheet(f"color:{C['text_main']};")
            fl.addWidget(lbl_msg)
            lbl_oneri = QLabel(f"💡 {u['oneri']}")
            lbl_oneri.setFont(QFont('Segoe UI', 11))
            lbl_oneri.setStyleSheet(f"color:{C['text_secondary']};")
            fl.addWidget(lbl_oneri)
            self.uyari_lay.addWidget(frame)
        self.uyari_lay.addStretch()

    def _refresh_notlar(self):
        data = self.db.get_analiz_notlari(self.selected_sporcu_id)
        self.tbl_notlar.setRowCount(0)
        oncelik_renk = {'Kritik':C['danger'],'Yüksek':C['warning'],
                        'Normal':C['text_main'],'Düşük':C['text_muted']}
        tip_renk = {'Uyarı':C['danger'],'Öneri':C['primary'],
                    'Gözlem':C['success'],'Bilgi':C['info']}
        for d in data:
            row = self.tbl_notlar.rowCount(); self.tbl_notlar.insertRow(row)
            vals = [str(d['not_id']), d.get('sporcu_adi','—'),
                    d['not_tipi'], d['baslik'], d['oncelik'], d['tarih']]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 2:
                    item.setForeground(QColor(tip_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 4:
                    item.setForeground(QColor(oncelik_renk.get(val, C['text_main'])))
                self.tbl_notlar.setItem(row, col, item)
        self.selected_not_id = None; self.btn_del_not.setEnabled(False)

    def _on_not_select(self):
        if self.tbl_notlar.selectedItems():
            self.selected_not_id = int(self.tbl_notlar.item(self.tbl_notlar.currentRow(), 0).text())
            # Detay göster
            data = self.db.get_analiz_notlari(self.selected_sporcu_id)
            for d in data:
                if d['not_id'] == self.selected_not_id:
                    self.txt_not_detay.setPlainText(d.get('icerik',''))
                    break
            self.btn_del_not.setEnabled(True)

    def _add_not(self):
        from PyQt5.QtWidgets import QInputDialog
        if not self.selected_sporcu_id:
            dark_msg(self,'Uyarı','Sporcu seçin!',QMessageBox.Warning); return
        dlg = AnalizNotuDialog(self.db, self.selected_sporcu_id, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_analiz_notu(dlg.result_data)
            dark_msg(self,'Başarılı','Not eklendi!')
            self._refresh_notlar()

    def _del_not(self):
        if not self.selected_not_id: return
        if dark_confirm(self,'Sil','Bu not silinecek. Emin misiniz?'):
            self.db.delete_analiz_notu(self.selected_not_id)
            self.txt_not_detay.clear(); self._refresh_notlar()

    def refresh(self):
        self._load_sporcular(); self._refresh_analiz(); self._refresh_notlar()


class AnalizNotuDialog(BaseDialog):
    def __init__(self, db, sporcu_id, parent=None):
        super().__init__('Analiz Notu Ekle', parent, 540)
        self.db = db; self.sporcu_id = sporcu_id; self._build()

    def _build(self):
        self.f_tip = self.add_row('Not Tipi', self.combo(['Gözlem','Öneri','Uyarı','Bilgi']))
        self.f_baslik = self.add_row('Başlık *', self.inp('Not başlığı...'))
        self.f_oncelik = self.add_row('Öncelik', self.combo(['Normal','Yüksek','Kritik','Düşük']))
        self.f_icerik = self.add_row('İçerik', self.txt('Detaylı açıklama...', h=100))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        if not self.f_baslik.text().strip():
            dark_msg(self,'Hata','Başlık zorunludur!',QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id': self.sporcu_id,
            'not_tipi':  self.f_tip.currentText(),
            'baslik':    self.f_baslik.text().strip(),
            'oncelik':   self.f_oncelik.currentText(),
            'icerik':    self.f_icerik.toPlainText().strip(),
        }
        self.accept()



class RadarChartWidget(QWidget):
    """QPainter spider/radar chart — 6 eksen."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.labels = ['Güç','Kardiyo','Esneklik','Denge','Dayanıklılık','Hız']
        self.values = [0, 0, 0, 0, 0, 0]
        self.compare = None   # İkinci seri (opsiyonel)
        self.setMinimumHeight(300)

    def set_data(self, values, compare=None):
        self.values  = values[:6] if values else [0]*6
        self.compare = compare
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        n = len(self.labels)
        cx, cy = w // 2, h // 2
        r_max = min(cx, cy) - 52

        # Ağ çizgileri (5 halka)
        for ring in range(1, 6):
            r = r_max * ring / 5
            points = []
            for i in range(n):
                angle = math.radians(90 + i * 360 / n)
                x = cx + r * math.cos(angle)
                y = cy - r * math.sin(angle)
                points.append(QPointF(x, y))
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))
            p.setBrush(Qt.NoBrush)
            path = QPainterPath()
            path.moveTo(points[0])
            for pt in points[1:]: path.lineTo(pt)
            path.closeSubpath(); p.drawPath(path)
            # Yüzde etiketi
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 8))
            p.drawText(QRectF(cx+3, cy - r - 10, 28, 14),
                       Qt.AlignLeft, f"{ring*20}%")

        # Eksen çizgileri + etiketler
        for i, lbl in enumerate(self.labels):
            angle = math.radians(90 + i * 360 / n)
            ex = cx + r_max * math.cos(angle)
            ey = cy - r_max * math.sin(angle)
            p.setPen(QPen(QColor(C['border_light']), 1))
            p.drawLine(QPointF(cx, cy), QPointF(ex, ey))
            # Etiket konumu
            lx = cx + (r_max + 18) * math.cos(angle)
            ly = cy - (r_max + 18) * math.sin(angle)
            p.setPen(QColor(C['text_secondary']))
            p.setFont(QFont('Segoe UI', 10, QFont.Bold))
            p.drawText(QRectF(lx - 40, ly - 10, 80, 20), Qt.AlignCenter, lbl)

        def polygon(vals, color, alpha=160):
            pts = []
            for i, v in enumerate(vals):
                angle = math.radians(90 + i * 360 / n)
                r = r_max * min(v, 100) / 100
                pts.append(QPointF(cx + r * math.cos(angle),
                                   cy - r * math.sin(angle)))
            path = QPainterPath()
            path.moveTo(pts[0])
            for pt in pts[1:]: path.lineTo(pt)
            path.closeSubpath()
            fill = QColor(color); fill.setAlpha(alpha // 3)
            p.setBrush(QBrush(fill))
            p.setPen(QPen(QColor(color), 2))
            p.drawPath(path)
            # Noktalar
            for pt in pts:
                p.setBrush(QColor(C['bg_card'])); p.setPen(QPen(QColor(color), 2))
                p.drawEllipse(pt, 4, 4)
                p.setBrush(QColor(color)); p.setPen(Qt.NoPen)
                p.drawEllipse(pt, 2.5, 2.5)

        # Karşılaştırma serisi
        if self.compare:
            polygon(self.compare, C['text_muted'], 80)

        # Ana seri
        polygon(self.values, C['primary'], 160)

        # Merkez nokta
        p.setBrush(QColor(C['primary'])); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(cx, cy), 3, 3)

        # Legend
        p.setBrush(QColor(C['primary'])); p.setPen(Qt.NoPen)
        p.drawRoundedRect(12, 12, 12, 12, 3, 3)
        p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 9))
        p.drawText(28, 12, 80, 14, Qt.AlignVCenter, 'Sporcu')
        if self.compare:
            p.setBrush(QColor(C['text_muted'])); p.setPen(Qt.NoPen)
            p.drawRoundedRect(12, 28, 12, 12, 3, 3)
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(28, 28, 80, 14, Qt.AlignVCenter, 'Karşılaştırma')

        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 7 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

class BiRMDialog(BaseDialog):
    """1RM hesaplama dialog'u."""
    def __init__(self, db, parent=None):
        super().__init__('1RM Hesapla & Kaydet', parent, 560)
        self.db = db; self._build()

    def _build(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(
            ['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)

        self.f_egzersiz = self.combo([
            'Bench Press','Squat','Deadlift','Overhead Press',
            'Barbell Row','Pull-up','Dip','Leg Press','Diğer'])
        self.add_row('Egzersiz *', self.f_egzersiz)

        self.f_agirlik = self.dspin(1, 500, 60, 1)
        self.f_agirlik.setSuffix(' kg')
        self.f_agirlik.valueChanged.connect(self._hesapla)
        self.add_row('Ağırlık *', self.f_agirlik)

        self.f_tekrar = self.spin(1, 30, 5)
        self.f_tekrar.valueChanged.connect(self._hesapla)
        self.add_row('Tekrar *', self.f_tekrar)

        self.f_formul = self.combo(['Epley', 'Brzycki', 'Lander'])
        self.f_formul.currentTextChanged.connect(self._hesapla)
        self.add_row('Formül', self.f_formul)

        # Sonuç alanı
        self.lbl_sonuc = QLabel('— Değer girin —')
        self.lbl_sonuc.setFont(QFont('Segoe UI', 20, QFont.Bold))
        self.lbl_sonuc.setStyleSheet(f"color:{C['primary_light']};")
        self.lbl_sonuc.setAlignment(Qt.AlignCenter)
        self._main_layout.addWidget(self.lbl_sonuc)

        # Formül açıklaması
        self.lbl_formul = QLabel('')
        self.lbl_formul.setStyleSheet(f"color:{C['text_muted']};font-size:11px;")
        self.lbl_formul.setAlignment(Qt.AlignCenter)
        self._main_layout.addWidget(self.lbl_formul)

        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)
        self._hesapla()

    def _hesapla(self):
        ag = self.f_agirlik.value()
        rep = self.f_tekrar.value()
        formul = self.f_formul.currentText()
        bir_rm = self.db.hesapla_1rm(ag, rep, formul)
        self.lbl_sonuc.setText(f"1RM = {bir_rm} kg")
        formul_aciklama = {
            'Epley':   f"{ag} × (1 + {rep}/30)",
            'Brzycki': f"{ag} × (36 / (37 - {rep}))",
            'Lander':  f"(100 × {ag}) / (101.3 - 2.67 × {rep})",
        }
        self.lbl_formul.setText(formul_aciklama.get(formul, ''))

    def _save(self):
        if self.f_sporcu.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Sporcu seçin!', QMessageBox.Warning); return
        ag = self.f_agirlik.value()
        rep = self.f_tekrar.value()
        self.result_data = {
            'sporcu_id': self._sporcu_ids[self.f_sporcu.currentIndex()],
            'egzersiz':  self.f_egzersiz.currentText(),
            'agirlik':   ag, 'tekrar': rep,
            'formul':    self.f_formul.currentText(),
        }
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 7 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class BiRMPage(QWidget):
    """1RM Hesaplama — temiz layout."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self._sporcu_ids = [None]
        self._selected_egz = 'Tümü'
        self._build(); self._load_sporcular()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # ── Header ──
        hdr = QHBoxLayout()
        title = QLabel('1RM Hesaplama')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_new = make_btn('+ 1RM Hesapla', C['success'], small=True)
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # ── Formül banner — sabit yükseklik ──
        banner = QFrame()
        banner.setFixedHeight(36)
        banner.setStyleSheet(f"background:{C['bg_card']};border-radius:8px;border-left:3px solid {C['primary']};")
        bl = QHBoxLayout(banner); bl.setContentsMargins(14,0,14,0)
        lbl_f = QLabel("📐  Epley: w×(1+r/30)   |   Brzycki: w×36/(37-r)   |   Lander: 100w/(101.3-2.67r)")
        lbl_f.setFont(QFont('Segoe UI', 11))
        lbl_f.setStyleSheet(f"color:{C['text_secondary']};")
        bl.addWidget(lbl_f)
        lay.addWidget(banner)

        # ── Filtreler — sabit yükseklik ──
        fil_w = QWidget(); fil_w.setFixedHeight(INPUT_H + 8)
        fil = QHBoxLayout(fil_w); fil.setContentsMargins(0,0,0,0); fil.setSpacing(10)
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 200)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Egzersiz:'))
        self.egz_cb = make_combo(['Tümü','Bench Press','Squat','Deadlift','Overhead Press','Barbell Row'], 160)
        self.egz_cb.currentTextChanged.connect(self._on_egz)
        fil.addWidget(self.egz_cb); fil.addStretch()
        lay.addWidget(fil_w)

        # ── Alt bölüm: sol (rekorlar + trend) | sağ (tüm kayıtlar) ──
        from PyQt5.QtWidgets import QSplitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")
        splitter.setChildrenCollapsible(False)

        # Sol panel
        left = QWidget()
        ll = QVBoxLayout(left); ll.setContentsMargins(0,0,8,0); ll.setSpacing(8)

        lbl_best = QLabel('En İyi 1RM Rekorları')
        lbl_best.setFont(QFont('Segoe UI', 12, QFont.Bold)); ll.addWidget(lbl_best)

        self.tbl_best = make_table(
            ['Egzersiz', '1RM (kg)', 'Ağırlık', 'Tekrar', 'Tarih'],
            [-1, 90, 85, 65, 95]
        )
        self.tbl_best.setFixedHeight(200)
        self.tbl_best.itemSelectionChanged.connect(self._on_best_select)
        ll.addWidget(self.tbl_best)

        lbl_trend = QLabel('Trend Grafiği')
        lbl_trend.setFont(QFont('Segoe UI', 12, QFont.Bold)); ll.addWidget(lbl_trend)

        frame_t = QFrame()
        frame_t.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        ft_l = QVBoxLayout(frame_t); ft_l.setContentsMargins(8,8,8,8)
        self.chart_trend = LineChartWidget()
        ft_l.addWidget(self.chart_trend)
        ll.addWidget(frame_t, 1)
        splitter.addWidget(left)

        # Sağ panel
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(8,0,0,0); rl.setSpacing(8)
        lbl_all = QLabel('Tüm Kayıtlar')
        lbl_all.setFont(QFont('Segoe UI', 12, QFont.Bold)); rl.addWidget(lbl_all)
        self.tbl_all = make_table(
            ['ID','Sporcu','Egzersiz','Ağırlık','Tekrar','1RM','Formül','Tarih'],
            [0, 120, -1, 75, 65, 90, 75, 90]
        )
        self.tbl_all.hideColumn(0)
        rl.addWidget(self.tbl_all, 1)
        splitter.addWidget(right)
        splitter.setSizes([460, 540])
        lay.addWidget(splitter, 1)

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear(); self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self.refresh()

    def _on_egz(self):
        self._selected_egz = self.egz_cb.currentText()
        self.refresh()
        if self.selected_sporcu_id and self._selected_egz != 'Tümü':
            trend = self.db.get_1rm_trend(self.selected_sporcu_id, self._selected_egz)
            self.chart_trend.set_data(trend, f"{self._selected_egz} — 1RM Trend", 'kg', C['danger'])

    def _on_best_select(self):
        if self.tbl_best.selectedItems() and self.selected_sporcu_id:
            egz = self.tbl_best.item(self.tbl_best.currentRow(), 0).text()
            trend = self.db.get_1rm_trend(self.selected_sporcu_id, egz)
            self.chart_trend.set_data(trend, f"{egz} — 1RM Trend", 'kg', C['danger'])

    def refresh(self):
        self._load_sporcular()
        egz_filtre = self._selected_egz if self._selected_egz != 'Tümü' else None

        egz_renk = {
            'Bench Press': C['danger'], 'Squat': C['primary'],
            'Deadlift': C['warning'], 'Overhead Press': C['success'],
            'Barbell Row': C['info'],
        }

        # Tüm kayıtlar
        data = self.db.get_1rm_kayitlari(self.selected_sporcu_id, egz_filtre)
        self.tbl_all.setRowCount(0)
        for d in data:
            row = self.tbl_all.rowCount(); self.tbl_all.insertRow(row)
            vals = [str(d['rm_id']), d.get('sporcu_adi','—'), d['egzersiz_adi'],
                    f"{d['agirlik']} kg", str(d['tekrar']),
                    f"{d['hesaplanan_1rm']} kg", d.get('formul','Epley'), d['tarih']]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 12))
                if col == 2:
                    item.setForeground(QColor(egz_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 12, QFont.Bold))
                if col == 5:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 12, QFont.Bold))
                self.tbl_all.setItem(row, col, item)

        # En iyi rekorlar
        self.tbl_best.setRowCount(0)
        if self.selected_sporcu_id:
            for d in self.db.get_1rm_en_iyi(self.selected_sporcu_id):
                row = self.tbl_best.rowCount(); self.tbl_best.insertRow(row)
                vals = [d['egzersiz_adi'], f"{d['max_1rm']} kg",
                        f"{d['agirlik']} kg", str(d['tekrar']), d['tarih']]
                for col, val in enumerate(vals):
                    item = QTableWidgetItem(val); item.setFont(QFont('Segoe UI', 12))
                    if col == 1:
                        item.setForeground(QColor(C['accent']))
                        item.setFont(QFont('Segoe UI', 12, QFont.Bold))
                    self.tbl_best.setItem(row, col, item)

    def _add(self):
        dlg = BiRMDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            d = dlg.result_data
            bir_rm = self.db.add_1rm_kaydi(
                d['sporcu_id'], d['egzersiz'], d['agirlik'], d['tekrar'], d['formul'])
            dark_msg(self, 'Hesaplandı! 💪', f"{d['egzersiz']} için 1RM: {bir_rm} kg")
            for i, s_id in enumerate(self._sporcu_ids):
                if s_id == d['sporcu_id']:
                    self.sporcu_cb.setCurrentIndex(i); break
            self.refresh()


class RadarAnalizPage(QWidget):
    """Sporcu radar analizi — güç, kardiyo, esneklik, denge, dayanıklılık, hız."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id  = None
        self.selected_sporcu2_id = None
        self._sporcu_ids  = [None]
        self._sporcu2_ids = [None]
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Radar Analizi')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_hesapla = make_btn('⚡ Otomatik Hesapla', C['primary'], small=True)
        btn_hesapla.clicked.connect(self._hesapla_otomatik)
        hdr.addWidget(btn_hesapla)
        lay.addLayout(hdr)

        # Sporcu seç — sabit yükseklik
        fil_container = QWidget(); fil_container.setFixedHeight(INPUT_H + 8)
        fil = QHBoxLayout(fil_container)
        fil.setContentsMargins(0,0,0,0); fil.setSpacing(10)
        fil.addWidget(QLabel('Sporcu 1:'))
        self.sporcu_cb = make_combo(['-- Sporcu 1 --'], 200)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Karşılaştır:'))
        self.sporcu2_cb = make_combo(['-- Sporcu 2 (opsiyonel) --'], 220)
        self.sporcu2_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu2_cb); fil.addStretch()
        lay.addWidget(fil_container)

        # Ana içerik: sol radar + sağ barlar — QSplitter ile
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")
        splitter.setChildrenCollapsible(False)

        # Sol: Radar chart — kare, büyümeyi engelle
        left = QFrame()
        left.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ll = QVBoxLayout(left); ll.setContentsMargins(10,10,10,10)
        self.radar = RadarChartWidget()
        self.radar.setMinimumSize(300, 300)
        self.radar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        ll.addWidget(self.radar)
        splitter.addWidget(left)

        # Sağ: 6 bar — scroll ile, sabit genişlik
        right = QFrame()
        right.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        right.setMinimumWidth(320); right.setMaximumWidth(480)
        rl = QVBoxLayout(right); rl.setContentsMargins(16,14,16,14); rl.setSpacing(8)

        lbl_det = QLabel('Performans Detayı')
        lbl_det.setFont(QFont('Segoe UI', 13, QFont.Bold))
        rl.addWidget(lbl_det)

        self.bar_labels = {}  # key → (lbl_widget, bar_widget)
        eksenler = [
            ('guc_puani',          'Güç',         C['danger']),
            ('kardiyo_puani',      'Kardiyo',      C['info']),
            ('esneklik_puani',     'Esneklik',     C['success']),
            ('denge_puani',        'Denge',        C['primary']),
            ('dayaniklilik_puani', 'Dayanıklılık', C['warning']),
            ('hiz_puani',          'Hız',          '#ec4899'),
        ]
        for key, ad, renk in eksenler:
            row_w = QWidget(); row_w.setFixedHeight(52)
            row_l = QVBoxLayout(row_w); row_l.setContentsMargins(0,2,0,2); row_l.setSpacing(2)

            # Başlık satırı
            hrow = QHBoxLayout(); hrow.setContentsMargins(0,0,0,0)
            lbl_ad = QLabel(ad)
            lbl_ad.setFont(QFont('Segoe UI', 11, QFont.Bold))
            lbl_ad.setStyleSheet(f"color:{C['text_main']};")
            hrow.addWidget(lbl_ad)
            hrow.addStretch()
            lbl_val = QLabel('0 / 100  (0%)')
            lbl_val.setFont(QFont('Segoe UI', 10))
            lbl_val.setStyleSheet(f"color:{renk};")
            hrow.addWidget(lbl_val)
            row_l.addLayout(hrow)

            # Progress bar (QPainter değil, QFrame tabanlı)
            bar_bg = QFrame(); bar_bg.setFixedHeight(12)
            bar_bg.setStyleSheet(f"background:{C['border']};border-radius:6px;")
            bar_container = QWidget(); bar_container.setFixedHeight(12)
            bar_cl = QHBoxLayout(bar_container); bar_cl.setContentsMargins(0,0,0,0)
            bar_fill = QFrame(); bar_fill.setFixedHeight(12)
            bar_fill.setStyleSheet(f"background:{renk};border-radius:6px;")
            bar_fill.setFixedWidth(0)
            bar_cl.addWidget(bar_fill); bar_cl.addStretch()
            row_l.addWidget(bar_bg)
            # overlay bar_fill üstünde — geometry kullan
            bar_fill.setParent(bar_bg)
            bar_fill.setGeometry(0, 0, 0, 12)

            rl.addWidget(row_w)
            self.bar_labels[key] = (lbl_val, bar_fill, bar_bg, renk)

        rl.addStretch()
        lbl_info = QLabel('💡 "Otomatik Hesapla" antrenman verilerinden puanları hesaplar.')
        lbl_info.setStyleSheet(f"color:{C['text_muted']};font-size:10px;")
        lbl_info.setWordWrap(True); rl.addWidget(lbl_info)

        splitter.addWidget(right)
        splitter.setSizes([600, 380])
        lay.addWidget(splitter, 1)

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        for cb, ids_attr, ilk in [
            (self.sporcu_cb,  '_sporcu_ids',  '-- Sporcu 1 --'),
            (self.sporcu2_cb, '_sporcu2_ids', '-- Sporcu 2 (opsiyonel) --'),
        ]:
            cb.blockSignals(True); cb.clear(); cb.addItem(ilk)
            ids = [None]
            for s in sporcular:
                cb.addItem(f"{s['ad']} {s['soyad']}")
                ids.append(s['sporcu_id'])
            setattr(self, ids_attr, ids)
            cb.blockSignals(False)

    def _on_sporcu(self):
        idx1 = self.sporcu_cb.currentIndex()
        idx2 = self.sporcu2_cb.currentIndex()
        self.selected_sporcu_id  = self._sporcu_ids[idx1]  if idx1 > 0 else None
        self.selected_sporcu2_id = self._sporcu2_ids[idx2] if idx2 > 0 else None
        self._refresh_data()

    def _hesapla_otomatik(self):
        if not self.selected_sporcu_id:
            dark_msg(self,'Uyarı','Sporcu seçin!',QMessageBox.Warning); return
        self.db.hesapla_radar_otomatik(self.selected_sporcu_id)
        dark_msg(self,'Hesaplandı','Radar değerleri güncellendi! ✅')
        self._refresh_data()

    def _refresh_data(self):
        KEYS = ['guc_puani','kardiyo_puani','esneklik_puani',
                'denge_puani','dayaniklilik_puani','hiz_puani']

        if not self.selected_sporcu_id:
            self.radar.set_data([0]*6)
            for lbl_val, bar_fill, bar_bg, renk in self.bar_labels.values():
                lbl_val.setText('0 / 100  (0%)'); bar_fill.setFixedWidth(0)
            return

        d = self.db.get_radar_degerleri(self.selected_sporcu_id) or {}
        vals = [float(d.get(k, 0)) for k in KEYS]

        comp = None
        if self.selected_sporcu2_id:
            d2 = self.db.get_radar_degerleri(self.selected_sporcu2_id) or {}
            comp = [float(d2.get(k, 0)) for k in KEYS]

        self.radar.set_data(vals, comp)

        # Barları güncelle — geometry resize ile
        for i, key in enumerate(KEYS):
            lbl_val, bar_fill, bar_bg, renk = self.bar_labels[key]
            v = float(d.get(key, 0))
            pct = min(1.0, v / 100.0)
            lbl_val.setText(f"{v:.0f} / 100  ({int(pct*100)}%)")
            # bar_bg genişliği sonraki paint'te hazır — QTimer ile al
            QTimer.singleShot(50, lambda bf=bar_fill, bg=bar_bg, p=pct:
                bf.setFixedWidth(int(bg.width() * p)))

    def refresh(self):
        self._load_sporcular()
        self._refresh_data()



class RozetKartWidget(QWidget):
    """Tek rozet görüntüleme — kazanıldı/kazanılmadı durumu."""
    def __init__(self, rozet_adi, icon, puan, aciklama,
                 kazanildi=False, tarih='', parent=None):
        super().__init__(parent)
        self.rozet_adi  = rozet_adi
        self.icon       = icon
        self.puan       = puan
        self.aciklama   = aciklama
        self.kazanildi  = kazanildi
        self.tarih      = tarih[:10] if tarih else ''
        self.setFixedSize(160, 140)

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()

        # Arka plan
        if self.kazanildi:
            grad = QLinearGradient(0, 0, w, h)
            grad.setColorAt(0, QColor('#1e1b4b'))
            grad.setColorAt(1, QColor('#312e81'))
            p.fillPath(_rounded(0, 0, w, h, 14), QBrush(grad))
            border_col = QColor(C['primary_light'])
        else:
            p.fillPath(_rounded(0, 0, w, h, 14), QBrush(QColor(C['bg_secondary'])))
            border_col = QColor(C['border'])

        # Border
        p.setPen(QPen(border_col, 1.5)); p.setBrush(Qt.NoBrush)
        p.drawRoundedRect(QRectF(1, 1, w-2, h-2), 13, 13)

        # Kilit overlay (kazanılmadıysa)
        if not self.kazanildi:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 28))
            p.drawText(QRectF(0, 14, w, 46), Qt.AlignCenter, '🔒')
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 10, QFont.Bold))
            p.drawText(QRectF(6, 64, w-12, 20), Qt.AlignCenter, self.rozet_adi)
            p.setFont(QFont('Segoe UI', 8))
            p.drawText(QRectF(6, 82, w-12, 30), Qt.AlignCenter | Qt.TextWordWrap, self.aciklama)
            p.setPen(QColor(C['border_light'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(6, h-22, w-12, 18), Qt.AlignCenter, f"+{self.puan} puan")
            p.end(); return

        # Kazanıldı — ikon
        p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI', 30))
        p.drawText(QRectF(0, 8, w, 48), Qt.AlignCenter, self.icon)

        # Parlak halo
        _c = QColor(C['primary_light']); _c.setAlpha(30); p.setBrush(_c)
        p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(w//2, 32), 28, 28)

        # İsim
        p.setPen(QColor(255, 255, 255))
        p.setFont(QFont('Segoe UI', 10, QFont.Bold))
        p.drawText(QRectF(6, 60, w-12, 20), Qt.AlignCenter, self.rozet_adi)

        # Puan badge
        p.setBrush(QColor(C['primary'])); p.setPen(Qt.NoPen)
        p.drawRoundedRect(QRectF(w//2-22, 84, 44, 18), 9, 9)
        p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
        p.drawText(QRectF(w//2-22, 84, 44, 18), Qt.AlignCenter, f"+{self.puan}p")

        # Tarih
        p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 8))
        p.drawText(QRectF(6, h-18, w-12, 16), Qt.AlignCenter, self.tarih)
        p.end()


def _rounded(x, y, w, h, r):
    path = QPainterPath(); path.addRoundedRect(QRectF(x,y,w,h), r, r); return path


class SeviyeBarWidget(QWidget):
    """Sporcu seviye widget — saf Qt layout, QPainter yok."""
    SEV_RENK = {1:'#6b7280',2:'#92400e',3:'#9ca3af',
                4:'#f59e0b',5:'#818cf8',6:'#06b6d4'}
    ESIKLER  = [0, 100, 300, 600, 1000, 2000, 9999]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(68)
        self.setStyleSheet("background:transparent;")

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 6, 12, 6)
        root.setSpacing(5)

        # Satır 1: seviye adı sol + puan sağ
        row1 = QHBoxLayout(); row1.setContentsMargins(0,0,0,0); row1.setSpacing(0)
        self.lbl_seviye = QLabel("⚡ Başlangıç")
        self.lbl_seviye.setFont(QFont('Segoe UI', 12, QFont.Bold))
        self.lbl_seviye.setStyleSheet("color:#6b7280;background:transparent;border:none;")
        self.lbl_puan = QLabel("0 puan")
        self.lbl_puan.setFont(QFont('Segoe UI', 11))
        self.lbl_puan.setStyleSheet(
            "color:" + C['text_secondary'] + ";background:transparent;border:none;")
        row1.addWidget(self.lbl_seviye, 1)
        row1.addWidget(self.lbl_puan, 0, Qt.AlignRight | Qt.AlignVCenter)
        root.addLayout(row1)

        # Satır 2: bar arka plan
        self.bar_bg = QFrame()
        self.bar_bg.setFixedHeight(10)
        self.bar_bg.setStyleSheet(
            "background:" + C['border'] + ";border-radius:5px;border:none;")
        # doluluk
        self.bar_fill = QFrame(self.bar_bg)
        self.bar_fill.setGeometry(0, 0, 0, 10)
        self.bar_fill.setStyleSheet(
            "background:#6b7280;border-radius:5px;border:none;")
        root.addWidget(self.bar_bg)

        # Satır 3: kalan puan sağda
        self.lbl_kalan = QLabel("Sonraki seviye: 100 puan")
        self.lbl_kalan.setFont(QFont('Segoe UI', 8))
        self.lbl_kalan.setStyleSheet(
            "color:" + C['text_muted'] + ";background:transparent;border:none;")
        self.lbl_kalan.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        root.addWidget(self.lbl_kalan)

    def set_data(self, puan, seviye, seviye_no):
        renk = self.SEV_RENK.get(seviye_no, C['primary'])
        self.lbl_seviye.setText("⚡ " + str(seviye))
        self.lbl_seviye.setStyleSheet("color:" + renk + ";background:transparent;font-weight:bold;font-size:12px;")
        self.lbl_puan.setText(str(puan) + " puan")

        idx = min(seviye_no, len(self.ESIKLER)-2)
        sonraki = self.ESIKLER[idx+1]
        onceki  = self.ESIKLER[idx]
        pct = min(1.0, (puan - onceki) / max(sonraki - onceki, 1))
        kalan = max(0, sonraki - puan)

        self.bar_fill.setStyleSheet(
            "background:" + renk + ";border-radius:6px;")
        self.lbl_kalan.setText("Sonraki seviye: " + str(kalan) + " puan")

        # bar genişliğini sonraki repaint'te ayarla
        QTimer.singleShot(30, lambda: self.bar_fill.setFixedWidth(
            int(self.bar_bg.width() * pct)))


# ═══════════════════════════════════════════════════════════════════
# TIER 6 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class RozetlerPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self._sporcu_ids = [None]
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QScrollArea, QGridLayout
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(8)

        # Header sabit
        hdr_w = QWidget(); hdr_w.setFixedHeight(44)
        hdr = QHBoxLayout(hdr_w); hdr.setContentsMargins(0,0,0,0)
        title = QLabel('Rozetler & Basarimlar')
        title.setFont(QFont('Segoe UI', 20, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_kontrol = make_btn('Rozet Kontrol', C['primary'], small=True)
        btn_kontrol.clicked.connect(self._rozet_kontrol)
        hdr.addWidget(btn_kontrol)
        lay.addWidget(hdr_w)

        # Sporcu sec sabit
        fil_w = QWidget(); fil_w.setFixedHeight(INPUT_H + 4)
        fil = QHBoxLayout(fil_w); fil.setContentsMargins(0,0,0,0); fil.setSpacing(8)
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Sec --'], 220)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb); fil.addStretch()
        lay.addWidget(fil_w)

        # Seviye bar sabit
        self.frame_sev = QFrame()
        self.frame_sev.setFixedHeight(80)
        self.frame_sev.setStyleSheet(
            "background:" + C['bg_card'] + ";border-radius:10px;border:1px solid " + C['border'] + ";")
        fs_l = QVBoxLayout(self.frame_sev); fs_l.setContentsMargins(10,6,10,6)
        self.seviye_bar = SeviyeBarWidget()
        self.seviye_bar.setFixedHeight(68)
        fs_l.addWidget(self.seviye_bar)
        lay.addWidget(self.frame_sev)

        # KPI sabit
        kpi_w = QWidget(); kpi_w.setFixedHeight(100)
        kpi = QHBoxLayout(kpi_w); kpi.setSpacing(SPACING); kpi.setContentsMargins(0,0,0,0)
        self.k_rozet = KPICard('Kazanilan Rozet', '-', '🏅', C['primary'], C['primary_dark'])
        self.k_puan  = KPICard('Toplam Puan',     '-', '⭐', '#f59e0b', '#d97706')
        self.k_kalan = KPICard('Kalan Rozet',     '-', '🔒', C['text_muted'], C['border'])
        for k in [self.k_rozet, self.k_puan, self.k_kalan]:
            kpi.addWidget(k)
        lay.addWidget(kpi_w)

        # Baslik sabit
        lbl_kol = QLabel('Rozet Koleksiyonu')
        lbl_kol.setFont(QFont('Segoe UI', 12, QFont.Bold))
        lbl_kol.setFixedHeight(22)
        lay.addWidget(lbl_kol)

        # Scroll - kalan alan
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:" + C['bg_main'] + ";}")
        self.grid_widget = QWidget()
        self.grid_widget.setStyleSheet("background:" + C['bg_main'] + ";")
        self.grid_lay = QGridLayout(self.grid_widget)
        self.grid_lay.setSpacing(10)
        self.grid_lay.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        scroll.setWidget(self.grid_widget)
        lay.addWidget(scroll, 1)

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear(); self.sporcu_cb.addItem('-- Sporcu Sec --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(s['ad'] + ' ' + s['soyad'])
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self.refresh()

    def refresh(self):
        self._load_sporcular()
        while self.grid_lay.count():
            item = self.grid_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        if not self.selected_sporcu_id:
            self.seviye_bar.set_data(0, 'Baslangic', 1)
            for k in [self.k_rozet, self.k_puan, self.k_kalan]: k.set_value('-')
            return

        puan_data = self.db.get_sporcu_puan(self.selected_sporcu_id)
        self.seviye_bar.set_data(puan_data['toplam_puan'], puan_data['seviye'], puan_data['seviye_no'])

        tum_rozetler  = self.db.get_rozetler()
        kazanilan_map = {r['rozet_id']: r for r in self.db.get_sporcu_rozetleri(self.selected_sporcu_id)}
        kazanilan_sayi = len(kazanilan_map)

        self.k_rozet.set_value(str(kazanilan_sayi))
        self.k_puan.set_value(str(puan_data['toplam_puan']))
        self.k_kalan.set_value(str(len(tum_rozetler) - kazanilan_sayi))

        sirali = sorted(tum_rozetler, key=lambda r: (0 if r['rozet_id'] in kazanilan_map else 1))
        cols = max(4, self.width() // 170)
        for i, r in enumerate(sirali):
            k = r['rozet_id'] in kazanilan_map
            tarih = kazanilan_map[r['rozet_id']].get('kazanma_tarihi', '') if k else ''
            kart = RozetKartWidget(r['rozet_adi'], r['icon'], r['puan'], r['aciklama'], k, tarih)
            self.grid_lay.addWidget(kart, i // cols, i % cols)

    def _rozet_kontrol(self):
        if not self.selected_sporcu_id:
            dark_msg(self, 'Uyari', 'Sporcu secin!', QMessageBox.Warning); return
        yeni = self.db.rozet_kontrol_ve_ver(self.selected_sporcu_id)
        if yeni:
            isimler = ', '.join(r['rozet_adi'] for r in yeni)
            dark_msg(self, 'Yeni Rozet!', 'Tebrikler! Yeni rozetler:\n' + isimler)
        else:
            dark_msg(self, 'Bilgi', 'Su an yeni rozet yok. Devam et!')
        self.refresh()


class LeaderboardPage(QWidget):
    """Puan tablosu — tüm sporcular sıralı."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Puan Tablosu')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_ref = make_btn('⟳  Yenile', C['primary'], small=True)
        btn_ref.setMinimumWidth(100); btn_ref.clicked.connect(self.refresh)
        hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        # Top 3 KPI
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k1 = KPICard('🥇 1. Sıra', '—', '👑', '#f59e0b', '#d97706')
        self.k2 = KPICard('🥈 2. Sıra', '—', '🥈', '#6366f1', '#4f46e5')
        self.k3 = KPICard('🥉 3. Sıra', '—', '🥉', '#10b981', '#059669')
        for k in [self.k1, self.k2, self.k3]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        # Leaderboard tablosu
        self.table = make_table(
            ['Sıra','Sporcu','Seviye','Puan','Rozet Sayısı','Durum'],
            [60, -1, 120, 110, 120, 110]
        )
        lay.addWidget(self.table)

        # Rozet özet tablosu
        lbl2 = QLabel('Rozet Dağılımı')
        lbl2.setFont(QFont('Segoe UI', 14, QFont.Bold)); lay.addWidget(lbl2)
        self.tbl_rozet = make_table(
            ['Sporcu','Rozet Sayısı','Rozet Puanı'],
            [-1, 150, 150]
        )
        self.tbl_rozet.setMaximumHeight(220)
        lay.addWidget(self.tbl_rozet)

    def refresh(self):
        lb = self.db.get_leaderboard()

        # Top 3 KPI — puan büyük, isim başlıkta
        kpi_titles = ['🥇 1. Sıra', '🥈 2. Sıra', '🥉 3. Sıra']
        for i, (k, d) in enumerate(zip([self.k1, self.k2, self.k3], lb[:3])):
            ad = f"{d['ad']} {d['soyad']}"
            k.title_text = f"{kpi_titles[i]}  •  {ad}"
            k.set_value(f"{d['toplam_puan']} puan")

        # Tablo
        self.table.setRowCount(0)
        sev_renk = {1:'#6b7280',2:'#92400e',3:'#9ca3af',
                    4:'#f59e0b',5:'#818cf8',6:'#06b6d4'}
        for sira, d in enumerate(lb, 1):
            row = self.table.rowCount(); self.table.insertRow(row)
            madalya = {1:'🥇',2:'🥈',3:'🥉'}.get(sira, str(sira))
            renk = sev_renk.get(d.get('seviye_no',1), C['text_main'])
            vals = [madalya, f"{d['ad']} {d['soyad']}", d['seviye'],
                    str(d['toplam_puan']), f"🏅 {d['rozet_sayisi']}",
                    '⭐'*min(d.get('seviye_no',1), 6)]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 0 and sira <= 3:
                    item.setFont(QFont('Segoe UI', 15))
                if col == 2:
                    item.setForeground(QColor(renk))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                if col == 3:
                    item.setForeground(QColor(C['accent']))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)

        # Rozet dağılımı
        ozet = self.db.get_tum_sporcu_rozetleri_ozet()
        self.tbl_rozet.setRowCount(0)
        for d in ozet:
            row = self.tbl_rozet.rowCount(); self.tbl_rozet.insertRow(row)
            vals = [d['ad'], f"🏅 {d['rozet_sayisi']}", f"⭐ {d['rozet_puani']}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 2:
                    item.setForeground(QColor(C['accent']))
                self.tbl_rozet.setItem(row, col, item)



class MultiLineChart(QWidget):
    """Birden fazla veri serisini tek grafikte göster."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.series = []   # [{'label': str, 'data': [(x,y)], 'color': str}]
        self.title  = ''
        self.setMinimumHeight(240)

    def set_series(self, series, title=''):
        self.series = series
        self.title  = title
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        if not self.series or not any(s['data'] for s in self.series):
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Sporcu seçin'); p.end(); return

        pad_l, pad_r, pad_t, pad_b = 55, 20, 38, 50
        cw = w - pad_l - pad_r; ch = h - pad_t - pad_b

        # Başlık
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 26), Qt.AlignCenter, self.title)

        # Tüm değerleri topla — global min/max
        all_vals = [v for s in self.series for _, v in s['data'] if s['data']]
        if not all_vals: p.end(); return
        mn, mx = min(all_vals), max(all_vals)
        rng = mx - mn if mx != mn else 1

        # Tüm x label'ları topla
        all_labels = []
        for s in self.series:
            for lbl, _ in s['data']:
                if lbl not in all_labels:
                    all_labels.append(lbl)
        all_labels.sort()
        n = len(all_labels)
        if n == 0: p.end(); return

        def cx(lbl):
            idx = all_labels.index(lbl) if lbl in all_labels else 0
            return pad_l + int(idx * cw / max(n - 1, 1))

        def cy(v):
            return pad_t + int((mx - v) / rng * ch)

        # Grid
        p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))
        for gi in range(5):
            gy = pad_t + ch * gi // 4
            p.drawLine(pad_l, gy, w - pad_r, gy)
            gval = mx - rng * gi / 4
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(2, gy - 8, pad_l - 4, 16),
                       Qt.AlignRight | Qt.AlignVCenter, f"{gval:.0f}")
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))

        # Her seri için çizgi + noktalar
        for s in self.series:
            if not s['data']: continue
            col = s['color']

            # Çizgi
            p.setPen(QPen(QColor(col), 2.5))
            p.setBrush(Qt.NoBrush)
            path = QPainterPath()
            pts = [(cx(lbl), cy(val)) for lbl, val in s['data'] if lbl in all_labels]
            if pts:
                path.moveTo(pts[0][0], pts[0][1])
                for x, y in pts[1:]:
                    path.lineTo(x, y)
                p.drawPath(path)

            # Noktalar
            for lbl, val in s['data']:
                if lbl not in all_labels: continue
                x, y = cx(lbl), cy(val)
                p.setBrush(QColor(C['bg_card'])); p.setPen(QPen(QColor(col), 2))
                p.drawEllipse(QPointF(x, y), 4, 4)
                p.setBrush(QColor(col)); p.setPen(Qt.NoPen)
                p.drawEllipse(QPointF(x, y), 2.5, 2.5)

        # X ekseni etiketleri (seyrek)
        step = max(1, n // 8)
        p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 9))
        for i, lbl in enumerate(all_labels):
            if i % step == 0:
                x = cx(lbl)
                p.drawText(QRectF(x - 20, h - pad_b + 5, 40, 16), Qt.AlignCenter, lbl)

        # Eksenler
        p.setPen(QPen(QColor(C['border_light']), 1))
        p.drawLine(pad_l, pad_t, pad_l, pad_t + ch)
        p.drawLine(pad_l, pad_t + ch, w - pad_r, pad_t + ch)

        # Legend
        lx = pad_l + 8; ly = pad_t + 6
        for s in self.series:
            p.setBrush(QColor(s['color'])); p.setPen(Qt.NoPen)
            p.drawRoundedRect(lx, ly, 12, 12, 3, 3)
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 9))
            p.drawText(lx + 16, ly - 1, 80, 14, Qt.AlignLeft | Qt.AlignVCenter, s['label'])
            lx += 100
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 5 — Haftalık Özet Kartı (QPainter)
# ═══════════════════════════════════════════════════════════════════

class OzetKartWidget(QWidget):
    """Haftalık özet — büyük sayılar + ikonlar."""
    def __init__(self, label, value, icon, color, sub='', parent=None):
        super().__init__(parent)
        self.label = label; self.value = str(value)
        self.icon  = icon;  self.color = color; self.sub = sub
        self.setFixedHeight(100)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

    def set_value(self, v, sub=''):
        self.value = str(v)
        if sub: self.sub = sub
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        grad = QLinearGradient(0, 0, w, h)
        grad.setColorAt(0, QColor(self.color))
        grad.setColorAt(1, QColor(QColor(self.color).darker(130).name()))
        path = QPainterPath()
        path.addRoundedRect(QRectF(0, 0, w, h), 12, 12)
        p.fillPath(path, QBrush(grad))
        # Dekor
        p.setBrush(QColor(255,255,255,15)); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(w-20, 20), 40, 40)
        # İkon
        p.setPen(QColor(255,255,255,200)); p.setFont(QFont('Segoe UI', 20))
        p.drawText(QRectF(w-48, 6, 42, 32), Qt.AlignCenter, self.icon)
        # Label
        p.setPen(QColor(255,255,255,170)); p.setFont(QFont('Segoe UI', 10))
        p.drawText(QRectF(10, 8, w-60, 18), Qt.AlignLeft | Qt.AlignVCenter, self.label)
        # Ayraç
        p.setPen(QPen(QColor(255,255,255,40), 1))
        p.drawLine(10, 34, w-10, 34)
        # Değer
        p.setPen(QColor(255,255,255)); p.setFont(QFont('Segoe UI', 20, QFont.Bold))
        p.drawText(QRectF(10, 38, w-20, 36), Qt.AlignLeft | Qt.AlignVCenter, self.value)
        # Alt bilgi
        if self.sub:
            p.setPen(QColor(255,255,255,150)); p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(10, 76, w-20, 18), Qt.AlignLeft, self.sub)
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 5 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class IlerlemeGrafikleriPage(QWidget):
    """30 günlük kilo + kalori + antrenman multi-line grafik."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('İlerleme Grafikleri')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # Filtreler
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 220)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Süre:'))
        self.sure_cb = make_combo(['Son 14 Gün','Son 30 Gün','Son 60 Gün','Son 90 Gün'], 150)
        self.sure_cb.currentTextChanged.connect(self._refresh)
        fil.addWidget(self.sure_cb); fil.addStretch()
        lay.addLayout(fil)

        # Kilo değişim özet kartları
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_baslangic = OzetKartWidget('Başlangıç Kilo', '—', '⚖️', C['primary'])
        self.k_son       = OzetKartWidget('Son Kilo', '—', '📉', C['info'])
        self.k_fark      = OzetKartWidget('Değişim', '—', '📊', C['success'])
        self.k_streak    = OzetKartWidget('Streak', '—', '🔥', C['warning'])
        for k in [self.k_baslangic, self.k_son, self.k_fark, self.k_streak]:
            kpi.addWidget(k)
        lay.addLayout(kpi)

        def chart_frame(widget, h=260):
            f = QFrame()
            f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
            fl = QVBoxLayout(f); fl.setContentsMargins(8,8,8,8)
            widget.setMinimumHeight(h); fl.addWidget(widget)
            return f

        # Kilo trendi
        self.chart_kilo = LineChartWidget()
        lay.addWidget(chart_frame(self.chart_kilo, 240))

        # Antrenman + Kalori multi-line
        self.chart_multi = MultiLineChart()
        lay.addWidget(chart_frame(self.chart_multi, 240))

        # Performans bar (haftalık yoğunluk)
        self.chart_perf = BarChartWidget()
        lay.addWidget(chart_frame(self.chart_perf, 200))

        self._sporcu_ids = [None]

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self._refresh()

    def _refresh(self):
        self._load_sporcular()
        if not self.selected_sporcu_id:
            for chart in [self.chart_kilo, self.chart_multi, self.chart_perf]:
                if hasattr(chart, 'set_data'): chart.set_data([], '')
                elif hasattr(chart, 'set_series'): chart.set_series([], '')
            return

        gun_map = {'Son 14 Gün': 14, 'Son 30 Gün': 30,
                   'Son 60 Gün': 60, 'Son 90 Gün': 90}
        gun = gun_map.get(self.sure_cb.currentText(), 30)

        data = self.db.get_ilerleme_grafik_data(self.selected_sporcu_id, gun)

        # Kilo trend
        kilo_data = [(d[0][-5:] if len(d[0]) > 5 else d[0], d[1])
                     for d in self.db.get_vucut_trend(self.selected_sporcu_id, 'kilo')
                     if d] if True else []
        self.chart_kilo.set_data(
            self.db.get_vucut_trend(self.selected_sporcu_id, 'kilo'),
            'Kilo Trendi (kg)', 'kg', C['primary'])

        # Multi-line: antrenman süresi + kalori
        ant_sure = [(d[0], d[2]) for d in data['antrenman']]
        ant_kalori = [(d[0], d[3]) for d in data['antrenman']]
        kal_data = data['kalori']
        self.chart_multi.set_series([
            {'label': 'Antrenman Süresi (dk)', 'data': ant_sure, 'color': C['primary']},
            {'label': 'Yakılan Kalori',        'data': ant_kalori, 'color': C['danger']},
            {'label': 'Alınan Kalori',         'data': kal_data,  'color': C['success']},
        ], 'Antrenman & Kalori Trendi')

        # Performans bar (haftalık)
        perf = self.db.get_performans_trend(self.selected_sporcu_id, gun // 7)
        self.chart_perf.set_data(
            [(d['hafta'], d['sayi']) for d in perf],
            'Haftalık Antrenman Sayısı')

        # KPI güncelle
        degisim = self.db.get_kilo_degisim(self.selected_sporcu_id)
        if degisim:
            self.k_baslangic.set_value(f"{degisim['ilk_kilo']} kg", degisim['ilk_tarih'])
            self.k_son.set_value(f"{degisim['son_kilo']} kg", degisim['son_tarih'])
            fark = degisim['fark']
            self.k_fark.set_value(f"{fark:+.1f} kg",
                                   '↓ İyi gidiyorsun!' if fark < 0 else '↑ Kas kazanıyor')
        streak = self.db.get_streak(self.selected_sporcu_id)
        self.k_streak.set_value(f"{streak} gün", 'Kesintisiz antrenman')

    def refresh(self):
        self._load_sporcular()
        self._refresh()


class HaftalikOzetPage(QWidget):
    """Son 7 günün tam haftalık özet raporu."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Haftalık Özet')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        lay.addLayout(hdr)

        # Sporcu seç
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 220)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu)
        fil.addWidget(self.sporcu_cb); fil.addStretch()
        lay.addLayout(fil)

        # Bilgi banner
        self.banner = QFrame()
        self.banner.setStyleSheet(f"""
            QFrame{{background:{C['bg_card']};border-radius:12px;
            border-left:4px solid {C['primary']};padding:4px;}}
        """)
        bl = QHBoxLayout(self.banner); bl.setContentsMargins(16,12,16,12)
        self.lbl_banner = QLabel('Sporcu seçin')
        self.lbl_banner.setFont(QFont('Segoe UI', 13)); bl.addWidget(self.lbl_banner)
        lay.addWidget(self.banner)

        # 6 özet kart
        kpi1 = QHBoxLayout(); kpi1.setSpacing(SPACING)
        self.k_ant   = OzetKartWidget('Antrenman', '—', '💪', C['primary'])
        self.k_sure  = OzetKartWidget('Toplam Süre', '—', '⏱', C['info'])
        self.k_kal   = OzetKartWidget('Yakılan Kalori', '—', '🔥', C['danger'])
        for k in [self.k_ant, self.k_sure, self.k_kal]:
            kpi1.addWidget(k)
        lay.addLayout(kpi1)

        kpi2 = QHBoxLayout(); kpi2.setSpacing(SPACING)
        self.k_zorluk = OzetKartWidget('Ort. Zorluk', '—', '⚡', C['warning'])
        self.k_bes    = OzetKartWidget('Ort. Kalori/Gün', '—', '🥗', C['success'])
        self.k_str    = OzetKartWidget('Streak', '—', '🏆', '#8b5cf6')
        for k in [self.k_zorluk, self.k_bes, self.k_str]:
            kpi2.addWidget(k)
        lay.addLayout(kpi2)

        # Kategori dağılımı + başarı mesajı
        row3 = QHBoxLayout(); row3.setSpacing(SPACING)

        # Sol: Kategori pie
        left_f = QFrame()
        left_f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ll = QVBoxLayout(left_f); ll.setContentsMargins(8,8,8,8)
        self.chart_kat = PieChartWidget()
        self.chart_kat.setMinimumHeight(220)
        ll.addWidget(self.chart_kat)
        row3.addWidget(left_f, 1)

        # Sağ: Detay tablo
        right_f = QFrame()
        right_f.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        rl = QVBoxLayout(right_f); rl.setContentsMargins(12,12,12,12); rl.setSpacing(10)
        lbl_det = QLabel('Detay Bilgi')
        lbl_det.setFont(QFont('Segoe UI', 13, QFont.Bold)); rl.addWidget(lbl_det)
        self.detail_table = make_table(['Özellik','Değer'], [-1, 150])
        self.detail_table.setMaximumHeight(240)
        rl.addWidget(self.detail_table)

        # Öneri alanı
        self.lbl_oneri = QLabel('')
        self.lbl_oneri.setWordWrap(True)
        self.lbl_oneri.setFont(QFont('Segoe UI', 11))
        self.lbl_oneri.setStyleSheet(f"""
            color:{C['text_secondary']};
            background:{C['bg_secondary']};
            border-radius:8px; padding:10px;
            border-left:3px solid {C['primary']};
        """)
        rl.addWidget(self.lbl_oneri)
        rl.addStretch()
        row3.addWidget(right_f, 1)
        lay.addLayout(row3)

        self._sporcu_ids = [None]

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self._refresh()

    def _refresh(self):
        self._load_sporcular()
        if not self.selected_sporcu_id:
            self.lbl_banner.setText('Sporcu seçin')
            self.detail_table.setRowCount(0); return

        ozet = self.db.get_haftalik_ozet(self.selected_sporcu_id)
        sporcu = ozet.get('sporcu', {})
        ad = f"{sporcu.get('ad','')} {sporcu.get('soyad','')}"

        # Banner
        tarih_str = datetime.now().strftime('%d.%m.%Y')
        self.lbl_banner.setText(
            f"🗓  {ad} — Haftalık Özet Raporu  |  {tarih_str}")

        # KPI kartlar
        self.k_ant.set_value(str(ozet['antrenman_sayisi']), 'antrenman')
        self.k_sure.set_value(f"{ozet['toplam_sure']} dk",
                               f"≈ {ozet['toplam_sure']//60}s {ozet['toplam_sure']%60}dk")
        self.k_kal.set_value(f"{ozet['toplam_kalori_yak']:,}", 'kcal')
        self.k_zorluk.set_value(f"{ozet['ort_zorluk']}/10", 'ortalama')
        self.k_bes.set_value(f"{ozet['ort_gunluk_kalori']:,}", 'kcal/gün')
        self.k_str.set_value(f"{ozet['streak']} gün", 'kesintisiz 🔥')

        # Kategori pie
        self.chart_kat.set_data(ozet['kategori_dagilimi'], 'Haftalık Antrenman Dağılımı')

        # Detay tablo
        hedef = ozet.get('kalori_hedef', {})
        rows = [
            ('Spor Yapılan Gün',   f"{ozet['antrenman_sayisi']} / 7"),
            ('En Çok Kategori',    ozet['en_cok_kategori']),
            ('Günlük Kalori Hedefi', f"{hedef.get('gunluk_kalori','—')} kcal"),
            ('Protein Hedefi',     f"{hedef.get('protein_gram','—')} g"),
            ('Karbonhidrat Hedefi',f"{hedef.get('karb_gram','—')} g"),
            ('Yağ Hedefi',         f"{hedef.get('yag_gram','—')} g"),
            ('Streak',             f"{ozet['streak']} gün"),
        ]
        self.detail_table.setRowCount(0)
        for lbl, val in rows:
            row = self.detail_table.rowCount()
            self.detail_table.insertRow(row)
            item_l = QTableWidgetItem(lbl); item_l.setFont(QFont('Segoe UI',12))
            item_l.setForeground(QColor(C['text_secondary']))
            item_v = QTableWidgetItem(str(val)); item_v.setFont(QFont('Segoe UI',12,QFont.Bold))
            self.detail_table.setItem(row, 0, item_l)
            self.detail_table.setItem(row, 1, item_v)

        # Akıllı öneri
        onerimler = []
        if ozet['antrenman_sayisi'] < 3:
            onerimler.append("💡 Bu hafta az antrenman yaptın. Haftada en az 3 antrenman hedefle.")
        elif ozet['antrenman_sayisi'] >= 5:
            onerimler.append("🏆 Harika! Haftada 5+ antrenman yaptın, dinlenme günlerine dikkat et.")
        if ozet['streak'] >= 7:
            onerimler.append("🔥 7 günlük streak! Şampiyonluk ruhu!")
        elif ozet['streak'] == 0:
            onerimler.append("😴 Bugün streak yok. Yarın başla!")
        if hedef and ozet['ort_gunluk_kalori'] > 0:
            oran = ozet['ort_gunluk_kalori'] / hedef.get('gunluk_kalori', 2000)
            if oran < 0.8:
                onerimler.append("⚠️ Kalori alımın hedefin altında. Beslenmeye dikkat et.")
            elif oran > 1.2:
                onerimler.append("⚠️ Kalori alımın hedefin çok üzerinde. Porsiyon kontrolü yap.")
        if not onerimler:
            onerimler.append("✅ Her şey yolunda görünüyor! Devam et.")

        self.lbl_oneri.setText('\n'.join(onerimler))

    def refresh(self):
        self._load_sporcular()
        self._refresh()



class MakroPieChart(QWidget):
    """Protein / Karb / Yağ — donut + 3 progress bar."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.protein = 0; self.karb = 0; self.yag = 0
        self.h_protein = 150; self.h_karb = 200; self.h_yag = 65
        self.setFixedHeight(130)

    def set_data(self, protein, karb, yag, h_protein=150, h_karb=200, h_yag=65):
        self.protein = protein; self.karb = karb; self.yag = yag
        self.h_protein = h_protein; self.h_karb = h_karb; self.h_yag = h_yag
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        total = self.protein + self.karb + self.yag
        colors = [C['danger'], C['warning'], C['info']]
        labels = ['Protein', 'Karb', 'Yağ']
        vals   = [self.protein, self.karb, self.yag]
        hvals  = [self.h_protein, self.h_karb, self.h_yag]

        if total == 0:
            p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Veri yok'); p.end(); return

        # ── Başlık ──
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 11, QFont.Bold))
        p.drawText(QRectF(0, 4, w, 20), Qt.AlignCenter, 'Makro Dağılımı')

        # ── Sol: Donut ──
        pie_size = min(h - 30, w // 3)          # kare boyut
        pie_r = pie_size // 2
        cx = pie_r + 12
        cy = 14 + pie_r + (h - 14 - pie_size) // 2

        start = 90 * 16
        for val, col in zip(vals, colors):
            span = int(val / total * 5760) if total else 0
            p.setBrush(QColor(col)); p.setPen(Qt.NoPen)
            p.drawPie(QRectF(cx - pie_r, cy - pie_r, pie_size, pie_size), start, span)
            start += span

        # İç beyaz daire (donut deliği)
        inner = int(pie_r * 0.55)
        p.setBrush(QColor(C['bg_card'])); p.setPen(Qt.NoPen)
        p.drawEllipse(QPointF(cx, cy), inner, inner)
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 10, QFont.Bold))
        p.drawText(QRectF(cx - inner, cy - 10, inner * 2, 20),
                   Qt.AlignCenter, f"{int(total)}g")

        # ── Sağ: 3 bar ──
        # Her bar için alan: başlık(16) + bar(12) + boşluk(6) = 34px → 3x = 102 < 130
        bar_area_x = cx + pie_r + 20
        bar_w = w - bar_area_x - 12
        row_h = (h - 24) // 3          # 3 satıra böl
        top_y = 22

        for i, (lbl, val, hval, col) in enumerate(zip(labels, vals, hvals, colors)):
            ry = top_y + i * row_h      # satır başlangıcı

            # Label sol, değer/hedef sağ
            p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 10, QFont.Bold))
            p.drawText(QRectF(bar_area_x, ry, bar_w // 2, 16),
                       Qt.AlignLeft | Qt.AlignVCenter, lbl)
            p.setPen(QColor(col)); p.setFont(QFont('Segoe UI', 9, QFont.Bold))
            p.drawText(QRectF(bar_area_x + bar_w // 2, ry, bar_w // 2, 16),
                       Qt.AlignRight | Qt.AlignVCenter, f"{val:.0f}/{hval}g")

            # Bar arka plan
            bar_y = ry + 18
            bh = 10
            p.setBrush(QColor(C['border'])); p.setPen(Qt.NoPen)
            p.drawRoundedRect(int(bar_area_x), bar_y, int(bar_w), bh, 5, 5)

            # Bar doluluk
            pct = min(1.0, val / hval) if hval else 0
            fw = int(bar_w * pct)
            if fw > 0:
                grad = QLinearGradient(bar_area_x, 0, bar_area_x + bar_w, 0)
                grad.setColorAt(0, QColor(col))
                grad.setColorAt(1, QColor(QColor(col).lighter(140).name()))
                p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen)
                p.drawRoundedRect(int(bar_area_x), bar_y, fw, bh, 5, 5)

        p.end()


class KaloriGaugeWidget(QWidget):
    """Günlük kalori doluluk göstergesi."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.kalori = 0; self.hedef = 2000
        self.setFixedHeight(130)

    def set_data(self, kalori, hedef):
        self.kalori = kalori; self.hedef = hedef; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        pct = min(1.0, self.kalori / self.hedef) if self.hedef else 0
        kalan = max(0, self.hedef - self.kalori)
        renk = C['success'] if pct < 0.85 else (C['warning'] if pct < 1.0 else C['danger'])
        pad = 16
        bw = w - pad * 2

        # Satır 1 (y=8): Başlık sol, yüzde sağ
        p.setPen(QColor(C['text_secondary'])); p.setFont(QFont('Segoe UI', 11))
        p.drawText(QRectF(pad, 8, bw, 20), Qt.AlignLeft | Qt.AlignVCenter, 'Günlük Kalori')
        p.setPen(QColor(renk)); p.setFont(QFont('Segoe UI', 11, QFont.Bold))
        p.drawText(QRectF(pad, 8, bw, 20), Qt.AlignRight | Qt.AlignVCenter, f"%{int(pct*100)}")

        # Satır 2 (y=34): Büyük kalori değeri
        p.setPen(QColor(C['text_main'])); p.setFont(QFont('Segoe UI', 22, QFont.Bold))
        p.drawText(QRectF(pad, 30, bw, 34), Qt.AlignLeft | Qt.AlignVCenter,
                   f"{int(self.kalori)} kcal")

        # Satır 3 (y=68): Bar
        by = 70; bh = 16
        p.setBrush(QColor(C['border'])); p.setPen(Qt.NoPen)
        p.drawRoundedRect(pad, by, bw, bh, 8, 8)
        fw = int(bw * pct)
        if fw > 0:
            grad = QLinearGradient(pad, 0, pad + bw, 0)
            grad.setColorAt(0.0, QColor(C['success']))
            grad.setColorAt(0.7, QColor(C['warning']))
            grad.setColorAt(1.0, QColor(C['danger']))
            p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen)
            p.drawRoundedRect(pad, by, fw, bh, 8, 8)

        # Satır 4 (y=92): Hedef sol, Kalan sağ
        p.setPen(QColor(C['text_muted'])); p.setFont(QFont('Segoe UI', 10))
        p.drawText(QRectF(pad, 92, bw // 2, 18),
                   Qt.AlignLeft | Qt.AlignVCenter, f"Hedef: {self.hedef} kcal")
        p.setPen(QColor(renk)); p.setFont(QFont('Segoe UI', 10, QFont.Bold))
        p.drawText(QRectF(pad + bw // 2, 92, bw // 2, 18),
                   Qt.AlignRight | Qt.AlignVCenter, f"Kalan: {int(kalan)} kcal")
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 4 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

class BesinEkleDialog(BaseDialog):
    def __init__(self, db, parent=None):
        super().__init__('Yeni Besin Ekle', parent, 560)
        self.db = db; self._build()

    def _build(self):
        self.f_ad  = self.add_row('Besin Adı *', self.inp('Örn: Yumurta, Tavuk Göğsü'))
        self.f_kat = self.add_row('Kategori', self.combo(
            ['Protein','Karbonhidrat','Yağ','Sebze','Meyve','Süt Ürünleri','Tahıl','Genel']))
        self.f_pors = self.add_row('Standart Porsiyon (g)', self.spin(1,1000,100))

        # Makro satırı
        def nw(mn=0.0, mx=9999.0, val=0.0, suf='g'):
            w=QDoubleSpinBox(); w.setRange(mn,mx); w.setValue(val)
            w.setDecimals(1); w.setSuffix(suf); w.setFixedHeight(INPUT_H)
            w.setFont(QFont('Segoe UI',13)); return w

        r1=QWidget(); l1=QHBoxLayout(r1); l1.setContentsMargins(0,0,0,0)
        self.f_kal=nw(0,9999,0,' kcal'); l1.addWidget(QLabel('Kalori:')); l1.addWidget(self.f_kal)
        self.f_prot=nw(0,999,0); l1.addWidget(QLabel('Protein:')); l1.addWidget(self.f_prot)
        self.add_row('Kalori / Protein', r1)

        r2=QWidget(); l2=QHBoxLayout(r2); l2.setContentsMargins(0,0,0,0)
        self.f_karb=nw(0,999,0); l2.addWidget(QLabel('Karb:')); l2.addWidget(self.f_karb)
        self.f_yag=nw(0,999,0); l2.addWidget(QLabel('Yağ:')); l2.addWidget(self.f_yag)
        self.add_row('Karbonhidrat / Yağ', r2)

        self.f_lif=self.add_row('Lif (g)', self.dspin(0,100,0,1))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        if not self.f_ad.text().strip():
            dark_msg(self,'Hata','Besin adı zorunludur!',QMessageBox.Warning); return
        self.result_data = {
            'ad': self.f_ad.text().strip(), 'kategori': self.f_kat.currentText(),
            'porsiyon_gram': self.f_pors.value(), 'kalori': self.f_kal.value(),
            'protein': self.f_prot.value(), 'karbonhidrat': self.f_karb.value(),
            'yag': self.f_yag.value(), 'lif': self.f_lif.value(),
        }
        self.accept()


class BeslenmeKaydiDialog(BaseDialog):
    def __init__(self, db, parent=None):
        super().__init__('Öğün Ekle', parent, 600)
        self.db = db; self._build()

    def _build(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(
            ['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)

        self.f_tarih = QDateEdit(); self.f_tarih.setCalendarPopup(True)
        self.f_tarih.setDate(QDate.currentDate()); self.f_tarih.setFixedHeight(INPUT_H)
        self.f_tarih.setFont(QFont('Segoe UI',13))
        self.add_row('Tarih', self.f_tarih)

        self.f_ogun = self.combo(['Kahvalti','Ogle','Aksam','Ara Ogun','Gece'])
        self.add_row('Öğün', self.f_ogun)

        besinler = self.db.get_besinler()
        self.f_besin = self.combo(
            ['-- Besin Seç --'] + [f"{b['ad']} ({b['kalori']:.0f}kcal/{b['porsiyon_gram']:.0f}g)" for b in besinler])
        self._besin_ids = [None] + [b['besin_id'] for b in besinler]
        self._besinler = [None] + besinler
        self.f_besin.currentIndexChanged.connect(self._on_besin_change)
        self.add_row('Besin *', self.f_besin)

        self.f_miktar = self.spin(1,2000,100); self.f_miktar.setSuffix(' g')
        self.f_miktar.valueChanged.connect(self._on_miktar_change)
        self.add_row('Miktar', self.f_miktar)

        # Makro önizleme
        self.lbl_makro = QLabel('— Besin seçin —')
        self.lbl_makro.setStyleSheet(f"color:{C['primary_light']};font-size:12px;")
        self.lbl_makro.setWordWrap(True)
        self._main_layout.addWidget(self.lbl_makro)

        self.f_notlar = self.txt('Notlar...')
        self.add_row('Notlar', self.f_notlar)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Ekle', C['success'], self._save)

    def _on_besin_change(self, idx):
        self._on_miktar_change(self.f_miktar.value())

    def _on_miktar_change(self, miktar):
        idx = self.f_besin.currentIndex()
        if idx == 0 or not self._besinler[idx]: return
        b = self._besinler[idx]; oran = miktar / (b['porsiyon_gram'] or 100)
        self.lbl_makro.setText(
            f"Kalori: {b['kalori']*oran:.0f} kcal  |  "
            f"Protein: {b['protein']*oran:.1f}g  |  "
            f"Karb: {b['karbonhidrat']*oran:.1f}g  |  "
            f"Yağ: {b['yag']*oran:.1f}g")

    def _save(self):
        if self.f_sporcu.currentIndex()==0:
            dark_msg(self,'Hata','Sporcu seçiniz!',QMessageBox.Warning); return
        if self.f_besin.currentIndex()==0:
            dark_msg(self,'Hata','Besin seçiniz!',QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id': self._sporcu_ids[self.f_sporcu.currentIndex()],
            'besin_id':  self._besin_ids[self.f_besin.currentIndex()],
            'tarih':     self.f_tarih.date().toString('yyyy-MM-dd'),
            'ogun':      self.f_ogun.currentText(),
            'miktar_gram': self.f_miktar.value(),
            'notlar':    self.f_notlar.toPlainText().strip(),
        }
        self.accept()


class KaloriHedefDialog(BaseDialog):
    def __init__(self, db, sporcu_id, mevcut=None, parent=None):
        super().__init__('Kalori Hedeflerini Ayarla', parent, 500)
        self.db = db; self.sporcu_id = sporcu_id; self.mevcut = mevcut
        self._build()

    def _build(self):
        self.f_kal   = self.add_row('Günlük Kalori (kcal)', self.spin(800,6000,2000))
        self.f_prot  = self.add_row('Protein (g)', self.spin(30,500,150))
        self.f_karb  = self.add_row('Karbonhidrat (g)', self.spin(30,700,200))
        self.f_yag   = self.add_row('Yağ (g)', self.spin(20,300,65))
        if self.mevcut:
            self.f_kal.setValue(self.mevcut.get('gunluk_kalori',2000))
            self.f_prot.setValue(self.mevcut.get('protein_gram',150))
            self.f_karb.setValue(self.mevcut.get('karb_gram',200))
            self.f_yag.setValue(self.mevcut.get('yag_gram',65))
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        self.result_data = {
            'gunluk_kalori': self.f_kal.value(),
            'protein_gram':  self.f_prot.value(),
            'karb_gram':     self.f_karb.value(),
            'yag_gram':      self.f_yag.value(),
        }
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 4 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class BeslenmePage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self.selected_tarih = datetime.now().strftime('%Y-%m-%d')
        self.selected_beslenme_id = None
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QTabWidget
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Beslenme Takibi')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_new = make_btn('+ Öğün Ekle', C['success'])
        btn_new.clicked.connect(self._add_kayit); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # Filtreler
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 200)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu_change)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Tarih:'))
        self.tarih_edit = QDateEdit(); self.tarih_edit.setCalendarPopup(True)
        self.tarih_edit.setDate(QDate.currentDate()); self.tarih_edit.setFixedHeight(INPUT_H)
        self.tarih_edit.setFont(QFont('Segoe UI',13)); self.tarih_edit.setFixedWidth(150)
        self.tarih_edit.dateChanged.connect(self._on_tarih_change)
        fil.addWidget(self.tarih_edit)
        btn_hedef = make_btn('⚙ Hedef Ayarla', C['primary'], small=True)
        btn_hedef.clicked.connect(self._hedef_ayarla); fil.addWidget(btn_hedef)
        fil.addStretch()
        lay.addLayout(fil)

        # Üst: Gauge + Makro Pie — SABİT YÜKSEKLIK CONTAINER
        kpi_container = QWidget()
        kpi_container.setFixedHeight(150)
        kpi = QHBoxLayout(kpi_container)
        kpi.setSpacing(SPACING); kpi.setContentsMargins(0,0,0,0)

        gauge_frame = QFrame()
        gauge_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        gfl = QVBoxLayout(gauge_frame); gfl.setContentsMargins(8,8,8,8)
        self.gauge = KaloriGaugeWidget()
        self.gauge.setFixedHeight(130)
        gfl.addWidget(self.gauge)
        kpi.addWidget(gauge_frame, 1)

        pie_frame = QFrame()
        pie_frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        pfl = QVBoxLayout(pie_frame); pfl.setContentsMargins(8,8,8,8)
        self.makro_pie = MakroPieChart()
        self.makro_pie.setFixedHeight(130)
        pfl.addWidget(self.makro_pie)
        kpi.addWidget(pie_frame, 1)

        lay.addWidget(kpi_container)

        # Haftalık kalori grafiği
        self.haftalik_chart = BarChartWidget()
        self.haftalik_chart.setFixedHeight(160)
        frame = QFrame()
        frame.setStyleSheet(f"background:{C['bg_card']};border-radius:10px;border:1px solid {C['border']};")
        fl = QVBoxLayout(frame); fl.setContentsMargins(8,8,8,8); fl.addWidget(self.haftalik_chart)
        lay.addWidget(frame)

        # Öğün tablosu
        lbl_t = QLabel('Günlük Öğünler')
        lbl_t.setFont(QFont('Segoe UI', 14, QFont.Bold)); lay.addWidget(lbl_t)

        self.table = make_table(
            ['ID','Öğün','Besin','Miktar','Kalori','Protein','Karb','Yağ'],
            [0, 130, -1, 80, 100, 85, 85, 75]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table, 1)

        bot = QHBoxLayout()
        self.btn_del = make_btn('Sil', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

        self._sporcu_ids = [None]

    def _load_sporcular(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu_change(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self.refresh()

    def _on_tarih_change(self, date):
        self.selected_tarih = date.toString('yyyy-MM-dd'); self.refresh()

    def refresh(self):
        self._load_sporcular()
        if not self.selected_sporcu_id:
            self.table.setRowCount(0)
            self.gauge.set_data(0, 2000)
            self.makro_pie.set_data(0,0,0); return

        # Makro özet
        ozet = self.db.get_gunluk_makro_ozet(self.selected_sporcu_id, self.selected_tarih)
        h = ozet['hedef']
        self.gauge.set_data(ozet['kalori'], h.get('gunluk_kalori',2000))
        self.makro_pie.set_data(ozet['protein'], ozet['karb'], ozet['yag'],
                                h.get('protein_gram',150), h.get('karb_gram',200), h.get('yag_gram',65))

        # Haftalık kalori grafiği
        haftalik = self.db.get_haftalik_kalori(self.selected_sporcu_id)
        self.haftalik_chart.set_data(haftalik, 'Son 7 Gün Kalori')

        # Öğün tablosu
        data = self.db.get_gunluk_beslenme(self.selected_sporcu_id, self.selected_tarih)
        self.table.setRowCount(0)
        ogun_renk = {
            'Kahvalti': C['warning'], 'Ogle': C['success'],
            'Aksam': C['primary'], 'Ara Ogun': C['info'], 'Gece': C['text_muted']
        }
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['beslenme_id']), d['ogun'], d['besin_adi'],
                    f"{d['miktar_gram']:.0f}g",
                    f"{d['kalori_toplam']:.0f} kcal",
                    f"{d['protein_toplam']:.1f}g",
                    f"{d['karb_toplam']:.1f}g",
                    f"{d['yag_toplam']:.1f}g"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI',13))
                if col == 1:
                    item.setForeground(QColor(ogun_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI',13,QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_beslenme_id = None; self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_beslenme_id = int(self.table.item(self.table.currentRow(),0).text())
            self.btn_del.setEnabled(True)

    def _add_kayit(self):
        dlg = BeslenmeKaydiDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_beslenme_kaydi(dlg.result_data)
            # Sporcu seç
            sid = dlg.result_data['sporcu_id']
            for i, s_id in enumerate(self._sporcu_ids):
                if s_id == sid: self.sporcu_cb.setCurrentIndex(i); break
            self.refresh()

    def _delete(self):
        if not self.selected_beslenme_id: return
        if dark_confirm(self, 'Sil', 'Bu öğün kaydı silinecek. Emin misiniz?'):
            self.db.delete_beslenme_kaydi(self.selected_beslenme_id); self.refresh()

    def _hedef_ayarla(self):
        if not self.selected_sporcu_id:
            dark_msg(self,'Uyarı','Önce sporcu seçin!',QMessageBox.Warning); return
        mevcut = self.db.get_set_kalori_hedefi(self.selected_sporcu_id)
        dlg = KaloriHedefDialog(self.db, self.selected_sporcu_id, mevcut, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.set_kalori_hedefi(self.selected_sporcu_id, dlg.result_data)
            dark_msg(self,'Başarılı','Kalori hedefleri güncellendi!')
            self.refresh()


class BesinlerPage(QWidget):
    """Besin veritabanı yönetimi."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Besin Veritabanı')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};"); hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Besin Ekle', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        fil = QHBoxLayout()
        self.search = make_search('Besin adı ara...')
        self.search.textChanged.connect(lambda: self._timer.start(300))
        fil.addWidget(self.search, 1)
        fil.addWidget(QLabel('Kategori:'))
        self.kat_cb = make_combo(
            ['Tümü','Protein','Karbonhidrat','Yağ','Sebze','Meyve','Süt Ürünleri','Tahıl','Genel'], 160)
        self.kat_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.kat_cb)
        lay.addLayout(fil)

        self.table = make_table(
            ['ID','Besin Adı','Kategori','Porsiyon','Kalori','Protein','Karb','Yağ','Lif'],
            [0, -1, 120, 90, 90, 85, 85, 75, 65]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_del = make_btn('Sil', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_besinler(self.search.text(), self.kat_cb.currentText())
        self.lbl_count.setText(f"{len(data)} besin")
        self.table.setRowCount(0)
        kat_renk = {
            'Protein': C['danger'], 'Karbonhidrat': C['warning'],
            'Yağ': C['info'], 'Sebze': C['success'],
            'Meyve': '#ec4899', 'Süt Ürünleri': C['primary_light'],
        }
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['besin_id']), d['ad'], d['kategori'],
                    f"{d['porsiyon_gram']:.0f}g",
                    f"{d['kalori']:.0f} kcal",
                    f"{d['protein']:.1f}g",
                    f"{d['karbonhidrat']:.1f}g",
                    f"{d['yag']:.1f}g",
                    f"{d['lif']:.1f}g"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI',13))
                if col == 2:
                    item.setForeground(QColor(kat_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI',13,QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_id = None; self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(),0).text())
            self.btn_del.setEnabled(True)

    def _add(self):
        dlg = BesinEkleDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_besin(dlg.result_data)
            dark_msg(self,'Başarılı','Besin eklendi!'); self.refresh()

    def _delete(self):
        if not self.selected_id: return
        name = self.table.item(self.table.currentRow(),1).text()
        if dark_confirm(self,'Sil',f'"{name}" silinecek. Emin misiniz?'):
            self.db.delete_besin(self.selected_id); self.refresh()



class LineChartWidget(QWidget):
    """QPainter çizgi grafik — vücut ölçüm trendleri için."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data   = []   # [(label, value), ...]
        self.title  = ''
        self.unit   = ''
        self.color  = C['primary']
        self.setMinimumHeight(200)

    def set_data(self, data, title='', unit='', color=None):
        self.data  = data
        self.title = title
        self.unit  = unit
        if color: self.color = color
        self.update()

    def paintEvent(self, event):
        if len(self.data) < 2:
            p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
            p.fillRect(0, 0, self.width(), self.height(), QColor(C['bg_card']))
            p.setPen(QColor(C['text_muted']))
            p.setFont(QFont('Segoe UI', 11))
            p.drawText(self.rect(), Qt.AlignCenter, 'Yeterli veri yok\n(en az 2 ölçüm)')
            p.end(); return

        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        pad_l, pad_r, pad_t, pad_b = 55, 20, 35, 45
        cw = w - pad_l - pad_r
        ch = h - pad_t - pad_b

        vals = [v for _, v in self.data]
        mn, mx = min(vals), max(vals)
        rng = mx - mn if mx != mn else 1

        # Başlık
        p.setPen(QColor(C['text_main']))
        p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(0, 5, w, 26), Qt.AlignCenter, self.title)

        # Grid yatay çizgiler
        p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))
        for gi in range(5):
            gy = pad_t + ch * gi // 4
            p.drawLine(pad_l, gy, w - pad_r, gy)
            gval = mx - rng * gi / 4
            p.setPen(QColor(C['text_muted']))
            p.setFont(QFont('Segoe UI', 9))
            p.drawText(QRectF(2, gy - 8, pad_l - 5, 16),
                       Qt.AlignRight | Qt.AlignVCenter, f"{gval:.1f}")
            p.setPen(QPen(QColor(C['border']), 1, Qt.DashLine))

        # Koordinat hesapla
        def cx(i): return pad_l + int(i * cw / (len(self.data) - 1))
        def cy(v): return pad_t + int((mx - v) / rng * ch)

        # Gradient fill altı
        grad = QLinearGradient(0, pad_t, 0, pad_t + ch)
        grad.setColorAt(0, QColor(self.color.replace('#', '#60') if '#' in self.color else self.color))
        grad.setColorAt(1, QColor(C['bg_card']))
        fill = QPainterPath()
        fill.moveTo(cx(0), pad_t + ch)
        for i, (_, v) in enumerate(self.data):
            fill.lineTo(cx(i), cy(v))
        fill.lineTo(cx(len(self.data)-1), pad_t + ch)
        fill.closeSubpath()
        p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen)
        p.drawPath(fill)

        # Çizgi
        p.setPen(QPen(QColor(self.color), 2.5))
        p.setBrush(Qt.NoBrush)
        path = QPainterPath()
        path.moveTo(cx(0), cy(vals[0]))
        for i in range(1, len(self.data)):
            path.lineTo(cx(i), cy(vals[i]))
        p.drawPath(path)

        # Noktalar + değerler
        for i, (lbl, v) in enumerate(self.data):
            x, y = cx(i), cy(v)
            # Dış halka
            p.setBrush(QColor(C['bg_card'])); p.setPen(QPen(QColor(self.color), 2))
            p.drawEllipse(QPointF(x, y), 5, 5)
            # İç nokta
            p.setBrush(QColor(self.color)); p.setPen(Qt.NoPen)
            p.drawEllipse(QPointF(x, y), 3, 3)
            # Değer etiketi
            p.setPen(QColor(C['text_main']))
            p.setFont(QFont('Segoe UI', 9, QFont.Bold))
            p.drawText(QRectF(x-20, y-22, 40, 16), Qt.AlignCenter, f"{v:.1f}")
            # Tarih etiketi
            p.setPen(QColor(C['text_muted']))
            p.setFont(QFont('Segoe UI', 9))
            lbl_short = lbl[-5:] if len(lbl) > 5 else lbl
            p.drawText(QRectF(x-22, h-pad_b+6, 44, 18), Qt.AlignCenter, lbl_short)

        # Eksen
        p.setPen(QPen(QColor(C['border_light']), 1))
        p.drawLine(pad_l, pad_t, pad_l, pad_t + ch)
        p.drawLine(pad_l, pad_t + ch, w - pad_r, pad_t + ch)

        # Değişim etiketi
        fark = vals[-1] - vals[0]
        fark_renk = C['success'] if fark <= 0 else C['danger']
        p.setPen(QColor(fark_renk))
        p.setFont(QFont('Segoe UI', 10, QFont.Bold))
        p.drawText(QRectF(pad_l + 6, pad_t + 4, 120, 18),
                   Qt.AlignLeft, f"Değişim: {fark:+.1f} {self.unit}")
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — DIALOG'LAR
# ═══════════════════════════════════════════════════════════════════

class VucutOlcumDialog(BaseDialog):
    def __init__(self, db, parent=None):
        super().__init__('Vücut Ölçümü Ekle', parent, 620)
        self.db = db; self._build()

    def _build(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(
            ['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)

        self.f_tarih = QDateEdit(); self.f_tarih.setCalendarPopup(True)
        self.f_tarih.setDate(QDate.currentDate()); self.f_tarih.setFixedHeight(INPUT_H)
        self.f_tarih.setFont(QFont('Segoe UI', 13))
        self.add_row('Ölçüm Tarihi', self.f_tarih)

        # Ölçüm alanları — iki sütun
        def dspin(val=0.0):
            w = QDoubleSpinBox(); w.setRange(0, 999); w.setValue(val)
            w.setDecimals(1); w.setSuffix(' cm'); w.setFixedHeight(INPUT_H)
            w.setFont(QFont('Segoe UI', 13)); return w

        self.f_kilo = self.dspin(70.0, 20, 300, 1); self.f_kilo.setSuffix(' kg')
        self.add_row('Kilo (kg)', self.f_kilo)

        row1 = QWidget(); rl1 = QHBoxLayout(row1); rl1.setContentsMargins(0,0,0,0)
        self.f_bel = dspin(75); rl1.addWidget(QLabel('Bel:')); rl1.addWidget(self.f_bel)
        self.f_gogus = dspin(90); rl1.addWidget(QLabel('Göğüs:')); rl1.addWidget(self.f_gogus)
        self.add_row('Bel / Göğüs (cm)', row1)

        row2 = QWidget(); rl2 = QHBoxLayout(row2); rl2.setContentsMargins(0,0,0,0)
        self.f_kalca = dspin(95); rl2.addWidget(QLabel('Kalça:')); rl2.addWidget(self.f_kalca)
        self.f_yag = dspin(18); self.f_yag.setSuffix(' %'); rl2.addWidget(QLabel('Yağ %:')); rl2.addWidget(self.f_yag)
        self.add_row('Kalça / Yağ%', row2)

        row3 = QWidget(); rl3 = QHBoxLayout(row3); rl3.setContentsMargins(0,0,0,0)
        self.f_sag_kol = dspin(35); rl3.addWidget(QLabel('Sağ:')); rl3.addWidget(self.f_sag_kol)
        self.f_sol_kol = dspin(35); rl3.addWidget(QLabel('Sol:')); rl3.addWidget(self.f_sol_kol)
        self.add_row('Kol (cm)', row3)

        row4 = QWidget(); rl4 = QHBoxLayout(row4); rl4.setContentsMargins(0,0,0,0)
        self.f_sag_bacak = dspin(58); rl4.addWidget(QLabel('Sağ:')); rl4.addWidget(self.f_sag_bacak)
        self.f_sol_bacak = dspin(58); rl4.addWidget(QLabel('Sol:')); rl4.addWidget(self.f_sol_bacak)
        self.add_row('Bacak (cm)', row4)

        self.f_notlar = self.txt('Notlar...')
        self.add_row('Notlar', self.f_notlar)

        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        if self.f_sporcu.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Sporcu seçiniz!', QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id':       self._sporcu_ids[self.f_sporcu.currentIndex()],
            'tarih':           self.f_tarih.date().toString('yyyy-MM-dd'),
            'kilo':            self.f_kilo.value(),
            'bel':             self.f_bel.value(),
            'gogus':           self.f_gogus.value(),
            'kalca':           self.f_kalca.value(),
            'sag_kol':         self.f_sag_kol.value(),
            'sol_kol':         self.f_sol_kol.value(),
            'sag_bacak':       self.f_sag_bacak.value(),
            'sol_bacak':       self.f_sol_bacak.value(),
            'vucut_yag_orani': self.f_yag.value(),
            'notlar':          self.f_notlar.toPlainText().strip(),
        }
        self.accept()


class HedefDialog(BaseDialog):
    def __init__(self, db, parent=None):
        super().__init__('Yeni Hedef Ekle', parent, 580)
        self.db = db; self._build()

    def _build(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(
            ['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)
        self.f_tip = self.combo(['Kilo','Güç','Kardiyo','Antrenman','Ölçüm','Genel'])
        self.add_row('Hedef Tipi', self.f_tip)
        self.f_ad = self.inp('Örn: 80kg Squat, 5km 25dk')
        self.add_row('Hedef Adı *', self.f_ad)
        self.f_bas = self.dspin(0, 9999, 0, 1)
        self.add_row('Başlangıç Değeri', self.f_bas)
        self.f_hedef = self.dspin(0, 9999, 0, 1)
        self.add_row('Hedef Değer *', self.f_hedef)
        self.f_mevcut = self.dspin(0, 9999, 0, 1)
        self.add_row('Mevcut Değer', self.f_mevcut)
        self.f_tarihi = QDateEdit(); self.f_tarihi.setCalendarPopup(True)
        self.f_tarihi.setDate(QDate.currentDate().addDays(90))
        self.f_tarihi.setFixedHeight(INPUT_H); self.f_tarihi.setFont(QFont('Segoe UI',13))
        self.add_row('Hedef Tarihi', self.f_tarihi)
        self.f_notlar = self.txt('Hedef notları...')
        self.add_row('Notlar', self.f_notlar)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        if self.f_sporcu.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Sporcu seçiniz!', QMessageBox.Warning); return
        if not self.f_ad.text().strip():
            dark_msg(self, 'Hata', 'Hedef adı zorunludur!', QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id':        self._sporcu_ids[self.f_sporcu.currentIndex()],
            'hedef_tipi':       self.f_tip.currentText(),
            'hedef_adi':        self.f_ad.text().strip(),
            'baslangic_degeri': self.f_bas.value(),
            'hedef_degeri':     self.f_hedef.value(),
            'mevcut_deger':     self.f_mevcut.value(),
            'hedef_tarihi':     self.f_tarihi.date().toString('yyyy-MM-dd'),
            'notlar':           self.f_notlar.toPlainText().strip(),
        }
        self.accept()


class HedefIlerlemeDialog(BaseDialog):
    def __init__(self, db, hedef, parent=None):
        super().__init__('İlerleme Güncelle', parent, 480)
        self.db = db; self.hedef = hedef; self._build()

    def _build(self):
        info = QLabel(f"🎯  {self.hedef['hedef_adi']}")
        info.setFont(QFont('Segoe UI', 14, QFont.Bold))
        info.setStyleSheet(f"color:{C['primary_light']};")
        self._main_layout.insertWidget(2, info)

        self.f_mevcut = self.dspin(0, 9999, self.hedef.get('mevcut_deger',0), 1)
        self.add_row('Mevcut Değer', self.f_mevcut)

        pct = 0
        if self.hedef['hedef_degeri'] > 0:
            pct = min(100, int(self.hedef.get('mevcut_deger',0) / self.hedef['hedef_degeri'] * 100))
        self.lbl_pct = QLabel(f"İlerleme: %{pct}")
        self.lbl_pct.setFont(QFont('Segoe UI', 13))
        self.lbl_pct.setStyleSheet(f"color:{C['success']};")
        self._main_layout.insertWidget(4, self.lbl_pct)

        self.f_mevcut.valueChanged.connect(self._update_pct)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Güncelle', C['primary'], self._save)

    def _update_pct(self, val):
        if self.hedef['hedef_degeri'] > 0:
            pct = min(100, int(val / self.hedef['hedef_degeri'] * 100))
            self.lbl_pct.setText(f"İlerleme: %{pct}")

    def _save(self):
        self.result_deger = self.f_mevcut.value()
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — QPainter Progress Bar Widget
# ═══════════════════════════════════════════════════════════════════
class ProgressBarWidget(QWidget):
    """Tek satır hedef progress bar — QPainter."""
    def __init__(self, hedef_adi, mevcut, hedef, tip, parent=None):
        super().__init__(parent)
        self.hedef_adi = hedef_adi
        self.mevcut    = mevcut
        self.hedef_val = hedef
        self.tip       = tip
        self.setFixedHeight(58)

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        p.fillRect(0, 0, w, h, QColor(C['bg_card']))

        pct = min(1.0, self.mevcut / self.hedef_val) if self.hedef_val else 0
        renk_map = {
            'Kilo': C['info'], 'Güç': C['danger'], 'Kardiyo': C['success'],
            'Antrenman': C['primary'], 'Ölçüm': C['warning'], 'Genel': C['primary_light']
        }
        renk = renk_map.get(self.tip, C['primary'])

        # Hedef adı
        p.setPen(QColor(C['text_main']))
        p.setFont(QFont('Segoe UI', 12, QFont.Bold))
        p.drawText(QRectF(8, 4, w * 0.6, 20), Qt.AlignLeft | Qt.AlignVCenter, self.hedef_adi)

        # Yüzde + değer
        p.setPen(QColor(renk))
        p.setFont(QFont('Segoe UI', 11, QFont.Bold))
        p.drawText(QRectF(w * 0.6, 4, w * 0.4 - 6, 20),
                   Qt.AlignRight | Qt.AlignVCenter,
                   f"{self.mevcut:.1f} / {self.hedef_val:.1f}  ({int(pct*100)}%)")

        # Bar arka plan
        bx, by, bw, bh = 8, 30, w - 16, 16
        p.setBrush(QColor(C['border'])); p.setPen(Qt.NoPen)
        p.drawRoundedRect(int(bx), by, bw, bh, 8, 8)

        # Bar dolu kısım
        fill_w = int(bw * pct)
        if fill_w > 0:
            grad = QLinearGradient(bx, 0, bx + bw, 0)
            grad.setColorAt(0, QColor(renk))
            grad.setColorAt(1, QColor(QColor(renk).lighter(130).name()))
            p.setBrush(QBrush(grad)); p.setPen(Qt.NoPen)
            p.drawRoundedRect(int(bx), by, fill_w, bh, 8, 8)

        # %100 ise tik
        if pct >= 1.0:
            p.setPen(QColor(C['success']))
            p.setFont(QFont('Segoe UI', 13))
            p.drawText(QRectF(w - 28, 4, 22, 22), Qt.AlignCenter, '✅')
        p.end()


# ═══════════════════════════════════════════════════════════════════
# TIER 3 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class VucutOlcumleriPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_sporcu_id = None
        self._build(); self._load_sporcular()

    def _build(self):
        from PyQt5.QtWidgets import QSplitter, QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Vücut Ölçümleri')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_new = make_btn('+ Ölçüm Ekle', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # Sporcu seç
        fil = QHBoxLayout()
        fil.addWidget(QLabel('Sporcu:'))
        self.sporcu_cb = make_combo(['-- Sporcu Seç --'], 220)
        self.sporcu_cb.currentIndexChanged.connect(self._on_sporcu_change)
        fil.addWidget(self.sporcu_cb)
        fil.addWidget(QLabel('Alan:'))
        self.alan_cb = make_combo(['kilo','bel','gogus','kalca','sag_kol','vucut_yag_orani'], 160)
        self.alan_cb.currentTextChanged.connect(self._refresh_chart)
        fil.addWidget(self.alan_cb); fil.addStretch()
        lay.addLayout(fil)

        # Split: grafik sol, tablo sağ
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C['border']};width:2px;}}")

        # Sol: Grafik
        left = QFrame()
        left.setStyleSheet(f"background:{C['bg_card']};border-radius:12px;border:1px solid {C['border']};")
        ll = QVBoxLayout(left); ll.setContentsMargins(10,10,10,10)
        self.chart = LineChartWidget()
        self.chart.setMinimumHeight(260)
        ll.addWidget(self.chart)
        splitter.addWidget(left)

        # Sağ: Tablo
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(8)
        lbl_t = QLabel('Ölçüm Geçmişi')
        lbl_t.setFont(QFont('Segoe UI', 14, QFont.Bold))
        rl.addWidget(lbl_t)
        self.table = make_table(
            ['ID','Tarih','Kilo','Bel','Göğüs','Kalça','Kol','Bacak','Yağ%'],
            [0, 95, 70, 65, 75, 70, 65, 75, 65]
        )
        self.table.hideColumn(0)
        rl.addWidget(self.table)
        splitter.addWidget(right)
        splitter.setSizes([420, 560])
        lay.addWidget(splitter)

    def _load_sporcular(self):
        self._sporcu_ids = [None]
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_cb.blockSignals(True)
        self.sporcu_cb.clear()
        self.sporcu_cb.addItem('-- Sporcu Seç --')
        for s in sporcular:
            self.sporcu_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])
        self.sporcu_cb.blockSignals(False)

    def _on_sporcu_change(self):
        idx = self.sporcu_cb.currentIndex()
        self.selected_sporcu_id = self._sporcu_ids[idx] if idx > 0 else None
        self.refresh()

    def refresh(self):
        self._load_sporcular()
        if not self.selected_sporcu_id:
            self.table.setRowCount(0); self.chart.set_data([], 'Sporcu seçin'); return
        data = self.db.get_vucut_olcumleri(self.selected_sporcu_id)
        self.table.setRowCount(0)
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['olcum_id']), d['tarih'],
                    f"{d['kilo']:.1f}" if d['kilo'] else '—',
                    f"{d['bel']:.1f}" if d['bel'] else '—',
                    f"{d['gogus']:.1f}" if d['gogus'] else '—',
                    f"{d['kalca']:.1f}" if d['kalca'] else '—',
                    f"{d['sag_kol']:.1f}" if d['sag_kol'] else '—',
                    f"{d['sag_bacak']:.1f}" if d['sag_bacak'] else '—',
                    f"{d['vucut_yag_orani']:.1f}%" if d['vucut_yag_orani'] else '—']
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                self.table.setItem(row, col, item)
        self._refresh_chart()

    def _refresh_chart(self):
        if not self.selected_sporcu_id: return
        alan = self.alan_cb.currentText()
        alan_label = {
            'kilo':'Kilo (kg)','bel':'Bel (cm)','gogus':'Göğüs (cm)',
            'kalca':'Kalça (cm)','sag_kol':'Kol (cm)','vucut_yag_orani':'Vücut Yağ%'
        }.get(alan, alan)
        trend = self.db.get_vucut_trend(self.selected_sporcu_id, alan)
        renk_map = {'kilo':C['primary'],'bel':C['warning'],'gogus':C['info'],
                    'kalca':C['accent'],'sag_kol':C['success'],'vucut_yag_orani':C['danger']}
        self.chart.set_data(trend, alan_label, '', renk_map.get(alan, C['primary']))

    def _add(self):
        dlg = VucutOlcumDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_vucut_olcumu(dlg.result_data)
            dark_msg(self, 'Başarılı', 'Ölçüm kaydedildi!')
            sid = dlg.result_data['sporcu_id']
            # Sporcuyu seçili yap
            for i, s_id in enumerate(self._sporcu_ids):
                if s_id == sid:
                    self.sporcu_cb.setCurrentIndex(i); break
            else:
                self.refresh()


class HedeflerPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        from PyQt5.QtWidgets import QScrollArea
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Hedefler')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Hedef', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # Filtreler
        fil = QHBoxLayout()
        self.search = make_search('Hedef adı veya sporcu ara...')
        self.search.textChanged.connect(lambda: self._timer.start(300))
        fil.addWidget(self.search, 1)
        fil.addWidget(QLabel('Durum:'))
        self.durum_cb = make_combo(['Tümü','Devam','Tamamlandi'], 150)
        self.durum_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.durum_cb)
        fil.addWidget(QLabel('Tip:'))
        self.tip_cb = make_combo(['Tümü','Kilo','Güç','Kardiyo','Antrenman','Ölçüm','Genel'], 130)
        self.tip_cb.currentTextChanged.connect(self.refresh)
        fil.addWidget(self.tip_cb)
        lay.addLayout(fil)

        # Progress barlar (scroll area)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet(f"QScrollArea{{border:none;background:{C['bg_main']};}}")
        self.scroll_widget = QWidget()
        self.scroll_widget.setStyleSheet(f"background:{C['bg_main']};")
        self.scroll_lay = QVBoxLayout(self.scroll_widget)
        self.scroll_lay.setSpacing(6)
        self.scroll_lay.setContentsMargins(0,0,0,0)
        self.scroll.setWidget(self.scroll_widget)
        lay.addWidget(self.scroll, 1)

        # Tablo (detay için)
        self.table = make_table(
            ['ID','Sporcu','Hedef','Tip','Başlangıç','Hedef','Mevcut','Tarih','Durum'],
            [0, 140, -1, 90, 85, 85, 85, 100, 100]
        )
        self.table.hideColumn(0)
        self.table.setMaximumHeight(200)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_update = make_btn('İlerleme Güncelle', C['primary']); self.btn_update.setEnabled(False)
        self.btn_update.clicked.connect(self._update_ilerleme)
        self.btn_del = make_btn('Sil', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_update); bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        search = self.search.text().lower()
        durum  = self.durum_cb.currentText()
        tip    = self.tip_cb.currentText()
        data   = self.db.get_hedefler(durum=durum)

        # Filtrele
        if tip != 'Tümü':
            data = [d for d in data if d['hedef_tipi'] == tip]
        if search:
            data = [d for d in data if search in d['hedef_adi'].lower()
                    or search in d.get('sporcu_adi','').lower()]

        self.lbl_count.setText(f"{len(data)} hedef")

        # Progress barları temizle
        while self.scroll_lay.count():
            item = self.scroll_lay.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        # Progress barları çiz
        for d in data:
            frame = QFrame()
            frame.setStyleSheet(f"""
                QFrame{{background:{C['bg_card']};border-radius:10px;
                border:1px solid {C['border']};margin:2px 0;}}
            """)
            fl = QHBoxLayout(frame); fl.setContentsMargins(10,4,10,4)
            # Sporcu adı
            lbl_s = QLabel(d.get('sporcu_adi',''))
            lbl_s.setFont(QFont('Segoe UI', 11))
            lbl_s.setStyleSheet(f"color:{C['text_muted']};min-width:120px;")
            fl.addWidget(lbl_s)
            # Progress bar
            pb = ProgressBarWidget(d['hedef_adi'], d.get('mevcut_deger',0),
                                   d['hedef_degeri'], d['hedef_tipi'])
            fl.addWidget(pb, 1)
            self.scroll_lay.addWidget(frame)

        self.scroll_lay.addStretch()

        # Tablo
        self.table.setRowCount(0)
        durum_renk = {'Devam': C['warning'], 'Tamamlandi': C['success']}
        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [str(d['hedef_id']), d.get('sporcu_adi','—'), d['hedef_adi'],
                    d['hedef_tipi'], f"{d.get('baslangic_degeri',0):.1f}",
                    f"{d['hedef_degeri']:.1f}", f"{d.get('mevcut_deger',0):.1f}",
                    d.get('hedef_tarihi','—'), d.get('durum','Devam')]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 8:
                    item.setForeground(QColor(durum_renk.get(val, C['text_muted'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_id = None
        self.btn_update.setEnabled(False); self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(), 0).text())
            self.btn_update.setEnabled(True); self.btn_del.setEnabled(True)

    def _add(self):
        dlg = HedefDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_hedef(dlg.result_data)
            dark_msg(self, 'Başarılı', 'Hedef eklendi!'); self.refresh()

    def _update_ilerleme(self):
        if not self.selected_id: return
        h = self.db.get_hedef(self.selected_id)
        if not h: return
        dlg = HedefIlerlemeDialog(self.db, h, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.update_hedef_ilerleme(self.selected_id, dlg.result_deger)
            self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        name = self.table.item(row, 2).text()
        if dark_confirm(self, 'Sil', f'"{name}" hedefi silinecek. Emin misiniz?'):
            self.db.delete_hedef(self.selected_id); self.refresh()


class StreakPage(QWidget):
    """Sporcu streak tablosu + mini streak gösterge."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Antrenman Streakları')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        btn_ref = make_btn('⟳  Yenile', C['primary'], small=True)
        btn_ref.clicked.connect(self.refresh); hdr.addWidget(btn_ref)
        lay.addLayout(hdr)

        # KPI - Top 3 streak (sabit yükseklik container)
        kpi_container = QWidget(); kpi_container.setFixedHeight(105)
        self.kpi_row = QHBoxLayout(kpi_container)
        self.kpi_row.setSpacing(SPACING); self.kpi_row.setContentsMargins(0,0,0,0)
        self.k1 = KPICard('🥇 En Uzun Streak', '—', '🔥', '#f59e0b', '#d97706')
        self.k2 = KPICard('🥈 2. Sıra', '—', '🔥', '#6366f1', '#4f46e5')
        self.k3 = KPICard('🥉 3. Sıra', '—', '🔥', '#10b981', '#059669')
        for k in [self.k1, self.k2, self.k3]:
            self.kpi_row.addWidget(k)
        lay.addWidget(kpi_container)

        # Tablo — kalan tüm alanı al
        self.table = make_table(
            ['Sıra','Sporcu','Streak (Gün)','Durum'],
            [80, -1, 150, 120]
        )
        lay.addWidget(self.table, 1)

        # Sporcu bazlı streak ekle
        bot = QHBoxLayout()
        bot.addWidget(QLabel('Bugün antrenman yaptı:'))
        self.sporcu_streak_cb = make_combo(['-- Sporcu Seç --'], 200)
        bot.addWidget(self.sporcu_streak_cb)
        btn_streak = make_btn('✅ Streak Ekle', C['success'])
        btn_streak.clicked.connect(self._add_streak)
        bot.addWidget(btn_streak); bot.addStretch()
        lay.addLayout(bot)

        self._sporcu_ids = [None]

    def refresh(self):
        data = self.db.get_tum_streakler()
        self.table.setRowCount(0)

        # KPI güncelle — gün sayısı büyük, sporcu adı başlıkta
        kpi_titles = ['🥇 En Uzun Streak', '🥈 2. Sıra', '🥉 3. Sıra']
        for i, (k, d) in enumerate(zip([self.k1, self.k2, self.k3], data[:3])):
            k.title_text = f"{kpi_titles[i]}  •  {d['ad']} {d['soyad']}"
            k.set_value(f"{d['streak']} gün")

        for sira, d in enumerate(data, 1):
            row = self.table.rowCount(); self.table.insertRow(row)
            streak = d['streak']
            # Alev emojisi streak'e göre
            emoji = '🔥' * min(3, max(1, streak // 7)) if streak > 0 else '😴'
            durum = f"{emoji} {'Aktif' if streak > 0 else 'Pasif'}"
            sira_lbl = {1:'🥇',2:'🥈',3:'🥉'}.get(sira, str(sira))
            vals = [sira_lbl, f"{d['ad']} {d['soyad']}", str(streak), durum]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 0:
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(QFont('Segoe UI', 15 if sira <= 3 else 13))
                if col == 2:
                    renk = C['success'] if streak >= 7 else (C['warning'] if streak >= 3 else C['text_muted'])
                    item.setForeground(QColor(renk))
                    item.setFont(QFont('Segoe UI', 14, QFont.Bold))
                self.table.setItem(row, col, item)

        # Sporcu combo güncelle
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.sporcu_streak_cb.clear()
        self.sporcu_streak_cb.addItem('-- Sporcu Seç --')
        self._sporcu_ids = [None]
        for s in sporcular:
            self.sporcu_streak_cb.addItem(f"{s['ad']} {s['soyad']}")
            self._sporcu_ids.append(s['sporcu_id'])

    def _add_streak(self):
        idx = self.sporcu_streak_cb.currentIndex()
        if idx == 0:
            dark_msg(self, 'Hata', 'Sporcu seçiniz!', QMessageBox.Warning); return
        sporcu_id = self._sporcu_ids[idx]
        eklendi = self.db.add_streak(sporcu_id)
        if eklendi:
            dark_msg(self, 'Başarılı', 'Bugün için streak eklendi! 🔥')
        else:
            dark_msg(self, 'Bilgi', 'Bu sporcu için bugün zaten streak eklenmiş.')
        self.refresh()



class TakipDialog(BaseDialog):
    """Antrenman takibi ekle — sporcu + antrenman seç, set/rep/kalori gir."""
    def __init__(self, db, parent=None):
        super().__init__('Yeni Antrenman Takibi', parent, 620)
        self.db = db; self._build()

    def _build(self):
        # Sporcu
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)
        # Antrenman
        antrenmanlar = self.db.get_antrenmanlar()
        self.f_antrenman = self.combo(['-- Antrenman Seç --'] + [f"{a['ad']} ({a['kategori']})" for a in antrenmanlar])
        self._ant_ids = [None] + [a['antrenman_id'] for a in antrenmanlar]
        self.add_row('Antrenman *', self.f_antrenman)
        # Tarih
        self.f_tarih = QDateEdit(); self.f_tarih.setCalendarPopup(True)
        self.f_tarih.setDate(QDate.currentDate()); self.f_tarih.setFixedHeight(INPUT_H)
        self.f_tarih.setFont(QFont('Segoe UI', 13))
        self.add_row('Tarih', self.f_tarih)
        # Süre
        self.f_sure = self.spin(0, 300, 45)
        self.f_sure.setSuffix(' dakika')
        self.add_row('Süre', self.f_sure)
        # Kalori
        self.f_kalori = self.spin(0, 2000, 300)
        self.f_kalori.setSuffix(' kcal')
        self.add_row('Yakılan Kalori', self.f_kalori)
        # Nabız
        nabiz_w = QWidget(); nabiz_l = QHBoxLayout(nabiz_w); nabiz_l.setContentsMargins(0,0,0,0)
        self.f_max_nabiz = self.spin(0, 220, 160); self.f_max_nabiz.setPrefix('Max: ')
        self.f_ort_nabiz = self.spin(0, 220, 130); self.f_ort_nabiz.setPrefix('Ort: ')
        nabiz_l.addWidget(self.f_max_nabiz); nabiz_l.addWidget(self.f_ort_nabiz)
        self.add_row('Nabız (bpm)', nabiz_w)
        # Zorluk hissi
        self.f_zorluk = self.spin(1, 10, 5)
        self.f_zorluk.setSuffix(' / 10')
        self.add_row('Zorluk Hissi', self.f_zorluk)
        # Notlar
        self.f_notlar = self.txt('Antrenman notları...')
        self.add_row('Notlar', self.f_notlar)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Kaydet', C['success'], self._save)

    def _save(self):
        if self.f_sporcu.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Sporcu seçiniz!', QMessageBox.Warning); return
        if self.f_antrenman.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Antrenman seçiniz!', QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id':      self._sporcu_ids[self.f_sporcu.currentIndex()],
            'antrenman_id':   self._ant_ids[self.f_antrenman.currentIndex()],
            'tarih':          self.f_tarih.date().toString('yyyy-MM-dd'),
            'sure_dakika':    self.f_sure.value(),
            'kalori_yakilan': self.f_kalori.value(),
            'max_nabiz':      self.f_max_nabiz.value(),
            'ort_nabiz':      self.f_ort_nabiz.value(),
            'zorluk_hissi':   self.f_zorluk.value(),
            'notlar':         self.f_notlar.toPlainText().strip(),
        }
        self.accept()


class ProgramDialog(BaseDialog):
    """Antrenman programı oluştur."""
    def __init__(self, db, parent=None):
        super().__init__('Yeni Antrenman Programı', parent, 580)
        self.db = db; self._build()

    def _build(self):
        sporcular = self.db.get_sporcular(durum='Aktif')
        self.f_sporcu = self.combo(['-- Sporcu Seç --'] + [f"{s['ad']} {s['soyad']}" for s in sporcular])
        self._sporcu_ids = [None] + [s['sporcu_id'] for s in sporcular]
        self.add_row('Sporcu *', self.f_sporcu)
        self.f_ad = self.inp('Örn: 4 Haftalık Kilo Verme Programı')
        self.add_row('Program Adı *', self.f_ad)
        self.f_hedef = self.combo(['Kilo Verme','Kas Kazanma','Form Koruma','Dayanıklılık','Genel Sağlık'])
        self.add_row('Hedef', self.f_hedef)
        self.f_sure = self.spin(1, 52, 4)
        self.f_sure.setSuffix(' hafta')
        self.add_row('Program Süresi', self.f_sure)
        self.f_notlar = self.txt('Program hakkında notlar...')
        self.add_row('Notlar', self.f_notlar)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Oluştur', C['success'], self._save)

    def _save(self):
        if self.f_sporcu.currentIndex() == 0:
            dark_msg(self, 'Hata', 'Sporcu seçiniz!', QMessageBox.Warning); return
        if not self.f_ad.text().strip():
            dark_msg(self, 'Hata', 'Program adı zorunludur!', QMessageBox.Warning); return
        self.result_data = {
            'sporcu_id':  self._sporcu_ids[self.f_sporcu.currentIndex()],
            'program_adi': self.f_ad.text().strip(),
            'hedef':      self.f_hedef.currentText(),
            'sure_hafta': self.f_sure.value(),
            'notlar':     self.f_notlar.toPlainText().strip(),
        }
        self.accept()


class ProgramGunuDialog(BaseDialog):
    """Programa gün/egzersiz ekle."""
    def __init__(self, db, program_id, parent=None):
        super().__init__('Programa Antrenman Ekle', parent, 560)
        self.db = db; self.program_id = program_id; self._build()

    def _build(self):
        self.f_gun = self.combo(['Pazartesi','Salı','Çarşamba','Perşembe','Cuma','Cumartesi','Pazar'])
        self.add_row('Gün', self.f_gun)
        antrenmanlar = self.db.get_antrenmanlar()
        self.f_ant = self.combo([f"{a['ad']} ({a['kategori']})" for a in antrenmanlar])
        self._ant_ids = [a['antrenman_id'] for a in antrenmanlar]
        self.add_row('Antrenman *', self.f_ant)
        self.f_set = self.spin(1, 20, 3)
        self.add_row('Set Sayısı', self.f_set)
        self.f_tekrar = self.spin(1, 100, 12)
        self.add_row('Tekrar', self.f_tekrar)
        self.f_agirlik = self.dspin(0, 500, 0, 1)
        self.f_agirlik.setSuffix(' kg')
        self.add_row('Ağırlık', self.f_agirlik)
        self.f_dinlenme = self.spin(15, 300, 60)
        self.f_dinlenme.setSuffix(' sn')
        self.add_row('Dinlenme Süresi', self.f_dinlenme)
        self.add_btn('İptal', C['text_muted'], self.reject)
        self.add_btn('Ekle', C['success'], self._save)

    def _save(self):
        if not self._ant_ids:
            dark_msg(self, 'Hata', 'Antrenman bulunamadı!', QMessageBox.Warning); return
        self.result_data = {
            'program_id':       self.program_id,
            'gun_adi':          self.f_gun.currentText(),
            'antrenman_id':     self._ant_ids[self.f_ant.currentIndex()],
            'set_sayisi':       self.f_set.value(),
            'tekrar':           self.f_tekrar.value(),
            'agirlik':          self.f_agirlik.value(),
            'dinlenme_saniye':  self.f_dinlenme.value(),
        }
        self.accept()


# ═══════════════════════════════════════════════════════════════════
# TIER 2 — SAYFALAR
# ═══════════════════════════════════════════════════════════════════

class TakiplerPage(QWidget):
    """Antrenman takip kaydı — tüm log listesi + yeni ekle."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None
        self._build()
        self._timer = QTimer(); self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.refresh)
        self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        # Header
        hdr = QHBoxLayout()
        title = QLabel('Antrenman Takipleri')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('0 takip')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Takip', C['success'])
        btn_new.clicked.connect(self._add); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        # KPI satırı
        kpi = QHBoxLayout(); kpi.setSpacing(SPACING)
        self.k_toplam  = KPICard('Toplam Takip',   '—', '📋', '#6366f1', '#4f46e5')
        self.k_hafta   = KPICard('Bu Hafta',        '—', '🗓', '#10b981', '#059669')
        self.k_kalori  = KPICard('Toplam Kalori',   '—', '🔥', '#f59e0b', '#d97706')
        self.k_sure    = KPICard('Ort. Süre',       '—', '⏱', '#3b82f6', '#2563eb')
        for w in [self.k_toplam, self.k_hafta, self.k_kalori, self.k_sure]:
            kpi.addWidget(w)
        lay.addLayout(kpi)

        # Filtreler
        filt = QHBoxLayout(); filt.setSpacing(10)
        self.search = make_search('Sporcu adı veya antrenman ara...')
        self.search.textChanged.connect(lambda: self._timer.start(300))
        filt.addWidget(self.search, 1)
        lay.addLayout(filt)

        # Tablo
        self.table = make_table(
            ['ID','Sporcu','Antrenman','Kategori','Tarih','Süre','Kalori','Maks Nabız','Zorluk'],
            [0, -1, 160, 100, 100, 80, 90, 100, 80]
        )
        self.table.hideColumn(0)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table)

        bot = QHBoxLayout()
        self.btn_del = make_btn('Sil', C['danger']); self.btn_del.setEnabled(False)
        self.btn_del.clicked.connect(self._delete)
        bot.addWidget(self.btn_del); bot.addStretch()
        lay.addLayout(bot)

    def refresh(self):
        data = self.db.get_gunluk_log(search=self.search.text())
        self.lbl_count.setText(f"{len(data)} takip")
        self.table.setRowCount(0)

        # KPI güncelle
        with self.db.get_connection() as conn:
            hafta = conn.execute(
                "SELECT COUNT(*) FROM gunluk_log WHERE durum='Tamamlandi' AND tarih>=date('now','-7 days')"
            ).fetchone()[0]
            toplam_k = conn.execute(
                "SELECT COALESCE(SUM(kalori_yakilan),0) FROM gunluk_log WHERE durum='Tamamlandi'"
            ).fetchone()[0]
            ort_s = conn.execute(
                "SELECT COALESCE(AVG(sure_dakika),0) FROM gunluk_log WHERE durum='Tamamlandi'"
            ).fetchone()[0]
        self.k_toplam.set_value(str(len(data)))
        self.k_hafta.set_value(str(hafta))
        self.k_kalori.set_value(f"{toplam_k:,}")
        self.k_sure.set_value(f"{round(ort_s)}dk")

        zorluk_renkler = {
            range(1,4): C['success'], range(4,7): C['warning'], range(7,11): C['danger']
        }
        def zorluk_rengi(z):
            for r, c in zorluk_renkler.items():
                if z in r: return c
            return C['text_main']

        for d in data:
            row = self.table.rowCount(); self.table.insertRow(row)
            vals = [
                str(d['log_id']), d['sporcu_adi'], d['antrenman_adi'],
                d.get('kategori','—'), d['tarih'],
                f"{d['sure_dakika']} dk", f"{d['kalori_yakilan']} kcal",
                f"{d['max_nabiz']} bpm" if d['max_nabiz'] else '—',
                f"{d['zorluk_hissi']}/10",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 8:  # zorluk rengi
                    z = d['zorluk_hissi']
                    item.setForeground(QColor(zorluk_rengi(z)))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.table.setItem(row, col, item)
        self.selected_id = None; self.btn_del.setEnabled(False)

    def _on_select(self):
        if self.table.selectedItems():
            self.selected_id = int(self.table.item(self.table.currentRow(), 0).text())
            self.btn_del.setEnabled(True)

    def _add(self):
        dlg = TakipDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_gunluk_log(dlg.result_data)
            dark_msg(self, 'Başarılı', 'Takip kaydedildi!')
            self.refresh()

    def _delete(self):
        if not self.selected_id: return
        row = self.table.currentRow()
        sporcu = self.table.item(row, 1).text()
        ant = self.table.item(row, 2).text()
        if dark_confirm(self, 'Sil', f'"{sporcu} — {ant}" takibi silinecek. Emin misiniz?'):
            self.db.delete_gunluk_log(self.selected_id); self.refresh()


class ProgramlarPage(QWidget):
    """Antrenman programları — oluştur, detay gör, gün ekle."""
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.selected_id = None; self.selected_gun_id = None
        self._build(); self.refresh()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(PADDING, PADDING, PADDING, PADDING)
        lay.setSpacing(SPACING)

        hdr = QHBoxLayout()
        title = QLabel('Antrenman Programları')
        title.setFont(QFont('Segoe UI', 24, QFont.Bold))
        hdr.addWidget(title); hdr.addStretch()
        self.lbl_count = QLabel('0 program')
        self.lbl_count.setStyleSheet(f"color:{C['text_muted']};")
        hdr.addWidget(self.lbl_count)
        btn_new = make_btn('+ Yeni Program', C['success'])
        btn_new.clicked.connect(self._add_program); hdr.addWidget(btn_new)
        lay.addLayout(hdr)

        from PyQt5.QtWidgets import QSplitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.setStyleSheet(f"QSplitter::handle {{ background: {C['border']}; width: 2px; }}")

        # Sol: Program listesi
        left = QWidget()
        left_lay = QVBoxLayout(left); left_lay.setContentsMargins(0,0,0,0); left_lay.setSpacing(8)
        lbl_prog = QLabel('Programlar')
        lbl_prog.setFont(QFont('Segoe UI', 14, QFont.Bold))
        left_lay.addWidget(lbl_prog)
        self.tbl_prog = make_table(
            ['ID','Sporcu','Program','Hedef','Süre','Aktif'],
            [0, 130, -1, 110, 70, 60]
        )
        self.tbl_prog.hideColumn(0)
        self.tbl_prog.itemSelectionChanged.connect(self._on_prog_select)
        left_lay.addWidget(self.tbl_prog)
        bot_prog = QHBoxLayout()
        self.btn_del_prog = make_btn('Sil', C['danger'], small=True)
        self.btn_del_prog.setEnabled(False)
        self.btn_del_prog.clicked.connect(self._del_program)
        bot_prog.addWidget(self.btn_del_prog); bot_prog.addStretch()
        left_lay.addLayout(bot_prog)
        splitter.addWidget(left)

        # Sağ: Program günleri
        right = QWidget()
        right_lay = QVBoxLayout(right); right_lay.setContentsMargins(0,0,0,0); right_lay.setSpacing(8)
        rh = QHBoxLayout()
        lbl_gun = QLabel('Program Günleri')
        lbl_gun.setFont(QFont('Segoe UI', 14, QFont.Bold))
        rh.addWidget(lbl_gun); rh.addStretch()
        self.btn_add_gun = make_btn('+ Gün Ekle', C['primary'], small=True)
        self.btn_add_gun.setEnabled(False)
        self.btn_add_gun.clicked.connect(self._add_gun)
        rh.addWidget(self.btn_add_gun)
        right_lay.addLayout(rh)
        self.tbl_gun = make_table(
            ['ID','Gün','Antrenman','Kategori','Set','Tekrar','Ağırlık','Dinlenme'],
            [0, 110, -1, 100, 55, 65, 80, 90]
        )
        self.tbl_gun.hideColumn(0)
        self.tbl_gun.itemSelectionChanged.connect(self._on_gun_select)
        right_lay.addWidget(self.tbl_gun)
        bot_gun = QHBoxLayout()
        self.btn_del_gun = make_btn('Günü Sil', C['danger'], small=True)
        self.btn_del_gun.setEnabled(False)
        self.btn_del_gun.clicked.connect(self._del_gun)
        bot_gun.addWidget(self.btn_del_gun); bot_gun.addStretch()
        right_lay.addLayout(bot_gun)
        splitter.addWidget(right)

        splitter.setSizes([450, 600])
        lay.addWidget(splitter)

    def refresh(self):
        data = self.db.get_programlar()
        self.lbl_count.setText(f"{len(data)} program")
        self.tbl_prog.setRowCount(0)
        for d in data:
            row = self.tbl_prog.rowCount(); self.tbl_prog.insertRow(row)
            vals = [str(d['program_id']), d['sporcu_adi'], d['program_adi'],
                    d.get('hedef','—'), f"{d.get('sure_hafta','?')} Hafta",
                    '✅' if d.get('aktif') else '⏸']
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 5:
                    item.setForeground(QColor(C['success'] if d.get('aktif') else C['text_muted']))
                self.tbl_prog.setItem(row, col, item)
        self.selected_id = None
        self.btn_del_prog.setEnabled(False)
        self.btn_add_gun.setEnabled(False)
        self.tbl_gun.setRowCount(0)

    def _load_gunler(self, program_id):
        gunler = self.db.get_program_gunleri(program_id)
        self.tbl_gun.setRowCount(0)
        gun_renk = {
            'Pazartesi':'#6366f1','Salı':'#10b981','Çarşamba':'#f59e0b',
            'Perşembe':'#3b82f6','Cuma':'#ec4899','Cumartesi':'#14b8a6','Pazar':'#8b5cf6'
        }
        for d in gunler:
            row = self.tbl_gun.rowCount(); self.tbl_gun.insertRow(row)
            ag = f"{d['agirlik']} kg" if d['agirlik'] else '—'
            vals = [str(d['gun_id']), d['gun_adi'], d['antrenman_adi'],
                    d.get('kategori','—'), str(d['set_sayisi']),
                    str(d['tekrar']), ag, f"{d['dinlenme_saniye']}sn"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFont(QFont('Segoe UI', 13))
                if col == 1:
                    item.setForeground(QColor(gun_renk.get(val, C['text_main'])))
                    item.setFont(QFont('Segoe UI', 13, QFont.Bold))
                self.tbl_gun.setItem(row, col, item)

    def _on_prog_select(self):
        if self.tbl_prog.selectedItems():
            self.selected_id = int(self.tbl_prog.item(self.tbl_prog.currentRow(), 0).text())
            self.btn_del_prog.setEnabled(True)
            self.btn_add_gun.setEnabled(True)
            self._load_gunler(self.selected_id)

    def _on_gun_select(self):
        if self.tbl_gun.selectedItems():
            self.selected_gun_id = int(self.tbl_gun.item(self.tbl_gun.currentRow(), 0).text())
            self.btn_del_gun.setEnabled(True)

    def _add_program(self):
        dlg = ProgramDialog(self.db, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_program(dlg.result_data)
            dark_msg(self, 'Başarılı', 'Program oluşturuldu!')
            self.refresh()

    def _del_program(self):
        if not self.selected_id: return
        row = self.tbl_prog.currentRow()
        name = self.tbl_prog.item(row, 2).text()
        if dark_confirm(self, 'Sil', f'"{name}" programı silinecek. Emin misiniz?'):
            self.db.delete_program(self.selected_id); self.refresh()

    def _add_gun(self):
        if not self.selected_id: return
        dlg = ProgramGunuDialog(self.db, self.selected_id, self)
        if dlg.exec_() == QDialog.Accepted:
            self.db.add_program_gunu(dlg.result_data)
            self._load_gunler(self.selected_id)

    def _del_gun(self):
        if not self.selected_gun_id: return
        if dark_confirm(self, 'Sil', 'Bu antrenman günü silinecek. Emin misiniz?'):
            self.db.delete_program_gunu(self.selected_gun_id)
            self._load_gunler(self.selected_id)
            self.selected_gun_id = None
            self.btn_del_gun.setEnabled(False)



class LoginWindow(QWidget):
    def __init__(self, db, on_login):
        super().__init__()
        self.db = db; self.on_login = on_login
        self.setWindowTitle('Fitness Takip Sistemi')
        self.setFixedSize(480, 520)
        self.setStyleSheet(f"background-color: {C['bg_main']};")
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(50, 40, 50, 40); lay.setSpacing(14)

        ico = QLabel('🏋️')
        ico.setFont(QFont('Segoe UI', 48))
        ico.setAlignment(Qt.AlignCenter); lay.addWidget(ico)

        t1 = QLabel('Fitness Takip Sistemi')
        t1.setFont(QFont('Segoe UI', 17, QFont.Bold)); t1.setAlignment(Qt.AlignCenter)
        t1.setStyleSheet(f"color: {C['text_main']};"); lay.addWidget(t1)

        t2 = QLabel('v1.0 — Dark Luxury Edition')
        t2.setFont(QFont('Segoe UI', 10)); t2.setAlignment(Qt.AlignCenter)
        t2.setStyleSheet(f"color: {C['text_muted']};"); lay.addWidget(t2)

        lay.addSpacing(14)

        def lbl(text):
            l = QLabel(text); l.setFont(QFont('Segoe UI', 12, QFont.Bold))
            l.setStyleSheet(f"color: {C['text_secondary']};"); return l

        inp_style = f"""
            QLineEdit {{
                background: {C['bg_secondary']}; color: {C['text_main']};
                border: 1.5px solid {C['border']}; border-radius: 10px;
                padding: 0 14px; font-size: 13px;
            }}
            QLineEdit:focus {{ border: 2px solid {C['primary']}; }}
        """

        lay.addWidget(lbl('Kullanıcı Adı'))
        self.inp_user = QLineEdit(); self.inp_user.setPlaceholderText('admin')
        self.inp_user.setFixedHeight(46); self.inp_user.setStyleSheet(inp_style)
        lay.addWidget(self.inp_user)

        lay.addWidget(lbl('Şifre'))
        self.inp_pass = QLineEdit(); self.inp_pass.setPlaceholderText('Şifrenizi girin')
        self.inp_pass.setEchoMode(QLineEdit.Password)
        self.inp_pass.setFixedHeight(46); self.inp_pass.setStyleSheet(inp_style)
        self.inp_pass.returnPressed.connect(self._login)
        lay.addWidget(self.inp_pass)

        self.lbl_err = QLabel(''); self.lbl_err.setAlignment(Qt.AlignCenter)
        self.lbl_err.setStyleSheet(f"color: {C['danger']}; font-size: 12px;")
        lay.addWidget(self.lbl_err)

        btn = QPushButton('Giriş Yap')
        btn.setFixedHeight(50); btn.setFont(QFont('Segoe UI', 13, QFont.Bold))
        btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {C['primary']}, stop:1 {C['primary_light']});
                color: white; border: none; border-radius: 10px;
                font-size: 14px; font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {C['primary_light']}, stop:1 {C['primary']});
            }}
        """)
        btn.clicked.connect(self._login); lay.addWidget(btn)
        lay.addStretch()

        info = QLabel('admin / admin123  |  personel / personel123')
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet(f"color: {C['text_muted']}; font-size: 10px;")
        lay.addWidget(info)

    def _login(self):
        u = self.inp_user.text().strip(); p = self.inp_pass.text().strip()
        if not u or not p:
            self.lbl_err.setText('Kullanıcı adı ve şifre gerekli!'); return
        user = self.db.authenticate(u, p)
        if user:
            self.on_login(user)
        else:
            self.lbl_err.setText('Kullanıcı adı veya şifre hatalı!')
            self.inp_pass.clear(); self.inp_pass.setFocus()


# ═══════════════════════════════════════════════════════════════════
# ANA PENCERE
# ═══════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self, db, user):
        super().__init__()
        self.db = db; self.user = user
        self.setWindowTitle(f"Fitness Takip Sistemi — {user['ad']} ({user['rol']})")
        self.setMinimumSize(1280, 800)
        self.showMaximized()
        self.setStyleSheet(f"QMainWindow {{ background-color: {C['bg_main']}; }}")
        self._build()
        self._nav(0)

    def _build(self):
        root = QWidget(); self.setCentralWidget(root)
        hl = QHBoxLayout(root); hl.setContentsMargins(0,0,0,0); hl.setSpacing(0)
        hl.addWidget(self._build_sidebar())
        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background: {C['bg_main']};")
        # Sayfaları ekle
        self.pg_dashboard  = GelismisDashboardPage(self.db)
        self.pg_sporcular  = SporcularPage(self.db)
        self.pg_antrenman  = AntrenmanlarPage(self.db)
        self.pg_takipler   = TakiplerPage(self.db)
        self.pg_programlar = ProgramlarPage(self.db)
        self.pg_vucut     = VucutOlcumleriPage(self.db)
        self.pg_hedefler  = HedeflerPage(self.db)
        self.pg_streak    = StreakPage(self.db)
        self.pg_beslenme  = BeslenmePage(self.db)
        self.pg_besinler  = BesinlerPage(self.db)
        self.pg_ilerleme  = IlerlemeGrafikleriPage(self.db)
        self.pg_haftalik  = HaftalikOzetPage(self.db)
        self.pg_rozetler  = RozetlerPage(self.db)
        self.pg_liderlik  = LeaderboardPage(self.db)
        self.pg_bir_rm    = BiRMPage(self.db)
        self.pg_radar     = RadarAnalizPage(self.db)
        self.pg_gelismis  = GelismisAnalizPage(self.db)
        self.pg_sablonlar  = SablonlarPage(self.db)
        self.pg_ayarlar    = SistemAyarlariPage(self.db)
        self.pg_excel      = ExcelExportPage(self.db)
        for pg in [self.pg_dashboard, self.pg_sporcular, self.pg_antrenman,
                   self.pg_takipler, self.pg_programlar,
                   self.pg_vucut, self.pg_hedefler, self.pg_streak,
                   self.pg_beslenme, self.pg_besinler,
                   self.pg_ilerleme, self.pg_haftalik,
                   self.pg_rozetler, self.pg_liderlik,
                   self.pg_bir_rm, self.pg_radar,
                   self.pg_gelismis,
                   self.pg_sablonlar, self.pg_ayarlar,
                   self.pg_excel]:
            self.stack.addWidget(pg)
        hl.addWidget(self.stack)

    def _build_sidebar(self):
        sb = QFrame(); sb.setFixedWidth(SIDEBAR_W)
        sb.setStyleSheet(f"""
            QFrame {{
                background-color: {C['sidebar_bg']};
                border-right: 1px solid {C['border']};
            }}
        """)
        lay = QVBoxLayout(sb); lay.setContentsMargins(0,0,0,0); lay.setSpacing(0)

        # Logo
        logo = QFrame(); logo.setFixedHeight(56)
        logo.setStyleSheet(f"background: {C['sidebar_bg']}; border-bottom: 1px solid {C['border']};")
        ll = QHBoxLayout(logo); ll.setContentsMargins(16,0,16,0)
        ico = QLabel('🏋️'); ico.setStyleSheet(f"font-size: 24px; border: none;")
        lbl = QLabel('FITNESS'); lbl.setFont(QFont('Segoe UI', 14, QFont.Bold))
        lbl.setStyleSheet(f"color: {C['text_main']}; border: none;")
        ll.addWidget(ico); ll.addWidget(lbl); ll.addStretch()
        lay.addWidget(logo)

        # Nav butonları
        self._nav_btns = []
        nav_items = [
            ('🏠', 'Dashboard',    0),
            ('🏃', 'Sporcular',    1),
            ('💪', 'Antrenmanlar', 2),
            ('📋', 'Takipler',      3),
            ('📅', 'Programlar',    4),
            ('📏', 'Vücut Ölçüm',  5),
            ('🎯', 'Hedefler',     6),
            ('🔥', 'Streak',       7),
            ('🥗', 'Beslenme',     8),
            ('🥦', 'Besinler',     9),
            ('📈', 'İlerleme',    10),
            ('📋', 'Haftalık Özet',11),
            ('🏅', 'Rozetler',    12),
            ('🏆', 'Liderlik',    13),
            ('🏋️', '1RM Hesap',   14),
            ('🕸️', 'Radar Analiz',15),
            ('🔬', 'Gelişmiş Analiz',16),
            ('📄', 'Şablonlar',   17),
            ('⚙️', 'Sistem',      18),
            ('📊', 'Excel/Rapor',  19),
        ]
        nav_w = QWidget(); nav_w.setStyleSheet("background: transparent;")
        nwl = QVBoxLayout(nav_w); nwl.setContentsMargins(8,8,8,8); nwl.setSpacing(3)
        for icon, label, idx in nav_items:
            btn = QPushButton(f'{icon}  {label}')
            btn.setFont(QFont('Segoe UI', 12)); btn.setFixedHeight(40)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda _, i=idx: self._nav(i))
            btn.setStyleSheet(self._sb_style(False))
            nwl.addWidget(btn); self._nav_btns.append(btn)
        nwl.addStretch(); lay.addWidget(nav_w, 1)

        # Kullanıcı bilgisi
        uf = QFrame(); uf.setFixedHeight(68)
        uf.setStyleSheet(f"background: {C['sidebar_bg']}; border-top: 1px solid {C['border']};")
        ufl = QHBoxLayout(uf); ufl.setContentsMargins(14,8,14,8)
        av = QLabel(self.user['ad'][0].upper())
        av.setFixedSize(34,34); av.setAlignment(Qt.AlignCenter)
        av.setFont(QFont('Segoe UI', 13, QFont.Bold))
        av.setStyleSheet(f"background:{C['primary']};color:white;border-radius:17px;border:none;")
        ufl.addWidget(av)
        info = QVBoxLayout(); info.setSpacing(1)
        nm = QLabel(self.user['kullanici_adi'].upper()); nm.setFont(QFont('Segoe UI', 11, QFont.Bold))
        nm.setStyleSheet(f"color:{C['text_main']};border:none;")
        rl = QLabel(self.user['rol'].upper()); rl.setFont(QFont('Segoe UI', 10))
        rl.setStyleSheet(f"color:{C['primary_light']};border:none;")
        info.addWidget(nm); info.addWidget(rl)
        ufl.addLayout(info); ufl.addStretch()
        btn_cikis = QPushButton('Çıkış'); btn_cikis.setFixedSize(72, 28)
        btn_cikis.setFont(QFont('Segoe UI', 10))
        btn_cikis.setStyleSheet(f"""
            QPushButton {{background:{C['danger']};color:white;border:none;
                border-radius:6px;font-size:10px;min-height:0;}}
            QPushButton:hover {{background:{C['danger_dark']};}}
        """)
        btn_cikis.clicked.connect(self._cikis); ufl.addWidget(btn_cikis)
        lay.addWidget(uf); return sb

    def _sb_style(self, active):
        if active:
            return f"""QPushButton {{
                background:{C['primary']}; color:white; border:none;
                border-radius:10px; text-align:left; padding:0 18px;
                font-size:13px; font-weight:bold; height:48px;
                margin:0 0px;
            }}"""
        return f"""QPushButton {{
            background:transparent; color:{C['text_secondary']}; border:none;
            border-radius:10px; text-align:left; padding:0 18px;
            font-size:13px; height:48px;
        }}
        QPushButton:hover {{
            background:{C['bg_hover']}; color:{C['text_main']};
        }}"""

    def _nav(self, idx):
        self.stack.setCurrentIndex(idx)
        for i, btn in enumerate(self._nav_btns):
            btn.setStyleSheet(self._sb_style(i == idx))
        page = self.stack.currentWidget()
        if hasattr(page, 'refresh'): page.refresh()

    def _cikis(self):
        if dark_confirm(self, 'Çıkış', 'Oturumu kapatmak istiyor musunuz?'):
            self.close(); start_app(self.db)


# ═══════════════════════════════════════════════════════════════════
# BAŞLATMA
# ═══════════════════════════════════════════════════════════════════
def start_app(db=None):
    if db is None: db = DatabaseManager()
    def on_login(user):
        login_win.close()
        main = MainWindow(db, user); main.show()
        app = QApplication.instance()
        if app: app._main = main; app.setQuitOnLastWindowClosed(True)
    login_win = LoginWindow(db, on_login)
    login_win.show()
    app = QApplication.instance()
    if app: app._login = login_win; app.setQuitOnLastWindowClosed(True)

def main():
    app = QApplication(sys.argv); app.setStyle('Fusion')
    app.setApplicationName('Fitness Takip Sistemi v1.0')
    from PyQt5.QtGui import QPalette, QColor
    palette = QPalette()
    palette.setColor(QPalette.Window,     QColor(C['bg_main']))
    palette.setColor(QPalette.Base,       QColor(C['bg_secondary']))
    palette.setColor(QPalette.WindowText, QColor(C['text_main']))
    palette.setColor(QPalette.Text,       QColor(C['text_main']))
    palette.setColor(QPalette.Button,     QColor(C['bg_card']))
    palette.setColor(QPalette.ButtonText, QColor(C['text_main']))
    palette.setColor(QPalette.Highlight,  QColor(C['primary']))
    palette.setColor(QPalette.HighlightedText, QColor('#ffffff'))
    app.setPalette(palette)
    app.setStyleSheet(GLOBAL_STYLESHEET)
    start_app(); sys.exit(app.exec_())

if __name__ == '__main__':
    main()

# ═══════════════════════════════════════════════════════════════════
# TIER 2 — Antrenman Günlüğü + Takip Sistemi
# ═══════════════════════════════════════════════════════════════════